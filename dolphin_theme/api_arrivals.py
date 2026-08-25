import json

import frappe
from frappe.utils import flt, cint, now_datetime
@frappe.whitelist()
def get_my_roles(): return frappe.get_roles()

# Tolerance aligned with the Measurement Variations work: a side passes if it is
# within 3 cm OR 3% of the dispatched value (whichever is larger).
DIM_TOL_ABS = 3.0
DIM_TOL_PCT = 0.03


def _within_tol(dispatched, arrived):
    d, a = flt(dispatched), flt(arrived)
    if not d:
        return True
    tol = max(DIM_TOL_ABS, d * DIM_TOL_PCT)
    return abs(d - a) <= tol


def _weights(b):
    """Per-block (CBM, M.Tonnes, Kgs) per the shipping-document convention:
    M.T = CBM x specific gravity; Kgs = M.T x 1000."""
    cbm = flt(b.cbm) or round(flt(b.length) * flt(b.width) * flt(b.height) / 1e6, 3)
    factor = flt(b.get("avg_factor")) or 2.6
    mt = round(cbm * factor, 3)
    return round(cbm, 3), mt, cint(round(mt * 1000))


def _edit_distance(a, b):
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[0], i
        for j in range(1, n + 1):
            cur = dp[j]
            dp[j] = min(dp[j] + 1, dp[j - 1] + 1, prev + (a[i - 1] != b[j - 1]))
            prev = cur
    return dp[n]


def _nearest(block_no, candidates):
    target, best, best_d = str(block_no).strip(), None, 99
    for c in candidates:
        d = _edit_distance(target, str(c))
        if d < best_d:
            best, best_d = c, d
    return best if best_d <= 2 else None


def _dispatched_index():
    rows = frappe.db.sql(
        """
        SELECT r.block, r.block_no, r.export_block_no, r.length_gross AS l,
               r.width_gross AS w, r.height_gross AS h, r.gross_volume AS vol, p.name AS dc
        FROM `tabDC Block Row` r
        JOIN `tabDelivery Challan` p ON p.name = r.parent
        WHERE p.docstatus = 1
        """,
        as_dict=True,
    )
    return {str(r.block).strip(): r for r in rows}


def _arrived_on_other(block_no, exclude_arrival):
    return bool(
        frappe.db.sql(
            """
            SELECT 1 FROM `tabPort Arrival Block` b
            JOIN `tabPort Arrival` p ON p.name = b.parent
            WHERE TRIM(b.block_no) = %s AND b.parent != %s
            LIMIT 1
            """,
            (str(block_no).strip(), exclude_arrival),
        )
    )


def _classify(pa):
    """Classify every block row in place. Does not save."""
    idx = _dispatched_index()
    keys = list(idx.keys())
    seen, flags = set(), 0
    for row in pa.blocks:
        bno = str(row.block_no or "").strip()
        row.suggested_block = None
        row.matched_dc = None
        if row.resolution_type:
            row.recon_status = "Resolved"
        elif not bno:
            row.recon_status = "Typo - not in DC"
            flags += 1
        elif bno in seen or _arrived_on_other(bno, pa.name):
            row.recon_status = "Duplicate"
            flags += 1
        elif bno not in idx:
            row.recon_status = "Typo - not in DC"
            row.suggested_block = _nearest(bno, keys)
            flags += 1
        else:
            dc = idx[bno]
            row.matched_dc = dc.dc
            ok = (
                _within_tol(dc.l, row.length)
                and _within_tol(dc.w, row.width)
                and _within_tol(dc.h, row.height)
            )
            row.recon_status = "Matched" if ok else "Dimension mismatch"
            if not ok:
                flags += 1
        seen.add(bno)
    awaiting = [k for k in keys if k not in seen]
    return {"flags": flags, "awaiting": awaiting, "total": len(pa.blocks)}


def run_reconcile(doc, method=None):
    """doc_events hook (validate) — mutate in place, framework saves. Phase 2 wires this."""
    try:
        _classify(doc)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Dolphin reconcile")


@frappe.whitelist()
def reconcile_arrival(name):
    """Run the cross-check on a Port Arrival and persist the per-row verdicts."""
    pa = frappe.get_doc("Port Arrival", name)
    res = _classify(pa)
    for row in pa.blocks:
        frappe.db.set_value("Port Arrival Block", row.name, {
            "recon_status": row.recon_status,
            "matched_dc": row.matched_dc,
            "suggested_block": row.suggested_block,
        }, update_modified=False)
    _mark_at_port(pa)
    return res


def _mark_at_port(pa):
    """A block confirmed arrived (Matched, or a non-duplicate resolution) is physically
    AT THE PORT -> flip its Quarry Block status to 'At Port'.

    REWRITTEN 17 Aug 2026. The previous version called
    ``frappe.db.exists("Quarry Block", bno)`` — it resolved a number typed by a
    shipping agency against the Quarry Block RECORD ID. Because the record id is
    an autoincrementing integer, every plausible number matched something, and
    56 blocks were moved to a port they had never reached; 25 of them were
    provably the wrong block.

    Now: the number is resolved against export number then quarry number only,
    a number that matches two blocks writes NOTHING, and every write is a
    versioned save carrying the arrival it came from as its reason.

    Returns a report so the caller can show what was refused instead of the
    refusal being invisible."""
    from dolphin_theme.block_resolve import try_resolve, set_status

    moved, skipped = [], []
    draft = (getattr(pa, "docstatus", 0) or 0) == 0
    for row in pa.blocks:
        bno = str(row.block_no or "").strip()
        arrived = (row.recon_status == "Matched") or (
            row.resolution_type and row.resolution_type != "Removed (duplicate)"
        )
        if not (arrived and bno):
            continue
        if draft:
            # B36's root cause, in its other guise: a DRAFT arrival must not move
            # live stock. The evidence is kept and shown, but nothing is written.
            skipped.append({"block": bno, "why": "draft-arrival"})
            continue
        hit, why = try_resolve(bno, allow_record_name=False)
        if not hit:
            skipped.append({"block": bno, "why": why})
            continue
        cur = _s(hit.get("status"))
        if cur in ("At Port", "Shipped", "Sold"):
            continue
        res = set_status(hit["name"], "At Port",
                         "arrived on {0} (row {1})".format(pa.name, row.idx),
                         machine="server (arrival reconcile)")
        (moved if res.get("ok") else skipped).append(
            {"block": bno, "name": hit["name"], "why": res.get("error")})
    return {"moved": len(moved), "skipped": skipped}


@frappe.whitelist()
def resolve_flag(arrival, row_name, resolution_type, updates=None, note=None, machine=None):
    """Clear one flag. `updates` (JSON fieldname->value) optionally corrects data.
    Accept-as-is requires a reason. Every resolution stamps who + which PC + when."""
    if resolution_type == "Accepted as-is" and not (note or "").strip():
        frappe.throw("A reason is required to accept a discrepancy as-is.")
    pa = frappe.get_doc("Port Arrival", arrival)
    row = next((r for r in pa.blocks if r.name == row_name), None)
    if not row:
        frappe.throw("Row not found.")
    if updates:
        data = json.loads(updates) if isinstance(updates, str) else updates
        for k, v in data.items():
            row.set(k, v)
    row.resolution_type = resolution_type
    row.resolution_note = note
    row.resolved_by = frappe.session.user
    row.resolved_machine = machine or "un-named device"
    row.resolved_on = now_datetime()
    row.recon_status = "Resolved"
    pa.save(ignore_permissions=True)
    _mark_at_port(pa)
    return {"ok": True}


@frappe.whitelist()
def get_token():
    """Return the current session's CSRF token via a GET call (no CSRF needed),
    so the www page can authorise its POSTs even when served from page cache."""
    return frappe.sessions.get_csrf_token()


@frappe.whitelist()
def count_open_flags():
    """Count unresolved reconciliation flags across all arrivals (desk banner)."""
    r = frappe.db.sql(
        """
        SELECT COUNT(*) FROM `tabPort Arrival Block`
        WHERE recon_status IN ('Typo - not in DC', 'Dimension mismatch', 'Duplicate')
          AND IFNULL(resolution_type, '') = ''
        """
    )
    return int(r[0][0]) if r and r[0] else 0


@frappe.whitelist()
def full_view():
    """One flat, colour-codeable sheet: every block across every arrival with its
    measurement, CBM/MT/Kgs, matched challan, live reconciliation status and resolver,
    plus the reverse-check list of blocks dispatched but not yet arrived."""
    idx = _dispatched_index()
    emap = _export_map()
    awaiting_keys = set(idx.keys())
    rows = []
    arrivals = frappe.get_all(
        "Port Arrival",
        fields=["name", "mark", "shipper", "arrival_date"],
        order_by="arrival_date desc",
    )
    for a in arrivals:
        pa = frappe.get_doc("Port Arrival", a.name)
        _classify(pa)
        for b in pa.blocks:
            bno = str(b.block_no or "").strip()
            awaiting_keys.discard(bno)
            cbm, mt, kgs = _weights(b)
            status = "Resolved" if b.resolution_type else (b.recon_status or "")
            dcrow = idx.get(bno)
            rows.append({
                "arrival": a.name,
                "mark": a.mark or "",
                "block_no": b.block_no or "",
                "quarry_block": (dcrow.block if dcrow else ""),
                "export_block_no": (_s(dcrow.export_block_no) if (dcrow and dcrow.get("export_block_no")) else "") or emap.get(_s(b.block_no)) or emap.get(_s(dcrow.block) if dcrow else "") or "",
                "length": cint(b.length),
                "width": cint(b.width),
                "height": cint(b.height),
                "cbm": cbm,
                "mt": mt,
                "kgs": kgs,
                "dc_l": (cint(dcrow.l) if dcrow else 0),
                "dc_w": (cint(dcrow.w) if dcrow else 0),
                "dc_h": (cint(dcrow.h) if dcrow else 0),
                "dc_cbm": (round(flt(dcrow.vol), 3) if dcrow else 0),
                "matched_dc": b.matched_dc or "",
                "status": status,
                "raw_status": b.recon_status or "",
                "suggested_block": b.suggested_block or "",
                "resolution_type": b.resolution_type or "",
                "resolved_by": b.resolved_by or "",
                "resolved_machine": b.resolved_machine or "",
            })
    awaiting = [{"block_no": k, "matched_dc": idx[k].dc} for k in sorted(awaiting_keys)]
    return {"rows": rows, "awaiting": awaiting}


@frappe.whitelist()
def generate_shipping_from_arrival(arrival):
    """Create a draft Shipping Document from a VERIFIED Port Arrival.
    Gated: refuses while any reconciliation flag is still open. Excludes
    blocks removed as duplicates. User completes consignee + rate after."""
    pa = frappe.get_doc("Port Arrival", arrival)
    _classify(pa)
    flagged = [
        b for b in pa.blocks
        if b.recon_status in ("Typo - not in DC", "Dimension mismatch", "Duplicate")
        and not b.resolution_type
    ]
    if flagged:
        frappe.throw(
            "Resolve all {0} open flag(s) before generating the shipping document.".format(len(flagged))
        )

    sd = frappe.new_doc("Shipping Document")
    sd.shipment_date = frappe.utils.today()
    if sd.meta.has_field("source_arrival"):
        sd.source_arrival = pa.name
    sd.shipping_mark = pa.get("mark")
    sd.marks_nos = pa.get("mark")
    sd.voyage_no = pa.get("vessel")           # Port Arrival.vessel is free text
    sd.bl_no = pa.get("booking_no")
    sd.goods_description = "Granite - Roughly Trimmed Blocks"
    sd.currency = "USD"
    sd.tax_treatment = "Export under LUT (No GST)"
    sd.rate_basis = "Per Kg"
    sd.country_of_origin = "INDIA"
    sd.pre_carriage_by = "ROAD"
    sd.terms_of_delivery = "F.O.B."

    total_cbm = 0.0
    total_mt = 0.0
    for b in pa.blocks:
        if b.resolution_type == "Removed (duplicate)":
            continue
        cbm, mt, kgs = _weights(b)
        vol = cbm
        row = sd.append("blocks", {})
        row.block = b.quarry_block
        row.block_no = b.block_no
        row.length = cint(b.length)
        row.width = cint(b.width)
        row.height = cint(b.height)
        row.net_volume = vol
        row.net_tonnage = mt
        row.net_kgs = kgs
        total_cbm += vol
        total_mt += mt

    sd.block_count = len(sd.blocks)
    sd.total_cbm = round(total_cbm, 2)
    sd.total_net_tonnage = round(total_mt, 3)
    sd.total_net_kgs = cint(round(total_mt * 1000))
    sd.flags.ignore_mandatory = True
    sd.insert(ignore_permissions=True)
    return sd.name


@frappe.whitelist()
def parse_check(arrival):
    """Did the consolidated xls parse cleanly? Compares row count + summed totals
    (file vs imported). file_rows/file_cbm/file_net are stored on the parent at parse."""
    pa = frappe.get_doc("Port Arrival", arrival)
    imported = len(pa.blocks)
    cbm = round(sum(flt(b.cbm) for b in pa.blocks), 2)
    net = round(sum(flt(b.net_wt) for b in pa.blocks), 2)
    file_rows = cint(pa.get("file_rows"))
    return {
        "imported_rows": imported,
        "imported_cbm": cbm,
        "imported_net": net,
        "file_rows": file_rows,
        "matches": file_rows in (0, imported),
    }


@frappe.whitelist()
def upsert_arrival(rows, source_file=None, current=None, meta=None):
    """Dedupe arrivals at parse time (P1, the real point-2 fix).

    `rows` is the JSON block list the form parsed from the xls (same fieldnames the
    Blocks child table uses). If these block numbers already live on ANOTHER Port
    Arrival, fold them into that record instead of letting a duplicate arrival be
    created -- upsert = update-or-insert, so the same arrival coming in again updates
    the one record rather than spawning PORT-ARR-0004, 0005, ... with the same blocks.

    Returns one of:
      {"action": "none"}                         -> no overlap; caller fills `current`
                                                    record as before (today's flow).
      {"action": "updated", "name": <arrival>,   -> exactly one existing arrival held
       "overlap": n, "total": m}                    these blocks; it was updated in
                                                    place. Caller routes to it and
                                                    drops the just-created draft.
      {"action": "ambiguous", "parents": [...],  -> blocks span >1 existing arrival;
       "overlap": {parent: n, ...}}                  no automatic merge -- caller warns
                                                    and lets the user resolve.
      {"action": "partial", "name": <arrival>,   -> the parse shares blocks with one
       "would_drop": [...]}                          arrival but is missing some of its
                                                    blocks; replacing would lose data,
                                                    so caller warns instead of writing.
    """
    data = json.loads(rows) if isinstance(rows, str) else (rows or [])
    nums = [str(r.get("block_no") or "").strip() for r in data]
    nums = [n for n in nums if n]
    if not nums:
        return {"action": "none"}

    cur = current or ""
    if cur.startswith("new-"):
        cur = ""

    # which OTHER Port Arrivals already hold any of these block numbers?
    existing = frappe.get_all(
        "Port Arrival Block",
        filters={"block_no": ["in", nums], "parenttype": "Port Arrival"},
        fields=["parent", "block_no"],
    )
    overlap = {}
    for e in existing:
        if e.parent == cur:
            continue
        overlap.setdefault(e.parent, set()).add(str(e.block_no).strip())

    if not overlap:
        return {"action": "none"}

    if len(overlap) > 1:
        return {
            "action": "ambiguous",
            "parents": list(overlap.keys()),
            "overlap": {p: len(s) for p, s in overlap.items()},
        }

    # exactly one existing arrival -> candidate for UPSERT
    target = next(iter(overlap))
    pa = frappe.get_doc("Port Arrival", target)

    # Safety: only rebuild the target's blocks when this parse would NOT drop any
    # block already on it -- i.e. the parse is the same arrival or a superset. If the
    # incoming file is missing blocks the target already has, replacing would lose
    # data (a different shipment that merely shares a block number), so warn instead.
    target_blocks = {
        str(b.block_no or "").strip() for b in pa.blocks if str(b.block_no or "").strip()
    }
    parsed_set = set(nums)
    if not target_blocks.issubset(parsed_set):
        return {
            "action": "partial",
            "name": target,
            "would_drop": sorted(target_blocks - parsed_set),
            "overlap": len(overlap[target]),
        }

    # safe to upsert: rebuild its blocks from this parse
    child_fields = {df.fieldname for df in frappe.get_meta("Port Arrival Block").fields}
    pa.set("blocks", [])
    for r in data:
        if not str(r.get("block_no") or "").strip():
            continue
        child = pa.append("blocks", {})
        for k, v in r.items():
            if k in child_fields:
                child.set(k, v)

    # refresh parent header from this parse where supplied
    m = json.loads(meta) if isinstance(meta, str) else (meta or {})
    for f in ("shipper", "mark", "arrival_date", "port", "vessel",
              "booking_no", "email_subject"):
        if m.get(f) not in (None, ""):
            pa.set(f, m.get(f))
    if source_file:
        pa.set("source_file", source_file)

    # keep the parent summary fields in sync (normally client-computed on the form)
    if pa.meta.has_field("total_blocks"):
        pa.total_blocks = len(pa.blocks)
    if pa.meta.has_field("total_cbm"):
        pa.total_cbm = round(sum(flt(b.cbm) for b in pa.blocks), 2)
    if pa.meta.has_field("total_net_wt"):
        pa.total_net_wt = round(sum(flt(b.net_wt) for b in pa.blocks), 2)

    # recompute the per-row reconciliation verdicts on the merged set
    try:
        _classify(pa)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Dolphin upsert reclassify")

    pa.flags.ignore_mandatory = True
    pa.save(ignore_permissions=True)
    return {
        "action": "updated",
        "name": target,
        "overlap": len(overlap[target]),
        "total": len(nums),
    }


@frappe.whitelist()
def create_shipment_lot(consignee=None, rows=None, mark=None, vessel=None):
    """Build a Export Shipment Lot (the final lot) from at-port blocks selected on the
    reconciliation sheet. `rows` is the JSON list of selected full_view rows."""
    data = json.loads(rows) if isinstance(rows, str) else (rows or [])
    if not data:
        frappe.throw("Select at least one at-port block to build a Export Shipment Lot.")
    lot = frappe.new_doc("Export Shipment Lot")
    lot.shipment_date = frappe.utils.today()
    if consignee:
        lot.export_consignee = consignee
    if mark:
        lot.shipping_mark = mark
    if vessel:
        lot.vessel = vessel
    lot.status = "Ready"
    arrivals, tc, tt = set(), 0.0, 0.0
    for r in data:
        _qb = _qb_by_any(r.get("quarry_block") or r.get("block_no"))
        _l = cint(r.get("length")); _w = cint(r.get("width")); _h = cint(r.get("height"))
        _cbm = flt(r.get("cbm")) or (round(_l * _w * _h / 1e6, 3) if (_l and _w and _h) else 0)
        _mt = flt(r.get("net_tonnage") or r.get("mt") or r.get("net_ton") or 0)
        if not _mt:
            _mt = round(_cbm * (flt(r.get("tonnage_factor")) or 2.7), 3)
        ch = lot.append("blocks", {})
        ch.block = r.get("quarry_block") or (_qb.get("name") if _qb else None)
        ch.block_no = r.get("block_no") or (_qb.get("block_number") if _qb else "")
        ch.length = _l
        ch.width = _w
        ch.height = _h
        ch.cbm = _cbm
        ch.net_tonnage = _mt
        ch.net_kgs = cint(r.get("kgs") or round(_mt * 1000))
        ch.grade = r.get("grade") or ""
        ch.source_dc = r.get("matched_dc") or r.get("source_dc") or ""
        ch.source_arrival = r.get("arrival") or ""
        if ch.meta.has_field("export_block_no"):
            ch.export_block_no = _s(r.get("export_block_no")) or (_s(_qb.get("export_block_no")) if _qb else "")
        if r.get("arrival"):
            arrivals.add(r.get("arrival"))
        tc += _cbm
        tt += _mt
    lot.block_count = len(lot.blocks)
    lot.total_cbm = round(tc, 2)
    lot.total_net_tonnage = round(tt, 3)
    lot.total_net_kgs = cint(round(tt * 1000))
    lot.source_arrivals = ", ".join(sorted(arrivals))
    lot.flags.ignore_mandatory = True
    lot.insert(ignore_permissions=True)
    return lot.name

# ===========================================================================
# Arrival .xls importer  (append to dolphin_theme/api_arrivals.py)
# Block number is the PRIMARY key. Only Dolphin's own sheet is ingested, so a
# workbook that also carries another company's tab (e.g. VARDHINI XG) cannot
# pollute the report. Idempotent: re-importing upserts blocks by block_no.
# ===========================================================================
import re as _re


def _xls_s(v):
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def _xls_num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _xls_is_dolphin_sheet(sheet, single):
    """Keep the Dolphin sheet: the only sheet, or a tab whose name / top title
    rows mention 'dolphin'. Drop tabs whose title names another firm (M/S. ...)."""
    if single:
        return True
    if "dolphin" in sheet.name.lower():
        return True
    for r in range(min(sheet.nrows, 3)):
        line = " ".join(_xls_s(sheet.cell_value(r, c)).lower() for c in range(sheet.ncols))
        if "dolphin" in line:
            return True
        if _re.search(r"m/?s\.", line):
            return False
    return False


def _xls_header(sheet):
    for r in range(min(sheet.nrows, 12)):
        cells = [_xls_s(sheet.cell_value(r, c)).lower() for c in range(sheet.ncols)]
        if _re.search(r"block\s*no", " ".join(cells)):
            cm = {}
            for c, t in enumerate(cells):
                t = t.strip()
                if _re.fullmatch(r"block\s*no\.?", t):
                    cm["block_no"] = c
                elif t == "cbm":
                    cm.setdefault("cbm", c)
                elif t in ("mark", "marking"):
                    cm["mark"] = c
                elif t in ("vehicle no.", "vehicle no", "way o transport", "way of transport"):
                    cm["vehicle"] = c
                elif t == "location":
                    cm["location"] = c
                elif t in ("line no", "line no.", "line no. "):
                    cm["line"] = c
                elif t in ("ado no", "ado no."):
                    cm["ado"] = c
                elif t in ("permit no", "permit no."):
                    cm["permit"] = c
                elif "weight" in t or t == "a/wt":
                    cm.setdefault("weight", c)
                elif t == "measurement":
                    cm["meas"] = c
            return r, cm
    return None, {}


def _xls_dims(sheet, r, start):
    vals, c = [], start
    while c < sheet.ncols and len(vals) < 3:
        v = _xls_num(sheet.cell_value(r, c))
        if v is not None:
            vals.append(v)
        c += 1
    return (vals + [None, None, None])[:3]


def _parse_arrival_xls(content):
    """content: bytes of a .xls. Returns (rows, sheet_name) for the Dolphin sheet."""
    try:
        import xlrd
    except Exception:
        frappe.throw(
            "The .xls reader (xlrd) is not installed on this bench. "
            "Add xlrd to the app's requirements and redeploy."
        )
    wb = xlrd.open_workbook(file_contents=content)
    single = wb.nsheets == 1
    rows, used_sheet = [], None
    for sh in wb.sheets():
        if not _xls_is_dolphin_sheet(sh, single):
            continue
        hr, cm = _xls_header(sh)
        if hr is None or "block_no" not in cm:
            continue
        used_sheet = sh.name
        for r in range(hr + 1, sh.nrows):
            bno = _xls_s(sh.cell_value(r, cm["block_no"]))
            if not bno or not _re.search(r"\d", bno):
                continue
            joined = " ".join(_xls_s(sh.cell_value(r, c)).lower() for c in range(sh.ncols))
            if "total" in joined and not _re.search(r"\bblock", joined):
                continue
            length = width = height = None
            if "meas" in cm:
                length, width, height = _xls_dims(sh, r, cm["meas"])

            def cell(k):
                return _xls_s(sh.cell_value(r, cm[k])) if k in cm else None

            rows.append({
                "block_no": bno,
                "mark": (cell("mark") or None),
                "cbm": _xls_num(sh.cell_value(r, cm["cbm"])) if "cbm" in cm else None,
                "weight": _xls_num(sh.cell_value(r, cm["weight"])) if "weight" in cm else None,
                "length": length, "width": width, "height": height,
                "vehicle_no": cell("vehicle"),
                "yard_location": cell("location"),
                "line_no": cell("line"),
                "ado_no": cell("ado"),
                "permit_no": cell("permit"),
            })
    return rows, used_sheet



# --------------------------------------------------------------------------
# THE MARK MUST NEVER ABORT AN IMPORT.  24 Aug 2026
#
# [stated] "the manual upload and parsing of xls is broken fix it"
#
# It was not the file picker and not the CSRF token. Posting a real agency
# spreadsheet returned 417:
#
#     LinkValidationError: Could not find Mark: YL/XMN
#
# `Port Arrival.mark` is a custom Link field pointing at Shipping Mark. The
# agency writes the marking as "YL/XMN" with a SLASH; the Shipping Mark record
# is "YL-XMN" with a HYPHEN. One character, and the whole 200-row import is
# rejected - every block lost because a descriptive label did not match.
#
# Two rules now:
#   1. resolve leniently - slash, hyphen and space are the same separator, and
#      case does not matter. "YL/XMN" finds "YL-XMN".
#   2. if it still does not resolve, LEAVE THE LINK EMPTY and carry on. The
#      marking is descriptive; the blocks are the payload. Losing 200 blocks to
#      protect a label is the wrong trade in every direction.
# --------------------------------------------------------------------------
def _mark_key(v):
    """A marking, reduced to what actually identifies it."""
    return "".join(ch for ch in _s(v).upper() if ch.isalnum())


def _resolve_mark(raw):
    """The Shipping Mark record this marking means, or None. Never raises."""
    want = _mark_key(raw)
    if not want:
        return None
    try:
        for name in frappe.get_all("Shipping Mark", pluck="name",
                                   limit_page_length=0):
            if _mark_key(name) == want:
                return name
    except Exception:
        return None
    return None

@frappe.whitelist()
def import_xls(arrival=None, mark=None, agency=None):
    """Import a Dolphin arrivals .xls (uploaded as multipart 'file').

    Block number is the primary key; only Dolphin's sheet is ingested.
    Idempotent - upserts blocks by block_no into the target Port Arrival:
    the one named in `arrival`, else an existing arrival with the same
    mark + source sheet, else a new Port Arrival. Runs reconciliation after."""
    f = frappe.request.files.get("file") if frappe.request else None
    if not f:
        frappe.throw("No file received. Attach an .xls file.")
    content = f.stream.read() if hasattr(f, "stream") else f.read()
    fname = getattr(f, "filename", "arrival.xls")

    rows, sheet = _parse_arrival_xls(content)
    if not rows:
        frappe.throw("No Dolphin block rows found (only the Dolphin sheet is read).")

    marks = [r["mark"] for r in rows if r.get("mark")]
    doc_mark = mark or (marks[0] if marks else None)

    pa = None
    if arrival and frappe.db.exists("Port Arrival", arrival):
        pa = frappe.get_doc("Port Arrival", arrival)
    else:
        existing = frappe.db.get_value(
            "Port Arrival", {"mark": doc_mark, "source_sheet": sheet, "docstatus": 0}, "name"
        ) if doc_mark else None
        if existing:
            pa = frappe.get_doc("Port Arrival", existing)
        else:
            pa = frappe.new_doc("Port Arrival")
            pa.arrival_date = frappe.utils.today()

    # The mark is a Link. An unknown value used to abort the whole import, so it
    # is resolved leniently and simply left blank when it cannot be matched.
    mark_note = None
    if doc_mark and pa.meta.has_field("mark"):
        resolved = _resolve_mark(doc_mark)
        if resolved:
            pa.mark = resolved
        else:
            mark_note = _s(doc_mark)
    if agency and pa.meta.has_field("shipper"):
        pa.shipper = agency
    if pa.meta.has_field("source_sheet"):
        pa.source_sheet = sheet
    subj = frappe.form_dict.get("subject") if frappe.form_dict else None
    sender = frappe.form_dict.get("sender") if frappe.form_dict else None
    if subj and pa.meta.has_field("email_subject"):
        pa.email_subject = subj
    if sender and pa.meta.has_field("email_sender"):
        pa.email_sender = sender

    existing_by_block = {str(b.block_no).strip(): b for b in pa.blocks}
    created = updated = 0
    for r in rows:
        b = existing_by_block.get(r["block_no"])
        if not b:
            b = pa.append("blocks", {})
            b.block_no = r["block_no"]
            created += 1
        else:
            updated += 1
        if r.get("mark"):
            b.mark = r["mark"]
        for k in ("length", "width", "height", "cbm",
                  "vehicle_no", "yard_location", "line_no", "ado_no", "permit_no"):
            if r.get(k) is not None and b.meta.has_field(k):
                b.set(k, r[k])
        if r.get("weight") is not None:
            if b.meta.has_field("net_wt"):
                b.net_wt = r["weight"]
            if b.meta.has_field("a_wt") and not b.get("a_wt"):
                b.a_wt = r["weight"]

    pa.total_blocks = len(pa.blocks)
    pa.total_cbm = round(sum(flt(b.cbm) for b in pa.blocks), 3)
    pa.total_net_wt = round(sum(flt(b.net_wt) for b in pa.blocks), 3)
    pa.flags.ignore_mandatory = True
    pa.save(ignore_permissions=True)

    try:
        _classify(pa)
        for row in pa.blocks:
            frappe.db.set_value("Port Arrival Block", row.name, {
                "recon_status": row.recon_status,
                "matched_dc": row.matched_dc,
                "suggested_block": row.suggested_block,
            }, update_modified=False)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "import_xls reconcile")

    frappe.db.commit()
    out = {
        "arrival": pa.name,
        "sheet": sheet,
        "file": fname,
        "created": created,
        "updated": updated,
        "duplicates": 0,
        "total_blocks": pa.total_blocks,
    }
    # Said plainly rather than swallowed: the blocks are in, and this one label
    # could not be matched to a Shipping Mark record.
    if mark_note:
        out["mark_not_matched"] = mark_note
        out["note"] = ("The blocks imported. The marking \"{0}\" does not match "
                       "any Shipping Mark record, so it was left blank on the "
                       "arrival - add it as a Shipping Mark if you want it "
                       "carried.").format(mark_note)
    return out


# ===========================================================================
# Workspace additions (append to dolphin_theme/api_arrivals.py)
#   lots_view()        -> read-only feed for the Shipment Lots tab + lets the
#                         Stock page compute in-lot / loaded / left-out.
#   move_to_at_port()  -> bulk "skip arrivals": take DC-submitted blocks that are
#                         awaiting arrival and place them At Port (draft, reversible).
# Both are defensive: field names are probed with meta.has_field so a schema
# difference degrades gracefully instead of writing bad data.
# ===========================================================================
import json as _json


@frappe.whitelist()
def lots_view():
    """Every Export Shipment Lot with its consignee/vessel/doc numbers and the
    list of block numbers it holds. Read-only."""
    out = []
    emap = _export_map()
    for name in frappe.get_all("Export Shipment Lot", pluck="name"):
        try:
            d = frappe.get_doc("Export Shipment Lot", name)
        except Exception:
            continue

        block_nos = []
        for tf in d.meta.get_table_fields():
            rows = d.get(tf.fieldname) or []
            picked = []
            for r in rows:
                bn = r.get("block_no") or r.get("block") or r.get("quarry_block")
                if bn:
                    _k = str(bn).strip()
                    picked.append(emap.get(_k, _k))
            if picked:
                block_nos = picked
                break

        def gv(*keys):
            for k in keys:
                if d.meta.has_field(k):
                    v = d.get(k)
                    if v not in (None, ""):
                        return v
            return None

        shipped = bool(gv("shipped")) or d.docstatus == 1
        stf = gv("status")
        if stf and str(stf).lower() in ("shipped", "dispatched", "sailed", "closed"):
            shipped = True

        out.append({
            "name": d.name,
            "title": gv("title", "lot_name", "lot_title") or d.name,
            "consignee": gv("consignee", "consignee_name", "customer", "shipper", "mark") or "",
            "vessel": gv("vessel", "vessel_name", "vessel_voyage", "voyage") or "",
            "packing_list": gv("packing_list", "packing_list_no", "pl_no"),
            "bl_no": gv("bl_no", "bill_of_lading", "bl_number", "bl"),
            "ship_date": (str(gv("ship_date", "shipment_date", "sailed_on") or "") or None),
            "total_blocks": gv("total_blocks") or len(block_nos),
            "total_cbm": gv("total_cbm"),
            "status": "ship" if shipped else "build",
            "shipped": shipped,
            "block_nos": block_nos,
        })
    return out


@frappe.whitelist()
def move_to_at_port(blocks=None, note=None):
    """Place DC-submitted 'awaiting arrival' blocks At Port without an arrival file.
    Writes into a single reusable DRAFT Port Arrival. Idempotent by block number."""
    if isinstance(blocks, str):
        blocks = _json.loads(blocks)
    blocks = blocks or []
    if not blocks:
        frappe.throw("No blocks supplied to move.")

    label = "AT-PORT (skipped arrivals)"
    meta = frappe.get_meta("Port Arrival")
    tag = next((f for f in ("source_sheet", "email_subject", "source_file")
                if meta.has_field(f)), None)

    pa = None
    if tag:
        tname = frappe.db.get_value("Port Arrival", {tag: label, "docstatus": 0}, "name")
        if tname:
            pa = frappe.get_doc("Port Arrival", tname)
    if pa is None:
        pa = frappe.new_doc("Port Arrival")
        if tag:
            pa.set(tag, label)
        if meta.has_field("arrival_date"):
            pa.arrival_date = frappe.utils.today()

    from dolphin_theme.block_resolve import try_resolve, set_status

    existing = {str(b.block_no).strip() for b in pa.blocks}
    moved, refused = 0, []
    for it in blocks:
        bn = str((it or {}).get("block_no") or "").strip()
        if not bn or bn in existing:
            continue
        # 17 Aug: resolve before writing. A number that means two blocks, or no
        # block, never reaches the arrival sheet — it is reported back instead.
        hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            refused.append({"block": bn, "why": why})
            continue
        row = pa.append("blocks", {})
        row.block_no = bn
        existing.add(bn)
        if row.meta.has_field("matched_dc") and it.get("dc"):
            row.matched_dc = it.get("dc")
        if row.meta.has_field("recon_status"):
            row.recon_status = "Resolved"
        if note not in (None, "") and row.meta.has_field("resolution_note"):
            row.resolution_note = note
            if row.meta.has_field("resolved_by"):
                row.resolved_by = frappe.session.user
            if row.meta.has_field("resolved_on"):
                row.resolved_on = now_datetime()
        moved += 1

    if meta.has_field("total_blocks"):
        pa.total_blocks = len(pa.blocks)
    pa.flags.ignore_mandatory = True
    pa.save(ignore_permissions=True)

    # A deliberate, recorded skip of the arrival step (B39). The status is written
    # here — not implied by the draft arrival — so the block's own history says
    # who decided to skip and why.
    skipped = []
    for it in blocks:
        bn = str((it or {}).get("block_no") or "").strip()
        hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            continue
        # 22 Aug 2026: remember what it was, so send_back_to_reconcile can put it
        # back exactly instead of guessing. Every forward move has a way back.
        try:
            _prev_field_ready()
            _cur = _s(frappe.db.get_value("Quarry Block", hit["name"], "status"))
            if _cur and _cur != "At Port":
                frappe.db.set_value("Quarry Block", hit["name"], "status_before_at_port", _cur)
        except Exception:
            pass
        res = set_status(hit["name"], "At Port",
                         "arrival step skipped: {0}".format(_s(note) or "no reason given"),
                         machine="server (skip arrivals)")
        if res.get("ok"):
            skipped.append(bn)
    frappe.db.commit()
    return {"arrival": pa.name, "moved": moved, "at_port": len(skipped),
            "refused": refused}


# ===========================================================================
# Holistic transported-block ledger + actions  (append to api_arrivals.py)
#
# KEY FIX: an arrival sheet's block number matches the Delivery Challan's
# `block` field (the quarry block no) — NOT dc_block_rows.block_no (internal
# series) nor export_block_no. Matching now tries block -> block_no -> export.
#
#   ledger_view()          one flat row per transported block, with backward
#                          trace (DC -> source BI -> consignee) + lot + state.
#   move_dc_to_at_port(dc) DC-wise "skip arrivals": whole challan -> At Port.
#   resolve_block(...)     accept / link / fix keyed by (arrival, block_no).
#   block_availability(..) which DC each block is already on (dialog + BI).
# ===========================================================================


def _s(v):
    return str(v).strip() if v not in (None, "") else ""


def _consignee_names():
    m = {}
    try:
        for c in frappe.get_all("Export Consignee", fields=["name", "consignee_name"]):
            m[c.name] = c.get("consignee_name") or c.name
    except Exception:
        pass
    return m


def _lot_membership():
    """block key -> {lot, st(build/ship), title}"""
    m = {}
    for l in frappe.get_all("Export Shipment Lot", fields=["name", "docstatus"]):
        try:
            d = frappe.get_doc("Export Shipment Lot", l.name)
        except Exception:
            continue
        shipped = (d.docstatus == 1) or bool(d.get("shipped"))
        title = d.get("title") or d.get("consignee") or d.name
        for tf in d.meta.get_table_fields():
            for r in (d.get(tf.fieldname) or []):
                for k in (r.get("block_no"), r.get("export_block_no"),
                          r.get("block"), r.get("quarry_block")):
                    if _s(k):
                        m[_s(k)] = {"lot": d.name, "st": "ship" if shipped else "build", "title": title}
    return m


_PAB_FIELDS = ["parent", "block_no", "mark", "length", "width", "height",
               "cbm", "net_wt", "recon_status", "match_status", "vehicle_no",
               "resolution_note"]


def _arrival_docstatus():
    """{arrival name: docstatus} — cheap, one query."""
    try:
        return {a.name: a.docstatus for a in
                frappe.get_all("Port Arrival", fields=["name", "docstatus"],
                               limit_page_length=0)}
    except Exception:
        return {}


def _arrived_index(include_drafts=False):
    """block_no -> Port Arrival Block row (physically at port).

    FIXED 17 Aug 2026. This had no docstatus filter, so all 849 rows of the five
    DRAFT arrivals counted as "at port" — which is the whole of the 178 the Port
    & Stock page was reporting. A draft is somebody's unfinished typing; it is
    evidence, not arrival.

    Confirmed arrivals only by default. `_evidence_index()` returns the drafts
    separately so they stay visible instead of vanishing (A3)."""
    ds = _arrival_docstatus()
    idx = {}
    for p in frappe.get_all("Port Arrival Block", fields=_PAB_FIELDS,
                            limit_page_length=0):
        k = _s(p.block_no)
        if not k:
            continue
        if not include_drafts and ds.get(p.parent, 0) != 1:
            continue
        idx.setdefault(k, p)
    return idx


def _evidence_index():
    """block_no -> Port Arrival Block row that sits on an UNSUBMITTED arrival.

    Arrival evidence, unconfirmed. Shown in its own bucket on the ledger so the
    number is honest without the rows disappearing from the screen."""
    ds = _arrival_docstatus()
    idx = {}
    for p in frappe.get_all("Port Arrival Block", fields=_PAB_FIELDS,
                            limit_page_length=0):
        k = _s(p.block_no)
        if k and ds.get(p.parent, 0) == 0:
            idx.setdefault(k, p)
    return idx


def _block_by_name():
    """Quarry Block keyed by its RECORD NAME only.

    22 Aug 2026 - found while chasing his identity question, and worse than the
    original fault. `_block_status_index` keys the SAME dictionary by export
    number, quarry number AND record name. On this site 62 numbers are two of
    those at once, so `idx["1353"]` resolves to whichever record the loop happened
    to reach first - block 1353 or block 1865 depending on iteration order. A
    block's status could therefore be read off the wrong block.

    A DC row carries a real Link to the block. A Link is unambiguous. So where a
    Link exists it is used, through this index, and nothing is inferred at all.
    """
    out = {}
    try:
        for qb in frappe.get_all("Quarry Block",
                                 fields=["name", "block_number", "export_block_no", "status"],
                                 limit_page_length=0):
            out[_s(qb.name)] = {"name": qb.name, "status": _s(qb.status),
                                "block_number": _s(qb.block_number),
                                "export_block_no": _s(qb.export_block_no)}
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Dolphin block by name")
    return out


def _block_status_index():
    """Every identifier a Quarry Block answers to -> (record name, real status).

    ledger_view never read Quarry Block.status at all; it inferred a state from
    arrival rows. So when the 56 blocks were corrected on 17 Aug the page went on
    saying 178 regardless. Now the block's own status is the spine of the row and
    the arrival is supporting evidence."""
    idx = {}
    try:
        for qb in frappe.get_all("Quarry Block",
                                 fields=["name", "block_number", "export_block_no", "status"],
                                 limit_page_length=0):
            rec = {"name": qb.name, "status": _s(qb.status),
                   "block_number": _s(qb.block_number),
                   "export_block_no": _s(qb.export_block_no)}
            for k in (qb.export_block_no, qb.block_number, qb.name):
                k = _s(k)
                if k:
                    idx.setdefault(k, rec)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Dolphin block status index")
    return idx


# Quarry Block.status -> ledger state. The block's own status wins over inferred
# state, so a correction made to the data shows on the page immediately (A3).
_LADDER_STATE = {
    "Dispatched/Transported": "await",
    # "In Delivery Challan" is deliberately ABSENT — see PRE_DISPATCH below.
    "At Port": "port",
    "At Bannikoppa Station yard": "port",
    "Reconciled": "recon",
    "Ready for Export Lot": "ready",
    "In Export Shipment Lot": "lot",
    "Shipped": "load",
    "Sold": "load",
}

# A BLOCK IS DISPATCHED ONLY WHEN ITS CHALLAN IS SUBMITTED (his rule, 21 Aug 2026).
#
# His words: "Blocks in Dc draft should not be listed under stock and port ... Dc drafts
# are created in advance since the DMG website demands all the details ready while issuing
# permits hence the draft. You consider blocks dispatched only after submission of DC".
#
# A draft challan is permit paperwork, not a dispatch. Until now a block whose own status
# still read "In Delivery Challan" was mapped to "await" and listed on Port & Stock as a
# transported block awaiting arrival — which is how block 247 appeared at the port while
# it was still sitting on a draft challan. These statuses mean the block has NOT left.
PRE_DISPATCH = {"In Stock", "Buyer Marked", "In Delivery Challan"}

# Labels the page shows for each state, kept server-side so every screen agrees.
STATE_LABELS = {
    "await": "Awaiting arrival",
    "unconfirmed": "Arrival evidence, unconfirmed",
    "port": "At port",
    "recon": "Reconciled",
    "ready": "Ready for export lot",
    "lot": "In shipment lot",
    "load": "Shipped",
    "mis": "Dimension mismatch",
    "held": "Held",
    "dmg": "Damaged",
    "orphan": "At port, not on any challan",
}


@frappe.whitelist()
def ledger_states():
    """The state vocabulary, for the page legend."""
    return STATE_LABELS


@frappe.whitelist()
def ledger_view():
    """One flat row per transported block: block -> DC -> BI/consignee, plus
    arrival/port reality, lot membership and a single state.

    REWRITTEN 17 Aug 2026 (A3). Three things were wrong and all three are fixed:
      * draft arrivals counted as arrivals — now filtered, with their own bucket
      * the block's real status was never read — now it is the spine of the row
      * a number was resolved against the record id — now never."""
    lots = _lot_membership()
    arrived = _arrived_index()          # submitted arrivals only (17 Aug fix)
    evidence = _evidence_index()        # draft arrivals — shown, never counted
    qbs = _block_status_index()         # the block's OWN status is the spine
    qbn = _block_by_name()              # unambiguous: keyed by record name only
    cons = _consignee_names()
    emap = _export_map()

    dcs = {x.name: x for x in frappe.get_all(
        "Delivery Challan", filters={"docstatus": 1},
        fields=["name", "export_consignee", "shipping_mark", "vehicle", "port_of_loading"])}

    ports = {p.name: p.port_code for p in frappe.get_all("Indian Port", fields=["name", "port_code"])}
    rows, seen = [], set()

    if dcs:
        for r in frappe.get_all(
            "DC Block Row",
            filters={"parenttype": "Delivery Challan", "parent": ["in", list(dcs.keys())]},
            fields=["parent", "block", "block_no", "export_block_no",
                    "source_inspection", "grade", "length_gross", "width_gross",
                    "height_gross", "gross_volume", "gross_tonnage"],
            limit_page_length=0,
        ):
            dc = dcs.get(r.parent)
            if not dc:
                continue
            # ================================================================
            # 22 Aug 2026 - THE END OF THE RECORD ID LEAK. His words:
            #   "can you put an end to internal ID reflecting again and again
            #    this is really hurting us"
            #
            # THIS LINE WAS THE SOURCE. `r.block` is the Link to Quarry Block -
            # the record id - and it was being emitted as the row's `block_no`.
            # Every screen and every payload downstream read `block_no` and got
            # a record id, which is why Move to At Port silently refused 219
            # blocks, why the challan comparison "lost" rows the agency had sent,
            # and why a port report once said 25 blocks never left the quarry.
            #
            # From here on the two are kept apart, permanently:
            #   block_no  = a number a PERSON uses (export number, else the
            #               quarry number). Never the record id.
            #   qb        = the record id, and the only field that carries it.
            # `keys` still holds all three for INDEX LOOKUPS - matching may use
            # any of them - but nothing that leaves this function may.
            # ================================================================
            keys = [_s(k) for k in (r.block, r.block_no, r.export_block_no) if _s(k)]
            if not keys:
                continue
            pa = next((arrived[k] for k in keys if k in arrived), None)
            ev = next((evidence[k] for k in keys if k in evidence), None)
            lot = next((lots[k] for k in keys if k in lots), None)
            # The Link on the row is authoritative. Only when there is no Link do
            # we fall back to matching by number, which is where collisions live.
            qb = qbn.get(_s(r.block)) or next((qbs[k] for k in keys if k in qbs), None)
            qstat = _s(qb["status"]) if qb else ""

            # A block that has not been dispatched does not belong on this page at all
            # (his rule, 21 Aug 2026 — see PRE_DISPATCH). It is on a submitted challan
            # here, but its own status still says it never left; that happens when a
            # later DRAFT challan re-added it and pulled it back to "In Delivery Challan".
            # Two exceptions, both of which must stay visible rather than be hidden:
            #   * a CONFIRMED arrival saying it did reach the port — a real contradiction
            #   * membership of an export shipment lot — 37 blocks are in a lot today
            #     while their status still reads In Stock; dropping those would empty
            #     the "In lot" count, which would be a worse lie than the one being fixed.
            if qstat in PRE_DISPATCH and pa is None and lot is None:
                continue

            # State, in order of authority:
            #   1. the block's own status, when it is one the ladder knows
            #   2. lot membership
            #   3. a CONFIRMED arrival row
            #   4. arrival evidence on a draft — its own bucket, never "at port"
            if qstat in _LADDER_STATE:
                state = _LADDER_STATE[qstat]
                # A3, second pass: the block's own status winning is right, but it
                # buried the thing worth seeing. When the block says it has NOT
                # reached the port and a draft arrival sheet says it has, that
                # disagreement is the whole point — surface it rather than letting
                # the block's status quietly swallow it.
                if state == "await" and pa is None and ev is not None:
                    state = "unconfirmed"
            elif lot and lot["st"] == "ship":
                state = "load"
            elif lot:
                state = "lot"
            elif pa is not None:
                rs = _s(pa.recon_status).lower()
                state = ("dmg" if "damage" in rs else
                         "held" if ("hold" in rs or "held" in rs) else
                         "mis" if ("mismatch" in rs or "dimension" in rs) else "port")
            elif ev is not None:
                state = "unconfirmed"
            else:
                state = "await"

            _export = _s(r.export_block_no) or _s(emap.get(_s(r.block_no))) or ""
            _quarry = _s(r.block_no) or _s((qb or {}).get("block_number"))
            _public = _export or _quarry
            if not _public:
                # No human-facing number exists for this row. Emitting the record
                # id here is exactly the fault being fixed, so the row is left out
                # and reported instead of quietly becoming a wrong match.
                continue

            rows.append({
                "block_no": _public,
                "export_block_no": _export,
                "quarry_block_no": _quarry,
                "mark": (pa.mark if pa else None) or dc.shipping_mark,
                "dc": dc.name,
                "consignee": cons.get(dc.export_consignee, dc.export_consignee),
                "source_bi": r.source_inspection,
                "grade": r.grade,
                "dc_l": r.length_gross, "dc_w": r.width_gross, "dc_h": r.height_gross,
                "dc_cbm": r.gross_volume, "ton": r.gross_tonnage,
                "pt_l": (pa.length if pa else None), "pt_w": (pa.width if pa else None),
                "pt_h": (pa.height if pa else None), "pt_cbm": (pa.cbm if pa else None),
                "net_wt": (pa.net_wt if pa else None),
                "arrival": (pa.parent if pa else (ev.parent if ev else None)),
                "arrival_confirmed": 1 if pa else 0,
                "recon_status": (pa.recon_status if pa else (ev.recon_status if ev else None)),
                "resolution_note": (pa.resolution_note if pa else
                                    (ev.resolution_note if ev else None)),
                "block_status": qstat,
                "qb": (qb["name"] if qb else None),
                "lot": (lot["lot"] if lot else None),
                "lot_title": (lot["title"] if lot else None),
                "truck": dc.vehicle, "port": dc.port_of_loading, "port_code": ports.get(dc.port_of_loading, dc.port_of_loading), "state": state, "source": "dc",
            })
            for k in keys:
                seen.add(k)

    # arrived but not on any submitted DC -> excess / opening / accepted.
    # Draft-arrival rows are included too, but flagged unconfirmed rather than
    # silently counted as stock at the port.
    merged = dict(evidence)
    merged.update(arrived)
    for k, pa in merged.items():
        if k in seen:
            continue
        confirmed = k in arrived
        lot = lots.get(k)
        qb = qbs.get(k)
        qstat = _s(qb["status"]) if qb else ""
        rs = _s(pa.recon_status).lower()
        if qstat in _LADDER_STATE:
            state = _LADDER_STATE[qstat]
        elif lot and lot["st"] == "ship":
            state = "load"
        elif lot:
            state = "lot"
        elif not confirmed:
            state = "unconfirmed"
        elif "resolved" in rs or "accept" in rs or "opening" in rs:
            state = "port"
        else:
            state = "orphan"
        # 22 Aug 2026: the same rule on this path. An agency sheet may name a
        # block by its record id - ARR-27Jul2026-NA does exactly that - so the key
        # `k` is not safe to publish. Resolve it to a number a person uses, and if
        # there isn't one, say so rather than leaking the id.
        _k_export = _s(emap.get(k)) or _s((qb or {}).get("export_block_no"))
        _k_quarry = _s((qb or {}).get("block_number"))
        _k_public = _k_export or _k_quarry
        if not _k_public:
            _k_public = k if not (qb and _s(qb.get("name")) == k) else ""
        if not _k_public:
            continue

        rows.append({
            "block_no": _k_public,
            "export_block_no": _k_export,
            "quarry_block_no": _k_quarry,
            "mark": pa.mark, "dc": None, "consignee": None,
            "arrival_confirmed": 1 if confirmed else 0,
            "block_status": qstat, "qb": (qb["name"] if qb else None),
            "source_bi": None, "grade": None,
            "dc_l": None, "dc_w": None, "dc_h": None, "dc_cbm": None, "ton": None,
            "pt_l": pa.length, "pt_w": pa.width, "pt_h": pa.height,
            "pt_cbm": pa.cbm, "net_wt": pa.net_wt,
            "arrival": pa.parent, "recon_status": pa.recon_status,
            "lot": (lot["lot"] if lot else None),
            "lot_title": (lot["title"] if lot else None),
            "truck": pa.vehicle_no, "port": None, "port_code": None, "state": state, "source": "arrival",
        })
    return rows


@frappe.whitelist()
def move_dc_to_at_port(dc=None):
    """DC-wise skip-arrivals: place every block on this challan At Port."""
    if not dc:
        frappe.throw("No challan given.")
    d = frappe.get_doc("Delivery Challan", dc)
    blocks = []
    for r in (d.get("dc_block_rows") or []):
        # 17 Aug: `block` is the LINK — the record id. Sending that downstream is
        # exactly the record-id leak that produced the wrong At Port matches, and
        # it also breaks the standing rule that after the DC only export numbers
        # travel. Prefer the export number, then the quarry number.
        bn = _s(r.get("export_block_no")) or _s(r.get("block_no")) or _s(r.get("block"))
        if bn:
            blocks.append({"block_no": bn, "dc": dc})
    if not blocks:
        frappe.throw("This challan has no blocks.")
    return move_to_at_port(blocks)


@frappe.whitelist()
def resolve_block(arrival=None, block_no=None, action="accept", dc=None, length=None, width=None, height=None, cbm=None, weight=None, note=None):
    """Resolve a flagged block, keyed by (arrival, block_no). Once accepted the
    block reads as Resolved and shows under All at port."""
    block_no = _s(block_no)
    if not block_no:
        frappe.throw("No block number.")

    alt = _pab_alt_keys(block_no) or {block_no}
    name = None
    if arrival:
        name = frappe.db.get_value("Port Arrival Block",
                                   {"parent": arrival, "block_no": ["in", list(alt)]}, "name")
    if not name:
        name = frappe.db.get_value("Port Arrival Block",
                                   {"block_no": ["in", list(alt)]}, "name")
    if not name:
        frappe.throw("Block {0} not found at port.".format(block_no))

    updates = {}
    if action in ("accept", "release", "accept_extra"):
        updates["recon_status"] = "Resolved"
    elif action == "use_dc":
        updates["recon_status"] = "Resolved"
    elif action == "hold":
        updates["recon_status"] = "Held"
    elif action == "link_dc" and dc:
        updates["matched_dc"] = dc
        updates["recon_status"] = "Resolved"
    elif action == "unresolve":
        updates["recon_status"] = ""
    else:
        updates["recon_status"] = "Resolved"

    if action in ("modify", "unresolve"):
        for _f, _v in (("length", length), ("width", width), ("height", height), ("cbm", cbm), ("weight", weight)):
            if _v not in (None, ""):
                updates[_f] = flt(_v)
    if note not in (None, ""):
        _bm = frappe.get_meta("Port Arrival Block")
        if _bm.has_field("resolution_note"):
            updates["resolution_note"] = note
        if _bm.has_field("resolved_by"):
            updates["resolved_by"] = frappe.session.user
        if _bm.has_field("resolved_on"):
            updates["resolved_on"] = now_datetime()
    frappe.db.set_value("Port Arrival Block", name, updates, update_modified=False)
    frappe.db.commit()
    return {"block_no": block_no, "action": action, "ok": True}


@frappe.whitelist()
def block_availability(blocks=None):
    """Given block numbers, return {block: {"dc": name, "draft": 0|1} or None}
    — which challan each block already sits on. Powers the Available / on-DC
    indicator in the add-blocks dialog and the Buyer Inspection screen.

    HIS RULE, 23 Aug 2026: "anything inside dc draft shouldnt be considered
    yet ... only after submit". A draft challan has not dispatched anything,
    so a block sitting on one is still AVAILABLE. It is worth saying out loud
    though — otherwise the same block quietly lands on two challans — so the
    draft is reported with draft=1 and the caller shows it as a note, never
    as a block. A SUBMITTED challan (draft=0) is the real claim.

    A submitted challan always wins over a draft for the same block.
    Cancelled challans (docstatus 2) are ignored entirely."""
    import json as _json
    if isinstance(blocks, str):
        blocks = _json.loads(blocks)
    blocks = [_s(b) for b in (blocks or []) if _s(b)]
    if not blocks:
        return {}
    out = {b: None for b in blocks}
    rows = frappe.get_all(
        "DC Block Row",
        filters={"parenttype": "Delivery Challan"},
        fields=["parent", "block", "block_no", "export_block_no"],
        limit_page_length=0,
    )
    state = {c.name: c.docstatus for c in frappe.get_all(
        "Delivery Challan", filters={"docstatus": ["<", 2]},
        fields=["name", "docstatus"])}
    for r in rows:
        ds = state.get(r.parent)
        if ds is None:            # cancelled, or gone
            continue
        is_draft = 1 if ds == 0 else 0
        for k in (_s(r.block), _s(r.block_no), _s(r.export_block_no)):
            if k not in out:
                continue
            cur = out[k]
            # first answer wins, except a submitted challan displaces a draft
            if cur is None or (cur.get("draft") and not is_draft):
                out[k] = {"dc": r.parent, "draft": is_draft}
    return out


@frappe.whitelist()
def active_lots():
    """Building/Ready (not shipped) Export Shipment Lots for the push picker."""
    out = []
    for l in lots_view():
        if l.get("status") != "ship" and not l.get("shipped"):
            out.append({
                "name": l["name"],
                "title": l.get("title"),
                "consignee": l.get("consignee"),
                "total_blocks": l.get("total_blocks"),
            })
    return out


@frappe.whitelist()
def add_blocks_to_lot(lot=None, rows=None):
    """Append at-port blocks to an existing (building) Export Shipment Lot."""
    if not lot:
        frappe.throw("No lot given.")
    data = _json.loads(rows) if isinstance(rows, str) else (rows or [])
    if not data:
        frappe.throw("No blocks to add.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    tf = None
    for t in d.meta.get_table_fields():
        tf = t.fieldname
        break
    if not tf:
        frappe.throw("Lot has no block table.")
    existing = set()
    for r in (d.get(tf) or []):
        for k in (r.get("block_no"), r.get("block"), r.get("quarry_block")):
            if _s(k):
                existing.add(_s(k))
    added = 0
    tc = flt(d.get("total_cbm"))
    tt = flt(d.get("total_net_tonnage"))
    for r in data:
        bn = _s(r.get("block_no"))
        if not bn or bn in existing:
            continue
        qb = _qb_by_any(r.get("quarry_block") or bn)
        _l = cint(r.get("length")); _w = cint(r.get("width")); _h = cint(r.get("height"))
        _cbm = flt(r.get("cbm")) or (round(_l * _w * _h / 1e6, 3) if (_l and _w and _h) else 0)
        _mt = flt(r.get("net_tonnage") or r.get("mt") or r.get("net_ton") or 0)
        if not _mt:
            _factor = flt(r.get("tonnage_factor")) or 2.7
            _mt = round(_cbm * _factor, 3)
        _kgs = cint(r.get("kgs") or r.get("net_kgs") or round(_mt * 1000))
        ch = d.append(tf, {})
        if ch.meta.has_field("block"):
            ch.block = r.get("quarry_block") or (qb.get("name") if qb else None)
        if ch.meta.has_field("block_no"):
            ch.block_no = r.get("block_no") or (qb.get("block_number") if qb else bn)
        for _fld, _v in (("length", _l), ("width", _w), ("height", _h)):
            if ch.meta.has_field(_fld):
                ch.set(_fld, _v)
        if ch.meta.has_field("cbm"):
            ch.cbm = _cbm
        if ch.meta.has_field("net_tonnage"):
            ch.net_tonnage = _mt
        if ch.meta.has_field("net_kgs"):
            ch.net_kgs = _kgs
        if ch.meta.has_field("source_dc"):
            ch.source_dc = r.get("matched_dc") or r.get("source_dc") or ""
        if ch.meta.has_field("source_arrival"):
            ch.source_arrival = r.get("arrival") or ""
        if ch.meta.has_field("export_block_no"):
            ex = _s(r.get("export_block_no")) or (_s(qb.get("export_block_no")) if qb else "")
            ch.export_block_no = ex
        existing.add(bn)
        added += 1
        tc += _cbm
        tt += _mt
    if d.meta.has_field("block_count"):
        d.block_count = len(d.get(tf) or [])
    if d.meta.has_field("total_cbm"):
        d.total_cbm = round(tc, 2)
    if d.meta.has_field("total_net_tonnage"):
        d.total_net_tonnage = round(tt, 3)
    if d.meta.has_field("total_net_kgs"):
        d.total_net_kgs = cint(round(tt * 1000))
    d.flags.ignore_mandatory = True
    d.save(ignore_permissions=True)
    frappe.db.commit()
    return {"lot": d.name, "added": added}


@frappe.whitelist()
def mark_lot_shipped(lot=None, vessel=None, ship_date=None, bl_no=None):
    """Mark an Export Shipment Lot as Shipped -> moves it to Exported Shipments.
    Block status cascades to Shipped via the Server Script on save."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    if d.meta.has_field("status"):
        d.status = "Shipped"
    if d.meta.has_field("shipped"):
        d.shipped = 1
    if vessel and d.meta.has_field("vessel"):
        d.vessel = vessel
    if ship_date:
        if d.meta.has_field("ship_date"):
            d.ship_date = ship_date
        elif d.meta.has_field("shipment_date"):
            d.shipment_date = ship_date
    if bl_no and d.meta.has_field("bl_no"):
        d.bl_no = bl_no
    d.flags.ignore_mandatory = True
    d.save(ignore_permissions=True)
    frappe.db.commit()
    return {"lot": d.name, "status": "Shipped"}


def _xls_tokens(content, file_url=""):
    """Every non-empty cell of an uploaded .xls/.xlsx as a list of string tokens."""
    toks = []
    name = (file_url or "").lower()

    def _num(v):
        try:
            fv = float(v)
            if fv == int(fv):
                return str(int(fv))
            return str(v)
        except Exception:
            return _s(v)

    if not name.endswith(".xlsx"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            for sh in wb.sheets():
                for r in range(sh.nrows):
                    for c in range(sh.ncols):
                        v = sh.cell_value(r, c)
                        if v is None or v == "":
                            continue
                        toks.append(_num(v))
            return toks
        except Exception:
            pass
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None or v == "":
                        continue
                    toks.append(_num(v))
    except Exception:
        pass
    return toks


@frappe.whitelist()
def import_lot_blocks_xls(lot=None, file_url=None):
    """Read block numbers (quarry OR export) from an uploaded .xls/.xlsx and add
    every matching At-Port block to the given Export Shipment Lot. Uses the same
    lookup/enrichment as the Add Blocks dialog, so system data stays authoritative."""
    if not lot:
        frappe.throw("No lot given.")
    if not file_url:
        frappe.throw("No file uploaded.")
    content = _arrival_file_bytes(file_url)
    if content is None:
        frappe.throw("Uploaded file could not be read.")
    tokens = _xls_tokens(content, file_url)
    if not tokens:
        return {"added": 0, "matched": 0, "not_found": [],
                "message": "No cells could be read from the file."}
    avail = at_port_available(lot)
    by_key = {}
    for b in avail:
        # EXPORT numbers only -- quarry block numbers are intentionally NOT matched
        ex = _s(b.get("export_block_no"))
        if ex:
            by_key[ex] = b
    used = set()
    rows = []
    not_found = []
    seen_tok = set()
    for tok in tokens:
        t = _s(tok)
        if not t or t in seen_tok:
            continue
        seen_tok.add(t)
        b = by_key.get(t)
        if not b:
            if any(ch.isdigit() for ch in t):
                not_found.append(t)
            continue
        bn = _s(b.get("block_no"))
        if bn in used:
            continue
        used.add(bn)
        qb = frappe.db.get_value("Quarry Block", {"block_number": bn},
            ["name", "delivery_challan", "port_net_wt", "gross_tonnage", "tonnage_factor"],
            as_dict=True) or {}
        fac = flt(qb.get("tonnage_factor")) or 2.7
        mt = flt(qb.get("port_net_wt")) or flt(qb.get("gross_tonnage")) or (flt(b.get("cbm")) * fac)
        rows.append({
            "block_no": bn, "quarry_block": qb.get("name"),
            "length": b.get("length"), "width": b.get("width"), "height": b.get("height"),
            "cbm": b.get("cbm"), "mt": mt, "matched_dc": qb.get("delivery_challan") or "",
        })
    if not rows:
        return {"added": 0, "matched": 0, "not_found": not_found[:60],
                "message": "None of the numbers in the file matched an available At-Port block."}
    res = add_blocks_to_lot(lot, rows)
    return {"added": (res or {}).get("added", len(rows)),
            "matched": len(rows), "not_found": not_found[:60]}


def _lot_table_field(d):
    for t in d.meta.get_table_fields():
        return t.fieldname
    return None


@frappe.whitelist()
def create_empty_lot(title=None, consignee=None, mark=None, vessel=None):
    """Create an empty Export Shipment Lot (status Ready); return its name.
    Optional title/consignee/mark/vessel are set from the New-lot dialog."""
    lot = frappe.new_doc("Export Shipment Lot")
    if lot.meta.has_field("shipment_date"):
        lot.shipment_date = frappe.utils.today()
    if lot.meta.has_field("status"):
        lot.status = "Ready"
    if title and lot.meta.has_field("lot_title"):
        lot.lot_title = title
    if consignee and lot.meta.has_field("export_consignee"):
        lot.export_consignee = consignee
    if mark and lot.meta.has_field("shipping_mark"):
        lot.shipping_mark = mark
    if vessel and lot.meta.has_field("vessel"):
        lot.vessel = vessel
    lot.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": lot.name}


@frappe.whitelist()
def lot_detail(lot=None):
    """Header + blocks for the lot-detail view."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    tf = _lot_table_field(d)
    blocks = []
    for r in (d.get(tf) or []):
        blocks.append({
            "block": r.get("block"),
            "block_no": r.get("block_no") or r.get("block"),
            "export_block_no": r.get("export_block_no") or "",
            "length": r.get("length"), "width": r.get("width"), "height": r.get("height"),
            "cbm": r.get("cbm"), "net_tonnage": r.get("net_tonnage"),
            "grade": r.get("grade"), "source_dc": r.get("source_dc"),
        })
    shipped = (d.get("status") == "Shipped") or bool(d.get("shipped"))
    return {
        "name": d.name, "status": d.get("status"), "shipped": 1 if shipped else 0,
        "vessel": d.get("vessel") or "", "bl_no": d.get("bl_no") or "",
        "ship_date": _s(d.get("shipment_date") or ""),
        "consignee": d.get("export_consignee") or "",
        "shipping_document": d.get("shipping_document") or "",
        "block_count": len(blocks), "blocks": blocks,
    }


@frappe.whitelist()
def remove_lot_block(lot=None, block_no=None):
    """Remove one block from a lot; if shipped, return that block to At Port."""
    if not lot or not block_no:
        frappe.throw("Lot and block required.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    tf = _lot_table_field(d)
    if not tf:
        frappe.throw("Lot has no block table.")
    target = _s(block_no)
    kept, removed_qb = [], None
    for r in (d.get(tf) or []):
        rn = _s(r.get("block_no")) or _s(r.get("block"))
        if rn == target and removed_qb is None:
            removed_qb = r.get("block") or r.get("block_no")
            continue
        kept.append(r)
    d.set(tf, kept)
    d.save(ignore_permissions=True)
    if removed_qb and ((d.get("status") == "Shipped") or bool(d.get("shipped"))):
        if frappe.db.exists("Quarry Block", removed_qb):
            frappe.db.set_value("Quarry Block", removed_qb, "status", "At Port")
    frappe.db.commit()
    return {"name": d.name, "removed": target}


@frappe.whitelist()
def reopen_lot(lot=None):
    """Reopen a lot: set status back to Ready AND return every block to At Port,
    removing the rows from the lot's block table (block_count -> 0) and unlinking
    the Shipping Document. Delegates the row removal + At Port flip to
    return_blocks_from_lot so both the desk 'Reopen' button and the Export Hub
    behave identically. The Quarry Block status is set idempotently and no Port
    Arrival row is created, so a block is never duplicated at port."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    if d.meta.has_field("status"):
        d.status = "Ready"
    if d.meta.has_field("shipped"):
        d.shipped = 0
    d.save(ignore_permissions=True)
    res = return_blocks_from_lot(lot=lot, blocks=None) or {}
    return {"name": lot, "status": "Ready", "returned": res.get("returned", 0)}


@frappe.whitelist()
def link_shipping_document(lot=None, shipping_document=None):
    """Link an existing Shipping Document to a lot."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    if not d.meta.has_field("shipping_document"):
        frappe.throw("Lot has no shipping_document field.")
    d.shipping_document = shipping_document or None
    d.save(ignore_permissions=True)
    frappe.db.commit()
    return {"name": d.name, "shipping_document": d.get("shipping_document") or ""}


@frappe.whitelist()
def list_shipping_documents():
    """Recent Shipping Documents for the link picker."""
    rows = frappe.get_all("Shipping Document", fields=["name"], order_by="creation desc", limit_page_length=50)
    out = []
    for r in rows:
        d = frappe.db.get_value("Shipping Document", r.name, ["export_consignee", "shipment_date", "vessel"], as_dict=True) or {}
        out.append({"name": r.name, "consignee": (d.get("export_consignee") or ""), "date": _s(d.get("shipment_date") or ""), "vessel": (d.get("vessel") or "")})
    return out


@frappe.whitelist()
def at_port_available(lot=None):
    """At-Port blocks not already in any lot - for the in-lot Add picker."""
    names = frappe.get_all("Quarry Block", filters={"status": "At Port"},
        fields=["name", "block_number", "export_block_no", "granite_quality_grade",
                "length_gross", "width_gross", "height_gross", "gross_volume"],
        limit_page_length=0)
    in_lot = set()
    for lb in frappe.get_all("Shipment Lot Block", fields=["block_no", "block"], limit_page_length=0):
        for k in (lb.get("block_no"), lb.get("block")):
            if k:
                in_lot.add(_s(k))
    out = []
    for b in names:
        bn = _s(b.get("block_number") or b.get("name") or "")
        if not bn or bn in in_lot:
            continue
        out.append({"block_no": bn, "export_block_no": b.get("export_block_no") or "",
            "grade": b.get("granite_quality_grade") or "", "length": b.get("length_gross"),
            "width": b.get("width_gross"), "height": b.get("height_gross"), "cbm": b.get("gross_volume")})
    return out


@frappe.whitelist()
def xls_source():
    """Imported arrival files as XLS sources with provenance: 'email' when the
    arrival carries an arrivals-inbox sender/subject, else 'direct' import."""
    meta = frappe.get_meta("Port Arrival")
    def has(f):
        return meta.has_field(f)
    fields = ["name", "arrival_date", "creation", "total_blocks"]
    for f in ("source_file", "source_sheet", "email_subject", "email_sender",
              "mark", "shipper", "total_cbm", "total_net_wt"):
        if has(f):
            fields.append(f)
    out = []
    for a in frappe.get_all("Port Arrival", fields=fields,
                            order_by="creation desc", limit_page_length=0):
        sender = a.get("email_sender") or ""
        subject = a.get("email_subject") or ""
        provenance = "email" if (sender or subject) else "direct"
        out.append({
            "arrival": a.name,
            "file": a.get("source_file") or "",
            "sheet": a.get("source_sheet") or "",
            "rows": a.get("total_blocks") or 0,
            "mark": a.get("mark") or "",
            "sender": sender,
            "subject": subject,
            "provenance": provenance,
            "date": str(a.get("arrival_date") or a.get("creation") or "")[:10],
            "cbm": a.get("total_cbm") or 0,
            "net_wt": a.get("total_net_wt") or 0,
        })
    return out



def _arrival_file_bytes(url):
    """Bytes of a stored File by its file_url, private or public."""
    if not url:
        return None
    try:
        fdoc = frappe.get_doc("File", {"file_url": url})
        return fdoc.get_content()
    except Exception:
        pass
    try:
        from frappe.utils.file_manager import get_file
        return get_file(url)[1]
    except Exception:
        return None


@frappe.whitelist()
def arrival_xls_grid(arrival=None, max_rows=5000):
    """Raw source sheet as the carrier sent it -- full grid of cell values PLUS
    each cell's original fill colour (when the .xls carries formatting), so the
    inline viewer looks like the shipping-agency file. Per-row parse tags are
    still returned for the parsed/skipped legend."""
    if not arrival:
        return {}
    pa = frappe.get_doc("Port Arrival", arrival)
    src = getattr(pa, "source_file", None) or ""
    if not src:
        return {"error": "No source file stored on this arrival."}
    content = _arrival_file_bytes(src)
    if content is None:
        return {"error": "Source file not found: " + src}
    try:
        import xlrd
    except Exception:
        return {"error": "xlrd not installed on this bench."}
    fmt = True
    try:
        wb = xlrd.open_workbook(file_contents=content, formatting_info=True)
    except Exception:
        fmt = False
        try:
            wb = xlrd.open_workbook(file_contents=content)
        except Exception as e:
            return {"error": "Could not read sheet: " + str(e)}

    def _bg(sh, r, c):
        if not fmt:
            return ""
        try:
            xf = wb.xf_list[sh.cell_xf_index(r, c)]
            if xf.background.fill_pattern != 1:
                return ""
            rgb = wb.colour_map.get(xf.background.pattern_colour_index)
            if not rgb:
                return ""
            return "#%02x%02x%02x" % rgb
        except Exception:
            return ""

    single = wb.nsheets == 1
    grid, tags, colors, used, hr = [], [], [], None, None
    for sh in wb.sheets():
        if not _xls_is_dolphin_sheet(sh, single):
            continue
        hr, cm = _xls_header(sh)
        bcol = cm.get("block_no") if cm else None
        used = sh.name
        n = min(int(max_rows), sh.nrows)
        for r in range(n):
            row = [_xls_s(sh.cell_value(r, c)) for c in range(sh.ncols)]
            grid.append(row)
            colors.append([_bg(sh, r, c) for c in range(sh.ncols)])
            if hr is None or r == hr:
                tag = "head"
            elif r < hr:
                tag = "pre"
            else:
                bno = _xls_s(sh.cell_value(r, bcol)) if bcol is not None else ""
                joined = " ".join(str(x).lower() for x in row)
                if not bno or not _re.search(r"\d", bno):
                    tag = "skip"
                elif "total" in joined:
                    tag = "skip"
                else:
                    tag = "parsed"
            tags.append(tag)
        break
    return {
        "sheet": used,
        "grid": grid,
        "tags": tags,
        "colors": colors,
        "has_colors": fmt,
        "header_row": hr,
        "file": src,
        "counts": {"parsed": tags.count("parsed"), "skipped": tags.count("skip")},
    }



@frappe.whitelist()
def sync_arrivals_email():
    """Manually pull incoming email accounts on demand (Check new mail), then
    auto-parse any arrival emails that came in but have no blocks yet."""
    n = 0
    try:
        for ea in frappe.get_all("Email Account", filters={"enable_incoming": 1}, pluck="name"):
            try:
                frappe.get_doc("Email Account", ea).receive()
                n += 1
            except Exception:
                frappe.log_error(frappe.get_traceback(), "sync_arrivals_email")
    except Exception:
        frappe.log_error(frappe.get_traceback(), "sync_arrivals_email")
    parsed = []
    try:
        parsed = parse_email_arrivals().get("parsed", [])
    except Exception:
        frappe.log_error(frappe.get_traceback(), "sync_arrivals_email parse")
    return {"accounts": n, "parsed": parsed}


@frappe.whitelist()
def parse_email_arrivals(limit=60):
    """Auto-import blocks for Port Arrivals created from an incoming arrival email
    that still have no blocks. Finds the .xls attached to the arrival's source
    email (Communication), sets it as source_file, parses it with the Dolphin
    parser and fills the block rows -- so arrivals from ANY agency whose sheet the
    parser recognises (Elite, Puyvast, ...) import automatically. Idempotent:
    skips arrivals that already have blocks or have no parseable .xls."""
    out = []
    arrivals = frappe.get_all("Port Arrival", filters=[["docstatus", "<", 2]],
        fields=["name"], order_by="creation desc", limit_page_length=int(limit))
    for a in arrivals:
        pa = frappe.get_doc("Port Arrival", a.name)
        if pa.get("blocks"):
            continue
        xls_url = pa.get("source_file") or None
        content = _arrival_file_bytes(xls_url) if xls_url else None
        if content is None:
            xls_url = None
            comms = frappe.get_all("Communication",
                filters={"reference_doctype": "Port Arrival", "reference_name": pa.name},
                pluck="name")
            for cm in comms:
                for f in frappe.get_all("File",
                        filters={"attached_to_doctype": "Communication", "attached_to_name": cm},
                        fields=["file_url", "file_name"]):
                    nm = (f.get("file_name") or f.get("file_url") or "").lower()
                    if nm.endswith(".xls") or nm.endswith(".xlsx"):
                        xls_url = f.get("file_url")
                        break
                if xls_url:
                    break
            if not xls_url:
                continue
            content = _arrival_file_bytes(xls_url)
            if content is None:
                continue
        try:
            rows, sheet = _parse_arrival_xls(content)
        except Exception:
            frappe.log_error(frappe.get_traceback(), "parse_email_arrivals")
            continue
        if not rows:
            continue
        if xls_url and pa.meta.has_field("source_file"):
            pa.source_file = xls_url
        if pa.meta.has_field("source_sheet"):
            pa.source_sheet = sheet
        marks = [r["mark"] for r in rows if r.get("mark")]
        # A block's/arrival's "mark" is a Link to the Mark master for some agencies
        # (Elite) but new agencies (Puyvast) ship marks with no master yet -- setting
        # an unknown Link would raise LinkValidationError and abort the whole import.
        # So: only gate Link-type mark fields by existence; Data marks pass through.
        valid_marks = set()
        for m in set(marks):
            try:
                # 24 Aug 2026: this checked "Mark", which has no records at all.
                # The real master is "Shipping Mark", and the agency writes the
                # separator differently ("YL/XMN" vs the record "YL-XMN"), so it
                # is matched on the identifying characters only.
                hit = _resolve_mark(m)
                if hit:
                    valid_marks.add(hit)
            except Exception:
                pass
        def _ok_mark(doc, field, value):
            if not value:
                return False
            fm = doc.meta.get_field(field)
            if fm and fm.fieldtype == "Link":
                return value in valid_marks
            return True
        if marks and pa.meta.has_field("mark") and not pa.get("mark") and _ok_mark(pa, "mark", marks[0]):
            pa.mark = marks[0]
        for r in rows:
            b = pa.append("blocks", {})
            b.block_no = r["block_no"]
            if r.get("mark") and _ok_mark(b, "mark", r["mark"]):
                b.mark = r["mark"]
            for k in ("length", "width", "height", "cbm",
                      "vehicle_no", "yard_location", "line_no", "ado_no", "permit_no"):
                if r.get(k) is not None and b.meta.has_field(k):
                    b.set(k, r[k])
            if r.get("weight") is not None:
                if b.meta.has_field("net_wt"):
                    b.net_wt = r["weight"]
                if b.meta.has_field("a_wt") and not b.get("a_wt"):
                    b.a_wt = r["weight"]
        pa.total_blocks = len(pa.blocks)
        if pa.meta.has_field("total_cbm"):
            pa.total_cbm = round(sum(flt(b.cbm) for b in pa.blocks), 3)
        if pa.meta.has_field("total_net_wt"):
            pa.total_net_wt = round(sum(flt(b.net_wt) for b in pa.blocks), 3)
        pa.flags.ignore_mandatory = True
        pa.save(ignore_permissions=True)
        try:
            _classify(pa)
            for row in pa.blocks:
                frappe.db.set_value("Port Arrival Block", row.name, {
                    "recon_status": row.get("recon_status"),
                    "matched_dc": row.get("matched_dc"),
                    "suggested_block": row.get("suggested_block"),
                }, update_modified=False)
        except Exception:
            frappe.log_error(frappe.get_traceback(), "parse_email_arrivals classify")
        out.append({"arrival": pa.name, "blocks": len(rows), "sheet": sheet})
    frappe.db.commit()
    # 24 Aug 2026. [stated] "can you avoid arrivals with empty sheet?"
    # An incoming mail with nothing parsable attached still created a Port
    # Arrival - ARR-19Aug2026-NA was one. It holds no blocks, so it can confirm
    # nothing and move nothing, but it sits in the list looking like a sheet.
    # After every sync, anything that ended up with no rows AND no spreadsheet is
    # cleared. A sheet that HAS a file but produced no rows is a parse failure and
    # is deliberately left alone - that one needs looking at, not deleting.
    swept = []
    try:
        state = empty_arrivals() or {}
        names = [_s(x.get("sheet")) for x in (state.get("removable") or [])]
        if names:
            res = delete_empty_arrivals(sheets=names) or {}
            swept = [_s(x.get("sheet")) for x in (res.get("detail") or [])]
    except Exception:
        swept = []

    return {"parsed": out, "empty_arrivals_removed": swept}


@frappe.whitelist()
def confirm_arrival_sheet(arrival=None, note=None, machine=None):
    """A person says: this is the sheet the agency sent us.

    22 Aug 2026. Every Port Arrival on this site is a draft - not one has ever been
    confirmed - and that single missing step is what leaves 58 challans without a
    weight verdict and blocks reading "arrival evidence, unconfirmed". The app is
    right to withhold those: on his rule, "even 1 block is error we had to pay huge
    penalty in Lakhs of rupees and dollars", so nothing is taken as proof of arrival
    until a person has looked at the sheet.

    This is that step, and nothing more. It does not move a single block; it only
    makes the sheet count. What the block figures then unlock is decided by the same
    tolerance and matching rules as before.
    """
    if not arrival:
        return {"error": "No sheet named."}
    try:
        pa = frappe.get_doc("Port Arrival", arrival)
    except Exception:
        return {"error": "That sheet no longer exists."}
    if pa.docstatus == 1:
        return {"ok": 1, "already": 1, "blocks": len(pa.blocks or []),
                "message": "That sheet was already confirmed."}
    if pa.docstatus == 2:
        return {"error": "That sheet is cancelled. Use 'take the confirmation back' "
                         "first - it puts the sheet into draft, and then it can be "
                         "confirmed again."}
    if not (pa.blocks or []):
        return {"error": "That sheet has no block rows on it - nothing to confirm."}

    who = _s(frappe.session.user)
    stamp = "Confirmed by {0}{1} on {2}{3}".format(
        who,
        (" (" + _s(machine) + ")") if _s(machine) else "",
        frappe.utils.now(),
        (" - " + _s(note)) if _s(note) else "")
    try:
        pa.flags.ignore_permissions = True
        pa.submit()
        frappe.db.commit()
    except Exception as e:
        frappe.db.rollback()
        return {"error": "Could not confirm it: {0}".format(e)}
    try:
        pa.add_comment("Comment", stamp)
        frappe.db.commit()
    except Exception:
        pass

    # ------------------------------------------------------------------
    # 25 Aug 2026.  [stated] "whatever tolerance of 1 ton more or less is set
    # if the blocks are within that range should be moved to at port straight
    # a way."
    #
    # Until today confirming a sheet moved nothing, and every block the agency
    # had already agreed with sat in a "ready" state waiting for somebody to
    # press a button that told them nothing they did not already know. That is
    # exactly the work he asked the app to stop making him do.
    #
    # Only the NARROW path runs here - blocks the agency's own row confirms and
    # whose challan total is inside the tonne. Anything out of tolerance, or
    # with no agency row at all, is still a person's decision with a reason.
    # Every block moved carries who moved it and why, and send_back_to_reconcile
    # puts any of them back exactly as they were.
    # ------------------------------------------------------------------
    settled = None
    try:
        settled = auto_settle_at_port(include_noconflict=0,
                                      person=who or "server (sheet confirmed)")
    except Exception as e:
        settled = {"error": str(e)}

    out = {"ok": 1, "arrival": pa.name, "blocks": len(pa.blocks or []),
           "confirmed_by": who, "note": _s(note)}
    if isinstance(settled, dict) and settled.get("ok"):
        out["settled_at_port"] = settled.get("moved") or 0
        out["settled_blocks"] = (settled.get("moved_blocks") or [])[:400]
        out["left_for_a_person"] = settled.get("left_for_a_person") or 0
        out["settle_refused"] = settled.get("refused") or []
    elif isinstance(settled, dict) and settled.get("error"):
        # The settle step must never swallow the confirmation itself.
        out["settle_error"] = settled.get("error")
    return out


@frappe.whitelist()
def unconfirm_arrival_sheet(arrival=None, reason=None, machine=None):
    """Take a confirmation back. The mirror of the button, so confirming is never
    a one-way door - his rule, "hope this can be reverted if needed"."""
    if not arrival:
        return {"error": "No sheet named."}
    if len(_s(reason)) < 4:
        return {"error": "Say why in a few words - it stays on the sheet."}
    try:
        pa = frappe.get_doc("Port Arrival", arrival)
    except Exception:
        return {"error": "That sheet no longer exists."}
    if pa.docstatus == 0:
        return {"ok": 1, "arrival": pa.name, "already": 1,
                "message": "That sheet is already back in draft."}

    # 23 Aug 2026 - WHY THIS DOES NOT CALL cancel().
    #
    # It did, on the first version, and a trial run proved that wrong: Frappe's
    # cancel puts the document into CANCELLED, not back into draft, and a
    # cancelled sheet cannot simply be confirmed again - it has to be amended into
    # a new document with a new name. "Take the confirmation back" has to leave
    # the sheet exactly as it was before it was confirmed, or it is not an undo at
    # all, and his rule is plain: "hope this can be reverted if needed or else it
    # will mess up more than required".
    #
    # A Port Arrival is the agency's sheet. It posts no ledger entry and no stock
    # movement, so its docstatus carries one meaning only - has a person confirmed
    # it. Putting that back to 0 is precisely the undo, on the parent and on every
    # row under it.
    try:
        frappe.db.set_value("Port Arrival", pa.name, "docstatus", 0,
                            update_modified=False)
        for tf in pa.meta.get_table_fields():
            for row in (pa.get(tf.fieldname) or []):
                frappe.db.set_value(tf.options, row.name, "docstatus", 0,
                                    update_modified=False)
        frappe.db.commit()
        frappe.get_doc("Port Arrival", pa.name).add_comment(
            "Comment", "Confirmation taken back by {0}{1} on {2} - {3}".format(
                _s(frappe.session.user),
                (" (" + _s(machine) + ")") if _s(machine) else "",
                frappe.utils.now(), _s(reason)))
        frappe.db.commit()
    except Exception as e:
        frappe.db.rollback()
        return {"error": "Could not take it back: {0}".format(e)}
    return {"ok": 1, "arrival": pa.name, "now": "draft"}


@frappe.whitelist()
def reparse_arrival(arrival=None):
    """Re-run the parser on an arrival's stored .xls and refresh its blocks in place."""
    if not arrival:
        return {"error": "No arrival."}
    pa = frappe.get_doc("Port Arrival", arrival)
    src = getattr(pa, "source_file", None) or ""
    if not src:
        return {"error": "No source file on this arrival."}
    content = _arrival_file_bytes(src)
    if content is None:
        return {"error": "Source file not found."}
    rows, sheet = _parse_arrival_xls(content)
    existing = {}
    for b in pa.blocks:
        existing[str(b.block_no)] = b.name
    updated, new = 0, 0
    fields = ("mark", "cbm", "weight", "length", "width", "height",
              "vehicle_no", "yard_location", "line_no", "ado_no", "permit_no")
    for row in rows:
        bno = str(row.get("block_no") or "").strip()
        if not bno:
            continue
        if bno in existing:
            vals = {}
            for f in fields:
                v = row.get(f)
                if v is not None:
                    vals[f] = v
            if vals:
                frappe.db.set_value("Port Arrival Block", existing[bno], vals, update_modified=False)
            updated += 1
        else:
            new += 1
    frappe.db.commit()
    return {"updated": updated, "new": new, "total": len(rows), "sheet": sheet}


@frappe.whitelist()
def export_arrival_xls(arrival=None):
    """Download a Port Arrival's blocks as an .xlsx."""
    if not arrival:
        frappe.throw("No arrival.")
    import openpyxl, io
    pa = frappe.get_doc("Port Arrival", arrival)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Blocks"
    cols = ["block_no", "length", "width", "height", "cbm", "weight", "mark", "matched_dc", "recon_status"]
    ws.append([c.upper() for c in cols])
    for b in pa.blocks:
        ws.append([b.get(c) for c in cols])
    buf = io.BytesIO(); wb.save(buf)
    frappe.response["filename"] = (arrival or "arrival") + ".xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "download"


@frappe.whitelist()
def export_doc_blocks_xls(doctype=None, name=None):
    """Download a document's block/row child table as .xlsx (QI/BI/DC)."""
    if not doctype or not name:
        frappe.throw("doctype and name required.")
    import openpyxl, io
    doc = frappe.get_doc(doctype, name)
    meta = frappe.get_meta(doctype)
    tf = None
    for df in meta.get_table_fields():
        fn = (df.fieldname or "").lower()
        if "block" in fn or "row" in fn:
            tf = df.fieldname
            break
    if not tf and meta.get_table_fields():
        tf = meta.get_table_fields()[0].fieldname
    rows = doc.get(tf) if tf else []
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = doctype[:31]
    if rows:
        cmeta = frappe.get_meta(rows[0].doctype)
        cols = [f.fieldname for f in cmeta.fields
                if f.fieldtype not in ("Section Break", "Column Break", "HTML", "Button")][:20]
        ws.append([c.upper() for c in cols])
        for r in rows:
            ws.append([r.get(c) for c in cols])
    buf = io.BytesIO(); wb.save(buf)
    frappe.response["filename"] = str(name) + ".xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "download"



def _qb_by_any(key):
    """Resolve a Quarry Block dict from ANY of its identifiers: docname,
    block_number, or export_block_no. Returns {name, block_number,
    export_block_no} or None. Used to keep block identity robust across the
    three overlapping number spaces.

    17 Aug 2026: the record-id branch used to come FIRST. That ordering is what
    matched 56 agency-typed numbers against autoincrementing record ids and moved
    25 wrong blocks to the port. Number spaces are now tried first and the record
    id is only a last resort — and never when the number is shared."""
    from dolphin_theme.block_resolve import try_resolve

    key = _s(key)
    if not key:
        return None
    hit, why = try_resolve(key, allow_record_name=False)
    if hit:
        return frappe._dict({"name": hit["name"],
                             "block_number": hit.get("block_number"),
                             "export_block_no": hit.get("export_block_no")})
    if why == "ambiguous":
        return None
    hit, _why = try_resolve(key, allow_record_name=True)
    if hit:
        return frappe._dict({"name": hit["name"],
                             "block_number": hit.get("block_number"),
                             "export_block_no": hit.get("export_block_no")})
    return None


def _pab_alt_keys(key):
    """All strings that could match a Port Arrival Block.block_no for one physical
    block: the given key plus the quarry docname / quarry number / export number."""
    key = _s(key)
    keys = {key} if key else set()
    qb = _qb_by_any(key)
    if qb:
        for v in (qb.get("name"), qb.get("block_number"), qb.get("export_block_no")):
            if _s(v):
                keys.add(_s(v))
    return keys


def _export_map():
    """quarry block_number -> export_block_no, for port displays."""
    m = {}
    try:
        allqb = frappe.get_all("Quarry Block",
                               fields=["name", "block_number", "export_block_no"],
                               limit_page_length=0)
        # Pass 1: quarry-number and docname keys win (they are the identifiers the
        # rest of the system passes around). Pass 2: export-number self-keys only
        # fill gaps, so a block_number that collides with another block's export
        # number still resolves to ITS OWN export number.
        for qb in allqb:
            v = str(qb.export_block_no or "").strip()
            if not v:
                continue
            for k in (qb.block_number, qb.name):
                k = str(k or "").strip()
                if k:
                    m[k] = v
        for qb in allqb:
            v = str(qb.export_block_no or "").strip()
            if v:
                m.setdefault(v, v)
    except Exception:
        pass
    return m


@frappe.whitelist()
def create_shipping_from_lot(lot=None):
    """Create a draft Shipping Document from an Export Shipment Lot and link it
    back to the lot. Reuses an already-linked Shipping Document if present.
    Header (consignee/vessel/BL/date) is carried from the lot; block rows and
    totals are copied from the lot's block table (export_block_no included)."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    tf = _lot_table_field(d)
    rows = d.get(tf) or []
    if not rows:
        frappe.throw("This lot has no blocks to ship.")

    existing = d.get("shipping_document")
    if existing and frappe.db.exists("Shipping Document", existing):
        return {"shipping_document": existing, "lot": d.name,
                "blocks": len(rows), "reused": True}

    sd = frappe.new_doc("Shipping Document")
    sd.shipment_date = _s(d.get("shipment_date")) or frappe.utils.today()
    if sd.meta.has_field("source_lot"):
        sd.source_lot = d.name
    real_mark = d.get("shipping_mark") or ""
    marks_text = d.get("export_consignee") or d.get("shipping_mark") or ""
    for fld, val in (("shipping_mark", real_mark), ("marks_nos", marks_text),
                     ("export_consignee", d.get("export_consignee")),
                     ("voyage_no", d.get("vessel")), ("bl_no", d.get("bl_no")),
                     ("lot_title", d.get("lot_title")), ("lot_description", d.get("lot_description")),
                     ("goods_description", "Granite - Roughly Trimmed Blocks"),
                     ("currency", "USD"), ("tax_treatment", "Export under LUT (No GST)"),
                     ("rate_basis", "Per Kg"), ("country_of_origin", "INDIA"),
                     ("pre_carriage_by", "ROAD"), ("terms_of_delivery", "F.O.B.")):
        if val and sd.meta.has_field(fld):
            sd.set(fld, val)

    total_cbm = 0.0
    total_mt = 0.0
    for r in rows:
        row = sd.append("blocks", {})
        if row.meta.has_field("block"):
            row.block = r.get("block")
        if row.meta.has_field("block_no"):
            row.block_no = r.get("block_no") or r.get("block")
        if row.meta.has_field("export_block_no"):
            row.export_block_no = r.get("export_block_no")
        for fld in ("length", "width", "height"):
            if row.meta.has_field(fld):
                row.set(fld, cint(r.get(fld)))
        vol = flt(r.get("cbm"))
        mt = flt(r.get("net_tonnage"))
        if row.meta.has_field("net_volume"):
            row.net_volume = vol
        if row.meta.has_field("net_tonnage"):
            row.net_tonnage = mt
        if row.meta.has_field("net_kgs"):
            row.net_kgs = cint(r.get("net_kgs") or round(mt * 1000))
        total_cbm += vol
        total_mt += mt

    if sd.meta.has_field("block_count"):
        sd.block_count = len(rows)
    if sd.meta.has_field("total_cbm"):
        sd.total_cbm = round(total_cbm, 2)
    if sd.meta.has_field("total_net_tonnage"):
        sd.total_net_tonnage = round(total_mt, 3)
    if sd.meta.has_field("total_net_kgs"):
        sd.total_net_kgs = cint(round(total_mt * 1000))
    sd.flags.ignore_mandatory = True
    sd.insert(ignore_permissions=True)

    if d.meta.has_field("shipping_document"):
        d.shipping_document = sd.name
        d.save(ignore_permissions=True)
    frappe.db.commit()
    return {"shipping_document": sd.name, "lot": d.name, "blocks": len(rows)}


@frappe.whitelist()
def paste_verify(lot=None, numbers=None):
    """Verify a pasted list of export/block numbers before adding them to a lot.
    Matches each against Quarry Block (export number first, then quarry block
    number), and returns per-row data + non-blocking warnings so the user can
    confirm / resolve / override."""
    nums = _json.loads(numbers) if isinstance(numbers, str) else (numbers or [])
    FIELDS = ["name", "block_number", "export_block_no", "status",
              "granite_quality_grade", "length_gross", "width_gross",
              "height_gross", "gross_volume", "buyer_inspection",
              "source_buyer_inspection", "delivery_challan"]
    # block -> lot membership map
    inlot = {}
    try:
        for r in frappe.get_all("Shipment Lot Block",
                fields=["parent", "block_no", "block", "export_block_no"],
                limit_page_length=0):
            for k in (r.get("block_no"), r.get("block"), r.get("export_block_no")):
                if _s(k):
                    inlot[_s(k)] = r.get("parent")
    except Exception:
        pass
    out = []
    seen = set()
    for raw in nums:
        n = _s(raw)
        if not n or n in seen:
            continue
        seen.add(n)
        qb = frappe.db.get_value("Quarry Block", {"export_block_no": n}, FIELDS, as_dict=True) \
            or frappe.db.get_value("Quarry Block", {"block_number": n}, FIELDS, as_dict=True)
        if not qb:
            out.append({"input": n, "found": 0, "warnings": ["Not found in Quarry Blocks"]})
            continue
        warns = []
        status = qb.get("status") or ""
        if "at port" not in status.lower():
            warns.append("Status is '" + (status or "-") + "' (not At Port)")
        if not (qb.get("buyer_inspection") or qb.get("source_buyer_inspection")):
            warns.append("No Buyer Inspection")
        loc = inlot.get(_s(qb.get("block_number"))) or inlot.get(_s(qb.get("export_block_no")))
        if loc:
            warns.append("Already in this lot" if loc == lot else ("Already in lot " + loc))
        out.append({
            "input": n, "found": 1, "block": qb.get("name"),
            "block_no": qb.get("block_number"), "export_no": qb.get("export_block_no"),
            "grade": qb.get("granite_quality_grade") or "",
            "L": cint(qb.get("length_gross")), "W": cint(qb.get("width_gross")),
            "H": cint(qb.get("height_gross")), "cbm": flt(qb.get("gross_volume")),
            "status": status, "in_lot": loc or "", "warnings": warns,
        })
    return out


@frappe.whitelist()
def backfill_bi_export():
    """One-time backfill: fill export_block_no on Buyer Inspection block rows
    from the Quarry Block master (matched by the row's block link, then by
    block number). Uses db.set_value so it works on submitted inspections.
    Returns before/after counts (idempotent - already-filled rows are skipped)."""
    rows = frappe.get_all(
        "Buyer Inspection Block",
        fields=["name", "block", "block_number_input", "export_block_no"],
        limit_page_length=0,
    )
    total = len(rows)
    had = 0
    filled = 0
    no_key = 0
    no_match = 0
    for r in rows:
        if _s(r.get("export_block_no")):
            had += 1
            continue
        key = r.get("block") or r.get("block_number_input")
        if not key:
            no_key += 1
            continue
        ex = frappe.db.get_value("Quarry Block", key, "export_block_no")
        if not ex:
            ex = frappe.db.get_value("Quarry Block", {"block_number": _s(key)}, "export_block_no")
        if ex:
            frappe.db.set_value("Buyer Inspection Block", r.get("name"),
                                "export_block_no", ex, update_modified=False)
            filled += 1
        else:
            no_match += 1
    frappe.db.commit()
    return {"total_rows": total, "had_before": had, "filled_now": filled,
            "no_block_key": no_key, "no_qb_match": no_match, "has_after": had + filled}


@frappe.whitelist()
def mark_lot_exported(lot=None):
    """Move a lot + its blocks to Exported (status Shipped) once the Bill of Lading
    number is present on the linked Shipping Document. Refuses without a B/L, so
    nothing is marked exported before the port issues the B/L and the invoice is
    printed with it."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)
    sd_name = d.get("shipping_document")
    bl = ""
    if sd_name and frappe.db.exists("Shipping Document", sd_name):
        bl = _s(frappe.db.get_value("Shipping Document", sd_name, "bl_no"))
    if not bl:
        frappe.throw("Enter the Bill of Lading number on the Shipping Document (from the port) before marking the lot as Exported.")
    if d.meta.has_field("status"):
        d.status = "Shipped"
    d.save(ignore_permissions=True)
    tf = _lot_table_field(d)
    n = 0
    for r in (d.get(tf) or []):
        bno = r.get("block")
        if bno and frappe.db.exists("Quarry Block", bno):
            frappe.db.set_value("Quarry Block", bno, "status", "Shipped", update_modified=False)
            n += 1
    frappe.db.commit()
    return {"lot": d.name, "status": "Shipped", "blocks_marked": n, "bl_no": bl}


@frappe.whitelist()
def return_blocks_from_lot(lot=None, blocks=None):
    """Return blocks from a lot back to 'ready for export' (Quarry Block -> At Port)
    and drop them from the lot table + the linked Shipping Document.
      - blocks empty -> return ALL (full unlink undo, also clears the SD link).
      - blocks given -> return only those (partial, like DC / BI)."""
    if not lot:
        frappe.throw("No lot given.")
    import json
    want = set()
    if blocks:
        want = set(str(x) for x in (blocks if isinstance(blocks, list) else json.loads(blocks)))
    d = frappe.get_doc("Export Shipment Lot", lot)
    tf = _lot_table_field(d)
    rows = d.get(tf) or []
    keep, returned = [], []
    for r in rows:
        bno = r.get("block")
        match = (not want) or (str(bno) in want) or (str(r.get("block_no")) in want) or (str(r.get("export_block_no")) in want)
        (returned if match else keep).append(r)
    kept_dicts = [r.as_dict() for r in keep]
    d.set(tf, [])
    for rd in kept_dicts:
        d.append(tf, rd)
    if d.meta.has_field("block_count"):
        d.block_count = len(kept_dicts)
    d.save(ignore_permissions=True)
    ret_ids = []
    for r in returned:
        bno = r.get("block")
        ret_ids.append(bno)
        if bno and frappe.db.exists("Quarry Block", bno):
            frappe.db.set_value("Quarry Block", bno, "status", "At Port", update_modified=False)
    sd_name = d.get("shipping_document")
    if sd_name and frappe.db.exists("Shipping Document", sd_name):
        sd = frappe.get_doc("Shipping Document", sd_name)
        keep_sd = [x.as_dict() for x in (sd.get("blocks") or []) if x.get("block") not in ret_ids]
        sd.set("blocks", [])
        for xd in keep_sd:
            sd.append("blocks", xd)
        if sd.meta.has_field("block_count"):
            sd.block_count = len(keep_sd)
        sd.save(ignore_permissions=True)
        if not keep_sd and not want and d.meta.has_field("shipping_document"):
            d.db_set("shipping_document", None)
    frappe.db.commit()
    return {"lot": d.name, "returned": len(returned), "remaining": len(kept_dicts)}


# ===========================================================================
# Port & Stock — the four capabilities that were missing  (B1, 17 Aug 2026)
#
#   "there is no way to delete duplicates or reject accepting or accept with
#    note and find with note block"
#
#   duplicate_rows()      what is actually duplicated, with both rows shown
#   remove_arrival_row()  soft-delete a duplicate row -> Trash, with a reason
#   reject_acceptance()   undo an accept; the block goes back to needing a person
#   accept_with_note()    accept, but the note travels with the BLOCK
#   find_by_note()        find every block carrying a note, by text
# ===========================================================================


@frappe.whitelist()
def duplicate_rows(arrival=None):
    """Every block number that appears on more than one arrival row, with both
    rows side by side. Previously the page said 617 rows were "Duplicate" and
    offered nothing to do about it."""
    filters = {}
    if arrival:
        filters["parent"] = arrival
    rows = frappe.get_all("Port Arrival Block",
                          filters=filters or None,
                          fields=["name", "parent", "block_no", "length", "width",
                                  "height", "cbm", "net_wt", "recon_status",
                                  "resolution_type", "resolution_note", "vehicle_no"],
                          limit_page_length=0)
    ds = _arrival_docstatus()
    by = {}
    for r in rows:
        k = _s(r.block_no)
        if k:
            by.setdefault(k, []).append(dict(r, confirmed=1 if ds.get(r.parent) == 1 else 0))
    out = []
    for k, group in by.items():
        if len(group) < 2:
            continue
        out.append({
            "block_no": k,
            "count": len(group),
            "rows": sorted(group, key=lambda x: (0 if x["confirmed"] else 1, x["parent"])),
            "identical": len({(_s(g["length"]), _s(g["width"]), _s(g["height"])) for g in group}) == 1,
        })
    out.sort(key=lambda x: (-x["count"], x["block_no"]))
    return out


@frappe.whitelist()
def remove_arrival_row(row=None, reason=None, machine=None, person=None):
    """Soft-delete one arrival row. Nothing is destroyed: the row's contents go
    into the block's Trash stamp and can be restored (B33). The reason is
    mandatory and it follows the BLOCK, so Trace shows it (B34)."""
    from dolphin_theme.lifecycle import remove_to_trash
    if not row:
        frappe.throw("No row given.")
    parent, block_no = frappe.db.get_value("Port Arrival Block", row,
                                           ["parent", "block_no"]) or (None, None)
    if not parent:
        frappe.throw("That row no longer exists.")
    return remove_to_trash(doctype="Port Arrival", parent=parent, row=row,
                           block=block_no, reason=reason, machine=machine,
                           person=person)


@frappe.whitelist()
def reject_acceptance(row=None, block_no=None, arrival=None, reason=None,
                      machine=None, person=None):
    """Undo an acceptance. The row goes back to unresolved and the block comes
    back off At Port, with the rejection written into its history — so an accept
    made in error is a correctable event, not a permanent one."""
    from dolphin_theme.block_resolve import try_resolve, set_status, log_event, machine_of
    reason = _s(reason)
    if len(reason) < 4:
        frappe.throw("Say why the acceptance is being rejected — that reason is the "
                     "only record of it.")
    name = row
    if not name:
        alt = _pab_alt_keys(block_no) or {_s(block_no)}
        flt_ = {"block_no": ["in", list(alt)]}
        if arrival:
            flt_["parent"] = arrival
        name = frappe.db.get_value("Port Arrival Block", flt_, "name")
    if not name:
        frappe.throw("No arrival row found for {0}.".format(_s(block_no)))

    bno = _s(frappe.db.get_value("Port Arrival Block", name, "block_no"))
    updates = {"recon_status": "", "resolution_type": None,
               "resolution_note": "REJECTED: " + reason}
    meta = frappe.get_meta("Port Arrival Block")
    for f, v in (("resolved_by", frappe.session.user), ("resolved_on", now_datetime())):
        if meta.has_field(f):
            updates[f] = v
    frappe.db.set_value("Port Arrival Block", name, updates, update_modified=False)

    hit, _why = try_resolve(bno, allow_record_name=False)
    if hit and _s(hit.get("status")) == "At Port":
        set_status(hit["name"], "Dispatched/Transported",
                   "acceptance rejected: " + reason, machine=machine_of(machine),
                   actor=person, allow_backwards=True)
    if hit:
        log_event(hit["name"], "acceptance-rejected", "Resolved", "", reason,
                  machine_of(machine), person)
    frappe.db.commit()
    return {"ok": True, "row": name, "block_no": bno}


@frappe.whitelist()
def accept_with_note(row=None, block_no=None, arrival=None, note=None,
                     machine=None, person=None, to_status="At Port"):
    """Accept an arrival row WITH a note that stays attached to the block.

    The old accept wrote `resolution_note` on the arrival row, where nobody ever
    looks again. This writes it there AND onto the block, so `find_by_note` and
    Trace can both find it afterwards."""
    from dolphin_theme.block_resolve import try_resolve, set_status, log_event, machine_of
    note = _s(note)
    if len(note) < 4:
        frappe.throw("An acceptance needs a note — that is the point of accepting "
                     "with a note rather than just accepting.")
    name = row
    if not name:
        alt = _pab_alt_keys(block_no) or {_s(block_no)}
        flt_ = {"block_no": ["in", list(alt)]}
        if arrival:
            flt_["parent"] = arrival
        name = frappe.db.get_value("Port Arrival Block", flt_, "name")
    if not name:
        frappe.throw("No arrival row found for {0}.".format(_s(block_no)))

    bno = _s(frappe.db.get_value("Port Arrival Block", name, "block_no"))
    meta = frappe.get_meta("Port Arrival Block")
    updates = {"recon_status": "Resolved", "resolution_note": note}
    if meta.has_field("resolution_type"):
        updates["resolution_type"] = "Accepted as-is"
    if meta.has_field("resolved_by"):
        updates["resolved_by"] = frappe.session.user
    if meta.has_field("resolved_on"):
        updates["resolved_on"] = now_datetime()
    if meta.has_field("resolved_machine"):
        updates["resolved_machine"] = machine_of(machine)
    frappe.db.set_value("Port Arrival Block", name, updates, update_modified=False)

    hit, why = try_resolve(bno, allow_record_name=False)
    if not hit:
        frappe.db.commit()
        return {"ok": True, "row": name, "block_no": bno, "block": None, "why": why,
                "message": "Row accepted, but {0} does not resolve to exactly one "
                           "block, so no block status was changed.".format(bno)}
    try:
        frappe.get_doc("Quarry Block", hit["name"]).add_comment(
            "Comment", "Accepted at port with note · {0} · {1} · {2}".format(
                person or frappe.session.user, machine_of(machine), note))
    except Exception:
        pass
    res = set_status(hit["name"], to_status, "accepted with note: " + note,
                     machine=machine_of(machine), actor=person)
    log_event(hit["name"], "accepted-with-note", None, to_status, note,
              machine_of(machine), person)
    frappe.db.commit()
    return {"ok": True, "row": name, "block_no": bno, "block": hit["name"],
            "status": res.get("status")}


@frappe.whitelist()
def find_by_note(q=None, limit=300):
    """Find every block carrying a note, optionally matching text.

    Looks in three places at once, because notes have been written into all
    three over the months: the arrival row's resolution_note, the block's
    comments, and the lifecycle stamps (trash / skip / reverse)."""
    q = _s(q)
    out, seen = [], set()

    def add(block, where, text, when, who=None):
        k = (str(block), where, (text or "")[:60])
        if k in seen:
            return
        seen.add(k)
        out.append({"block": block, "where": where, "note": text,
                    "at": _s(when)[:19], "by": who})

    like = "%{0}%".format(q) if q else "%"
    try:
        for r in frappe.get_all("Port Arrival Block",
                                filters={"resolution_note": ["like", like]},
                                fields=["block_no", "parent", "resolution_note",
                                        "resolved_on", "resolved_by"],
                                limit_page_length=int(limit or 300)):
            if _s(r.resolution_note):
                add(r.block_no, "arrival " + str(r.parent), r.resolution_note,
                    r.resolved_on, r.resolved_by)
    except Exception:
        pass

    try:
        for c in frappe.get_all("Comment",
                                filters={"reference_doctype": "Quarry Block",
                                         "comment_type": "Comment",
                                         "content": ["like", like]},
                                fields=["reference_name", "content", "creation", "owner"],
                                order_by="creation desc",
                                limit_page_length=int(limit or 300)):
            add(c.reference_name, "block note",
                frappe.utils.strip_html(c.content or "").strip(), c.creation, c.owner)
    except Exception:
        pass

    return out


@frappe.whitelist()
def reconciliation_view():
    """Did the agency transcribe our numbers correctly?  (B30, reframed by B54)

    B54, and this is the whole point of the screen: **the port agency never
    measures a block.** Weight and tonnage are their only concern. The
    measurements are ours and the Buyer Inspection measurement is FINAL.

    So this is not measurement-versus-measurement. It is a transcription check:
    where the agency has typed a size, does it match what we gave them? A row
    where they typed nothing is NORMAL — they were never asked to — and it is
    never counted as a fault.

    Buckets: match · within-tolerance · mismatch (mistyped) · not-entered ·
    not-on-a-challan."""
    disp = _dispatched_index()
    ds = _arrival_docstatus()
    rows, counts = [], {"match": 0, "tol": 0, "mismatch": 0, "nodim": 0, "unknown": 0}

    for p in frappe.get_all("Port Arrival Block", fields=_PAB_FIELDS + ["name"],
                            limit_page_length=0):
        k = _s(p.block_no)
        if not k:
            continue
        d = disp.get(k)
        if not d:
            for alt in _pab_alt_keys(k):
                if alt in disp:
                    d = disp[alt]
                    break
        # Compare a side ONLY when both ends actually carry a number.
        #
        # The first cut of this still called 471 rows a mismatch, because a row
        # with a cbm but no length was treated as "has measurements" and then
        # every side compared against a zero. Zero is not a measurement of zero;
        # it is an absence. Only genuinely comparable sides count, and a row with
        # none of them is 'nothing to compare' rather than a failure.
        pairs = [(d.l, p.length), (d.w, p.width), (d.h, p.height)] if d else []
        comparable = [(a, b) for a, b in pairs if flt(a) and flt(b)]

        if not d:
            bucket = "unknown"
        elif not comparable:
            bucket = "nodim"
        else:
            exact = all(flt(a) == flt(b) for a, b in comparable)
            ok = all(_within_tol(a, b) for a, b in comparable)
            bucket = "match" if exact else ("tol" if ok else "mismatch")
        counts[bucket] += 1
        rows.append({
            "row": p.name, "block_no": k, "arrival": p.parent,
            "confirmed": 1 if ds.get(p.parent) == 1 else 0,
            "dc": (d.dc if d else None),
            "dc_l": (d.l if d else None), "dc_w": (d.w if d else None),
            "dc_h": (d.h if d else None), "dc_cbm": (d.vol if d else None),
            "pt_l": p.length, "pt_w": p.width, "pt_h": p.height, "pt_cbm": p.cbm,
            "net_wt": p.net_wt, "recon_status": p.recon_status,
            "note": p.resolution_note, "bucket": bucket,
            "compared": len(comparable),
        })

    labels = {
        "match": "Agency figures match ours",
        "tol": "Within tolerance (3 cm or 3%)",
        "mismatch": "Mistyped — does not match what we gave",
        "nodim": "Not entered by the agency (normal)",
        "unknown": "Not on any submitted challan",
    }
    labels["_note"] = ("The port agency does not measure — weight and tonnage are their "
                       "only concern, and the Buyer Inspection measurement is final. "
                       "This table only asks whether what they typed matches what we gave "
                       "them. A blank is normal and is never a fault.")
    return {"rows": rows, "counts": counts, "labels": labels,
            "total": len(rows)}


# ---------------------------------------------------------------------------
# 19 Aug 2026 - CORRECTION, appended deliberately so it overrides the earlier
# definitions above (Python keeps the last definition in a module).
#
# _dispatched_index() was keyed on str(r.block) - the Quarry Block DOCNAME, a
# bare integer such as 1363 - while _classify() looks rows up by the ARRIVAL's
# block NUMBER. Those are two different numbering spaces and they overlap on
# this site: 25 docnames collide with an export number, 37 with a block number,
# 103 block numbers collide with an export number.
#
# Block 1363 was therefore matched to DC-DAFG-021 (challan 0021) and to another
# block's dimensions 297x190x59, when export number 1363 belongs to quarry block
# 1508 on DC-GCFG-049 (challan 0121), dimensions 331x138x133. Reported by the
# owner on 19 Aug 2026. Simulated over all 231 arrival numbers: every one of
# them changed verdict, and 25 had been paired with a genuinely wrong challan.
#
# Now keyed on the numbers the arrival actually quotes. A number that appears on
# more than one challan is left OUT of the index entirely, so it is reported as
# unmatched for a human to decide instead of being silently paired with one of
# them. _nearest() is tightened from edit distance 2 (which is guesswork on a
# 3-4 digit number: "021" vs "121" is distance 1) to distance 1 with exactly one
# candidate, and it only ever suggests.
# ---------------------------------------------------------------------------
def _dispatched_index():
    """SUPERSEDES the _dispatched_index earlier in this file (last definition wins)."""
    rows = frappe.db.sql(
        """
        SELECT r.block, r.block_no, r.export_block_no, r.length_gross AS l,
               r.width_gross AS w, r.height_gross AS h, r.gross_volume AS vol, p.name AS dc
        FROM `tabDC Block Row` r
        JOIN `tabDelivery Challan` p ON p.name = r.parent
        WHERE p.docstatus = 1
        """,
        as_dict=True,
    )
    idx, ambiguous = {}, set()
    for r in rows:
        for key in (r.get("block_no"), r.get("export_block_no")):
            k = str(key or "").strip()
            if not k:
                continue
            if k in idx:
                if idx[k].dc != r.dc:
                    ambiguous.add(k)
            else:
                idx[k] = r
    for k in ambiguous:
        idx.pop(k, None)
    return idx


def _nearest(block_no, candidates):
    """SUPERSEDES the _nearest earlier in this file (last definition wins)."""
    target = str(block_no).strip()
    if len(target) < 3:
        return None
    at_one = [c for c in candidates if _edit_distance(target, str(c)) == 1]
    return at_one[0] if len(at_one) == 1 else None


# ---------------------------------------------------------------------------
# 20 Aug 2026 - RECONCILIATION, REBUILT ON HIS TWO RULES.
#
# Appended, not edited: Python keeps the LAST definition, so nothing above is
# touched. Two things were wrong and both came from his own standing notes.
#
#   1. The screen compared the port against `DC Block Row.length/width/height`.
#      His instruction of 19 Aug: "not to consider measurements from DC since at
#      times it will be adjusted due to DMG permit". The Delivery Challan says
#      WHICH BLOCKS travelled and nothing else. The measurement authority is the
#      Buyer Inspection, carried on the Quarry Block master.
#      Measured 20 Aug: against the DC the screen claimed 5 mismatches; against
#      the master, 491 rows match to the centimetre and 0 differ.
#
#   2. The screen listed ROWS, not BLOCKS, and never looked at `recon_status`.
#      The port agency re-sends a cumulative manifest, so 849 rows exist for 231
#      blocks; 617 of them are already flagged Duplicate and none of them ever
#      left the queue. His rule, in his words: "we are supposed to skip the block
#      numbers received already and add only the new ones always".
#
# Identity is resolved in the order he confirmed - Delivery Challan link first
# (a stored record pointer, not a number), then export number, then quarry
# number. A number that resolves to more than one block is NEVER guessed.
# ---------------------------------------------------------------------------

def _qb_master_index():
    """export number / quarry number -> Quarry Block master (BI-final sizes).

    A number claimed by two different blocks is dropped, never guessed."""
    rows = frappe.get_all(
        "Quarry Block",
        fields=["name", "block_number", "export_block_no",
                "length_gross", "width_gross", "height_gross", "gross_volume"],
        limit_page_length=0,
    )
    by_export, by_quarry, by_id, ambiguous = {}, {}, {}, set()
    for q in rows:
        by_id[_s(q.name)] = q
        for bucket, key in ((by_export, q.get("export_block_no")),
                            (by_quarry, q.get("block_number"))):
            k = _s(key)
            if not k:
                continue
            if k in bucket and _s(bucket[k].name) != _s(q.name):
                ambiguous.add((id(bucket), k))
            else:
                bucket[k] = q
    for bucket in (by_export, by_quarry):
        for _b, k in list(ambiguous):
            if _b == id(bucket):
                bucket.pop(k, None)
    return by_export, by_quarry, by_id


def _dc_block_link_index():
    """block number / export number -> the Quarry Block RECORD the submitted
    challan points at. This is the strong witness: a link, not an integer."""
    rows = frappe.db.sql(
        """
        SELECT r.block, r.block_no, r.export_block_no, p.name AS dc
        FROM `tabDC Block Row` r
        JOIN `tabDelivery Challan` p ON p.name = r.parent
        WHERE p.docstatus = 1
        """,
        as_dict=True,
    )
    idx, ambiguous = {}, set()
    for r in rows:
        for key in (r.get("block_no"), r.get("export_block_no")):
            k = _s(key)
            if not k:
                continue
            if k in idx and _s(idx[k]["block"]) != _s(r.get("block")):
                ambiguous.add(k)
            else:
                idx[k] = {"block": r.get("block"), "dc": r.dc}
    for k in ambiguous:
        idx.pop(k, None)
    return idx


@frappe.whitelist()
def reconciliation_view():
    """SUPERSEDES the reconciliation_view earlier in this file (last definition wins).

    One row per BLOCK. Sizes checked against the Buyer Inspection, never the
    Delivery Challan. Rows already flagged Duplicate never reach the screen."""
    by_export, by_quarry, by_id = _qb_master_index()
    dc_link = _dc_block_link_index()
    ds = _arrival_docstatus()

    # newest manifest first, so the keeper carries the port's latest figures
    pabs = frappe.get_all(
        "Port Arrival Block", fields=_PAB_FIELDS + ["name"],   # _PAB_FIELDS already carries recon_status / vehicle_no
        limit_page_length=0, order_by="creation desc",
    )

    grouped, superseded = {}, 0
    for p in pabs:
        k = _s(p.block_no)
        if not k:
            continue
        if _s(p.get("recon_status")) == "Duplicate":
            superseded += 1
            continue
        grouped.setdefault(k, []).append(p)

    rows, counts = [], {"match": 0, "tol": 0, "mismatch": 0, "nodim": 0, "unknown": 0}

    for k, copies in grouped.items():
        p = copies[0]                      # newest surviving copy
        link = dc_link.get(k)
        master = None
        how = None
        if link and _s(link.get("block")) in by_id:
            master, how = by_id[_s(link["block"])], "challan link"
        elif k in by_export:
            master, how = by_export[k], "export number"
        elif k in by_quarry:
            master, how = by_quarry[k], "quarry number"

        pairs = []
        if master:
            pairs = [(master.length_gross, p.length),
                     (master.width_gross, p.width),
                     (master.height_gross, p.height)]
        comparable = [(a, b) for a, b in pairs if flt(a) and flt(b)]

        if not link and not master:
            bucket = "unknown"
        elif not link:
            bucket = "unknown"             # at port, no submitted challan yet
        elif not comparable:
            bucket = "nodim"               # agency recorded weight only - normal
        else:
            exact = all(flt(a) == flt(b) for a, b in comparable)
            ok = all(_within_tol(a, b) for a, b in comparable)
            bucket = "match" if exact else ("tol" if ok else "mismatch")
        counts[bucket] += 1

        weights = sorted({flt(c.net_wt) for c in copies if flt(c.net_wt)})
        rows.append({
            "row": p.name, "block_no": k, "arrival": p.parent,
            "confirmed": 1 if ds.get(p.parent) == 1 else 0,
            "dc": (link.get("dc") if link else None),
            "quarry_block": (master.name if master else None),
            "quarry_no": (master.block_number if master else None),
            "identified_by": how,
            # ours - the Buyer Inspection figure carried on the master
            "ref_l": (master.length_gross if master else None),
            "ref_w": (master.width_gross if master else None),
            "ref_h": (master.height_gross if master else None),
            # kept ONLY so old markup does not break. Never a measurement.
            "dc_l": (master.length_gross if master else None),
            "dc_w": (master.width_gross if master else None),
            "dc_h": (master.height_gross if master else None),
            "dc_cbm": (master.gross_volume if master else None),
            "pt_l": p.length, "pt_w": p.width, "pt_h": p.height, "pt_cbm": p.cbm,
            "net_wt": p.net_wt, "vehicle_no": p.get("vehicle_no"),
            "copies": len(copies),
            "weight_spread_kg": (round((weights[-1] - weights[0]) * 1000) if len(weights) > 1 else 0),
            "weights_seen": weights,
            "recon_status": p.get("recon_status"),
            "note": p.resolution_note, "bucket": bucket,
            "compared": len(comparable),
        })

    rows.sort(key=lambda r: (r["bucket"] != "mismatch", -r["weight_spread_kg"], r["block_no"]))

    labels = {
        "match": "Agency figures match ours",
        "tol": "Within tolerance (3 cm or 3%)",
        "mismatch": "Does not match the Buyer Inspection",
        "nodim": "Not entered by the agency (normal)",
        "unknown": "Not on any submitted challan",
        "_note": ("Sizes are checked against the Buyer Inspection, never the Delivery "
                  "Challan - the challan figure is adjusted for the DMG permit and is "
                  "not a measurement. The port agency does not measure; weight is "
                  "their only concern, so a blank size is normal. One row per block: "
                  "copies from re-sent manifests are folded in, not listed."),
    }
    return {
        "rows": rows, "counts": counts, "labels": labels,
        "blocks": len(rows), "superseded_rows_hidden": superseded,
        "weight_disagreements": len([r for r in rows if r["weight_spread_kg"] > 0]),
    }


# ---------------------------------------------------------------------------
# 20 Aug 2026 - THE ARRIVAL ABSORB. Built on four things he said today:
#
#  1. "we are supposed to skip the block numbers received already and add only
#      the new ones always"
#  2. "search for blocks in the arrivals emails as per our DC rather than other
#      way around to filter earlier older shipped block numbers etc"
#  3. "we cannot afford to make mistakes or else there will be chain reaction
#      and ripple effects"
#  4. "verification method of the blocks and measurement should be multilayered
#      2-3 methods to validate everything to make sure we are not taking into
#      account wrong block, measurements, weights etc"
#
# So nothing here decides anything on ONE piece of evidence. Every row is put
# through three independent witnesses for identity and three for weight, and a
# row is only absorbed when they agree. Where they disagree the row is held,
# never guessed - because guessing a number is what produced the 1363 incident.
#
# Appended: nothing above is touched. Dry run unless apply=1.
#
# IDENTITY - three witnesses
#   A. The submitted Delivery Challan's LINK to the Quarry Block record. A stored
#      pointer, not an integer, so the four overlapping numbering systems cannot
#      confuse it. This is also what reverses the direction he asked for: we look
#      for OUR dispatched blocks in the file, not the file's rows in our data.
#   B. The measurement. Port-typed L/W/H against the Buyer Inspection figure held
#      on the Quarry Block master. An exact three-way match is decisive.
#   C. Uniqueness of the number across all three numbering systems - export
#      number, quarry number, record id. If a bare number resolves to more than
#      one block, this witness abstains rather than voting.
#
# WEIGHT - three witnesses
#   1. The latest manifest figure (the agency's current position; theirs is final).
#   2. Stability across manifests - a figure that settles is a re-weigh, a figure
#      that moves twice is not trusted.
#   3. A sanity band against our own SG-derived tonnage (CBM x 2.6). This never
#      overrides the agency, it only catches a transcription error.
# ---------------------------------------------------------------------------

_SG_FACTOR = 2.6
_SG_BAND = 0.30          # generous - only catches gross errors, never nags


def _dispatched_block_index():
    """What WE say left the quarry: every block on a SUBMITTED Delivery Challan.

    number -> {block, dc}. A number two challans claim for two different blocks
    is dropped, never guessed."""
    rows = frappe.db.sql(
        """
        SELECT r.block, r.block_no, r.export_block_no, p.name AS dc
        FROM `tabDC Block Row` r
        JOIN `tabDelivery Challan` p ON p.name = r.parent
        WHERE p.docstatus = 1
        """,
        as_dict=True,
    )
    idx, ambiguous = {}, set()
    for r in rows:
        for key in (r.get("block_no"), r.get("export_block_no")):
            k = _s(key)
            if not k:
                continue
            if k in idx and _s(idx[k]["block"]) != _s(r.get("block")):
                ambiguous.add(k)
            else:
                idx[k] = {"block": r.get("block"), "dc": r.dc}
    for k in ambiguous:
        idx.pop(k, None)
    return idx


def _master_index():
    """Quarry Block masters, indexed every way a number can be written."""
    rows = frappe.get_all(
        "Quarry Block",
        fields=["name", "block_number", "export_block_no", "status",
                "length_gross", "width_gross", "height_gross", "gross_volume"],
        limit_page_length=0,
    )
    by_id, by_export, by_quarry = {}, {}, {}
    dup_export, dup_quarry = set(), set()
    for q in rows:
        by_id[_s(q.name)] = q
        e, n = _s(q.export_block_no), _s(q.block_number)
        if e:
            if e in by_export and _s(by_export[e].name) != _s(q.name):
                dup_export.add(e)
            else:
                by_export[e] = q
        if n:
            if n in by_quarry and _s(by_quarry[n].name) != _s(q.name):
                dup_quarry.add(n)
            else:
                by_quarry[n] = q
    for k in dup_export:
        by_export.pop(k, None)
    for k in dup_quarry:
        by_quarry.pop(k, None)
    return by_id, by_export, by_quarry


def _identity_witnesses(number, row, disp, by_id, by_export, by_quarry):
    """Three independent opinions on which Quarry Block this line is.

    Returns (verdict, quarry_block, detail). Nothing is ever guessed."""
    votes, detail = {}, {}

    # A - the challan link
    hit = disp.get(number)
    if hit and _s(hit.get("block")):
        votes["challan link"] = _s(hit["block"])
        detail["dc"] = hit.get("dc")

    # B - the measurement, against the Buyer Inspection figure on the master
    pl, pw, ph = flt(row.get("length")), flt(row.get("width")), flt(row.get("height"))
    if pl and pw and ph:
        matches = []
        seen = set()
        for cand in (by_export.get(number), by_quarry.get(number), by_id.get(number)):
            if not cand or _s(cand.name) in seen:
                continue
            seen.add(_s(cand.name))
            if (flt(cand.length_gross) == pl and flt(cand.width_gross) == pw
                    and flt(cand.height_gross) == ph):
                matches.append(_s(cand.name))
        if len(matches) == 1:
            votes["measurement"] = matches[0]
        elif len(matches) > 1:
            detail["measurement"] = "matched more than one block - abstained"
    else:
        detail["measurement"] = "port typed no size - normal, witness abstains"

    # C - uniqueness of the bare number across all three numbering systems
    cands = set()
    for cand in (by_export.get(number), by_quarry.get(number), by_id.get(number)):
        if cand:
            cands.add(_s(cand.name))
    if len(cands) == 1:
        votes["unique number"] = list(cands)[0]
    elif len(cands) > 1:
        detail["number"] = "means %d different blocks - witness abstains" % len(cands)

    if not votes:
        return "no witness", None, detail

    picked = set(votes.values())
    if len(picked) > 1:
        detail["votes"] = dict(votes)
        return "CONFLICT", None, detail

    qb = list(picked)[0]
    detail["agreed_by"] = sorted(votes.keys())
    if not hit:
        # nothing we dispatched carries this number
        return "not expected", qb, detail
    if len(votes) >= 2:
        return "confirmed", qb, detail
    return "single witness", qb, detail


def _weight_verdict(seq, master):
    """Three opinions on the weight. seq is oldest-to-newest, blanks removed."""
    out = {"latest": (seq[-1] if seq else None), "seen": seq}
    if not seq:
        out["verdict"] = "none given"
        return out
    changes = [i for i in range(1, len(seq)) if seq[i] != seq[i - 1]]
    out["changed_times"] = len(changes)
    if master and flt(master.get("gross_volume")):
        est = flt(master["gross_volume"]) * _SG_FACTOR
        out["our_estimate_mt"] = round(est, 3)
        if est:
            off = abs(seq[-1] - est) / est
            out["off_our_estimate_pct"] = round(off * 100, 1)
            if off > _SG_BAND:
                out["verdict"] = "CHECK - far from our own tonnage estimate"
                return out
    if len(changes) > 1:
        out["verdict"] = "CHECK - moved more than once"
        return out
    out["verdict"] = "settled" if changes else "steady"
    return out


@frappe.whitelist()
def absorb_arrivals(apply=0):
    """One row per block, newest weight winning, nothing decided on one witness.

    apply=0 (default) writes nothing and returns exactly what it would do."""
    apply = cint(apply)
    disp = _dispatched_block_index()
    by_id, by_export, by_quarry = _master_index()

    arrivals = frappe.get_all("Port Arrival", filters=[["docstatus", "<", 2]],
                              fields=["name"], order_by="creation asc",
                              limit_page_length=0)
    order = {a["name"]: i for i, a in enumerate(arrivals)}

    rows = frappe.get_all("Port Arrival Block",
                          fields=["name", "parent", "block_no", "net_wt",
                                  "length", "width", "height", "recon_status",
                                  "quarry_block"],
                          limit_page_length=0)
    rows = [r for r in rows if r.parent in order]
    rows.sort(key=lambda r: order.get(r.parent, 0))

    groups, verdicts, held = {}, {}, []
    for r in rows:
        k = _s(r.block_no)
        if not k:
            continue
        v, qb, detail = _identity_witnesses(k, r, disp, by_id, by_export, by_quarry)
        verdicts[r.name] = v
        if v in ("CONFLICT", "no witness"):
            held.append({"row": r.name, "block": k, "arrival": r.parent,
                         "verdict": v, "detail": detail})
            continue
        if v == "not expected":
            held.append({"row": r.name, "block": k, "arrival": r.parent,
                         "verdict": v, "why": "no submitted challan carries this number"})
            continue
        groups.setdefault(qb, []).append((r, k, detail))

    plan = {"keep": [], "repeat": [], "weight_updated": [], "weight_check": [], "linked": []}
    for qb, items in groups.items():
        keeper, kn, kdetail = items[0]
        seq = [flt(x[0].net_wt) for x in items if flt(x[0].net_wt)]
        master = by_id.get(_s(qb))
        wv = _weight_verdict(seq, master)

        plan["keep"].append({"quarry_block": qb, "block": kn, "row": keeper.name,
                             "arrival": keeper.parent, "copies": len(items),
                             "identified_by": kdetail.get("agreed_by"),
                             "dc": kdetail.get("dc"), "weight": wv})
        if str(wv.get("verdict", "")).startswith("CHECK"):
            plan["weight_check"].append({"quarry_block": qb, "block": kn, "weight": wv})

        latest = wv.get("latest")
        if latest is not None and flt(keeper.net_wt) != flt(latest):
            plan["weight_updated"].append({
                "quarry_block": qb, "block": kn, "row": keeper.name,
                "was": flt(keeper.net_wt), "now": flt(latest),
                "change_kg": round(abs(flt(latest) - flt(keeper.net_wt)) * 1000)})
            if apply and not str(wv.get("verdict", "")).startswith("CHECK"):
                frappe.db.set_value("Port Arrival Block", keeper.name,
                                    "net_wt", flt(latest), update_modified=False)

        if apply and _s(keeper.recon_status) != "Matched":
            frappe.db.set_value("Port Arrival Block", keeper.name,
                                "recon_status", "Matched", update_modified=False)
        # 23 Aug 2026: STORE the block this row resolved to. Until today the
        # resolver did all this work and threw the answer away - quarry_block was
        # empty on all 849 rows - so every screen downstream had to match the
        # number as text all over again, and that is where the crossed weights and
        # the wrong-block At Port moves came from. Three witnesses agreed on this
        # block; write it down.
        plan["linked"].append({"row": keeper.name, "block": kn, "quarry_block": qb})
        if apply and _s(keeper.get("quarry_block")) != _s(qb):
            frappe.db.set_value("Port Arrival Block", keeper.name,
                                "quarry_block", qb, update_modified=False)
        for other, on, _d in items[1:]:
            plan["repeat"].append({"row": other.name, "block": on, "quarry_block": qb,
                                   "arrival": other.parent})
            if apply and _s(other.recon_status) != "Duplicate":
                frappe.db.set_value("Port Arrival Block", other.name,
                                    "recon_status", "Duplicate", update_modified=False)
            plan["linked"].append({"row": other.name, "block": on, "quarry_block": qb})
            if apply and _s(other.get("quarry_block")) != _s(qb):
                frappe.db.set_value("Port Arrival Block", other.name,
                                    "quarry_block", qb, update_modified=False)

    if apply:
        for h in held:
            if h["verdict"] == "not expected":
                frappe.db.set_value("Port Arrival Block", h["row"],
                                    "recon_status", "Typo - not in DC", update_modified=False)
        frappe.db.commit()

    tally = {}
    for v in verdicts.values():
        tally[v] = tally.get(v, 0) + 1

    return {
        "applied": bool(apply),
        "arrival_files": len(arrivals),
        "rows_seen": len(rows),
        "blocks_after": len(groups),
        "identity_verdicts": tally,
        "counts": {"keep": len(plan["keep"]), "repeat": len(plan["repeat"]),
                   "held": len(held), "weight_updated": len(plan["weight_updated"]),
                   "weight_needs_check": len(plan["weight_check"]),
                   "rows_linked_to_a_block": len(plan["linked"])},
        "weight_needs_check": plan["weight_check"][:40],
        "weight_updates": sorted(plan["weight_updated"], key=lambda x: -x["change_kg"])[:40],
        "held": held[:60],
        "note": ("Dry run - nothing changed. apply=1 writes. Even with apply=1 a block "
                 "whose weight verdict starts with CHECK is never overwritten, a row is "
                 "never created or deleted, and a CONFLICT is never resolved."),
    }


# ---------------------------------------------------------------------------
# 20 Aug 2026 - UNDO GOES EXACTLY ONE RUNG. His design, his words:
#
#   "my idea is to undo style one step backward so if you want to reopen an
#    exported lot then shouldnt it be export shipment lot draft?"
#   "from export shipment lot one step undo is to return all the blocks to at
#    port isnt it ?"
#
# He is right, and the code did not work that way. The old reopen_lot did THREE
# things in one click: put the lot back to Ready, emptied its block table, and
# returned every block to At Port - so reopening a 56-block lot to correct a BL
# number destroyed the 56-block grouping and it had to be rebuilt by hand. That
# is the chain reaction he is trying to avoid.
#
# The ladder, one rung per click, each landing in the state it was in just before:
#
#   Exported            -> undo -> Shipping Document, unlocked, blocks untouched
#                                  (sd_return_from_exported - already correct)
#   Shipping Document   -> undo -> Export Shipment Lot back to DRAFT, the
#                                  Shipping Document unlinked, THE BLOCKS STAY
#                                  IN THE LOT                    <- reopen_lot
#   Export Shipment Lot -> undo -> every block back to At Port, lot emptied
#                                  (return_blocks_from_lot - already correct)
#
# Appended, so nothing above is edited. reopen_lot below replaces the earlier
# definition and no longer touches a single block.
# ---------------------------------------------------------------------------

@frappe.whitelist()
def reopen_lot(lot=None):
    """One rung back from a shipped/exported lot: the lot returns to draft and
    the Shipping Document is unlinked. Blocks are NOT touched and stay in the
    lot - emptying the lot is the NEXT rung down, return_blocks_from_lot."""
    if not lot:
        frappe.throw("No lot given.")
    d = frappe.get_doc("Export Shipment Lot", lot)

    was = {
        "status": d.get("status"),
        "shipped": d.get("shipped"),
        "shipping_document": d.get("shipping_document"),
    }

    if d.meta.has_field("status"):
        d.status = "Ready"
    if d.meta.has_field("shipped"):
        d.shipped = 0
    for f in ("ship_date", "bl_no"):
        if d.meta.has_field(f):
            d.set(f, None)
    if d.meta.has_field("shipping_document"):
        d.shipping_document = None
    d.flags.ignore_mandatory = True
    d.save(ignore_permissions=True)
    frappe.db.commit()

    kept = len(d.get("blocks") or [])
    return {
        "name": lot,
        "status": "Ready",
        "blocks_kept_in_lot": kept,
        "shipping_document_unlinked": was.get("shipping_document") or None,
        "blocks_returned_to_port": 0,
        "message": ("Lot %s is back to draft with its %d block(s) still in it. "
                    "To send those blocks back to At Port, use the next undo - "
                    "Return all blocks to At Port." % (lot, kept)),
    }


@frappe.whitelist()
def reopen_lot(lot=None, reason=None):
    """SUPERSEDES the reopen_lot above (last definition wins).

    20 Aug 2026. His words: "exported: it went back without any confirmation
    dialog box and reasons names".

    Reversing an export is a reversal on a customs-linked document. It is now
    REFUSED without a reason, and every reversal is written into the timeline of
    both the lot and the Shipping Document it was unlinked from, with who did it
    and when. The rung itself is unchanged: the lot returns to Ready, ship date
    and BL are cleared, the Shipping Document is unlinked, and the blocks STAY
    in the lot.

    The requirement lives here rather than in the buttons on purpose - desk, the
    portal pages and any future caller are all covered by this one gate.
    """
    if not lot:
        frappe.throw("No lot given.")

    why = (reason or "").strip()
    if len(why) < 4:
        frappe.throw(
            "An export cannot be undone without a reason. "
            "Type why this shipment is being reversed - it is written into the "
            "lot's history against your name."
        )

    d = frappe.get_doc("Export Shipment Lot", lot)

    was_sd = d.get("shipping_document")
    was_status = d.get("status")
    was_bl = d.get("bl_no")
    was_date = d.get("ship_date")

    if d.meta.has_field("status"):
        d.status = "Ready"
    if d.meta.has_field("shipped"):
        d.shipped = 0
    for f in ("ship_date", "bl_no"):
        if d.meta.has_field(f):
            d.set(f, None)
    if d.meta.has_field("shipping_document"):
        d.shipping_document = None
    d.flags.ignore_mandatory = True
    d.save(ignore_permissions=True)

    who = frappe.session.user
    when = frappe.utils.now_datetime().strftime("%d-%b-%Y %H:%M")
    kept = len(d.get("blocks") or [])

    note = (
        "<b>Export undone</b> &mdash; %s by %s.<br>"
        "Reason: <b>%s</b><br>"
        "Was: status %s, BL %s, shipment date %s, Shipping Document %s.<br>"
        "The %d block(s) stayed in the lot."
        % (when, who, frappe.utils.escape_html(why),
           was_status or "-", was_bl or "-", was_date or "-", was_sd or "-", kept)
    )

    def _log(dt, dn):
        if not dn:
            return
        try:
            c = frappe.get_doc({
                "doctype": "Comment",
                "comment_type": "Comment",
                "reference_doctype": dt,
                "reference_name": dn,
                "content": note,
                "comment_email": who,
                "comment_by": who,
            })
            c.insert(ignore_permissions=True)
        except Exception:
            frappe.log_error(frappe.get_traceback(), "reopen_lot audit note")

    _log("Export Shipment Lot", lot)
    _log("Shipping Document", was_sd)

    frappe.db.commit()

    return {
        "name": lot,
        "status": "Ready",
        "blocks_kept_in_lot": kept,
        "shipping_document_unlinked": was_sd or None,
        "blocks_returned_to_port": 0,
        "logged_by": who,
        "reason": why,
        "message": ("Lot %s is back to Export Shipment Lot with its %d block(s) still in it. "
                    "Recorded against %s. To send those blocks back to At Port, use the next "
                    "undo - Return all blocks to At Port." % (lot, kept, who)),
    }


# ---------------------------------------------------------------------------
# SHOW THE ARRIVAL SHEET AS THE PORT SENT IT  (21 Aug 2026)
# His words: "this message should show the arrival xls as is as pop up inside",
# and earlier: "if you add as is xls ... it will be best to see there itself to
# understand and either accept or remove, confirm, push to next step".
#
# The attachments are real legacy BIFF .xls (OLE2, magic D0 CF 11 E0), so the
# browser cannot render them and openpyxl cannot read them. xlrd can, and this
# app already uses it in _parse_arrival_xls, so the same reader is reused here.
#
# This returns the grid VERBATIM - every row, every cell, no interpretation,
# no matching, no filtering. It is a viewer, not an importer: it creates
# nothing and changes nothing.
# ---------------------------------------------------------------------------

def _xls_cell_out(sh, r, c, book):
    """One cell, rendered the way a person reading the sheet would expect."""
    try:
        ct = sh.cell_type(r, c)
        v = sh.cell_value(r, c)
    except Exception:
        return ""
    if ct in (0, 6):
        return ""
    if ct == 1:
        return _xls_s(v)
    if ct == 2:
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(round(f, 3))
        except Exception:
            return _xls_s(v)
    if ct == 3:
        try:
            import xlrd
            d = xlrd.xldate.xldate_as_datetime(v, book.datemode)
            return d.strftime("%d-%b-%Y") if (d.hour == 0 and d.minute == 0) else d.strftime("%d-%b-%Y %H:%M")
        except Exception:
            return _xls_s(v)
    if ct == 4:
        return "TRUE" if v else "FALSE"
    if ct == 5:
        return "#ERR"
    return _xls_s(v)


@frappe.whitelist()
def arrival_xls_sheets(file=None, communication=None, arrival=None, max_rows=600, max_cols=30):
    """The arrival .xls exactly as received, every sheet in it. Read-only.

    Renamed 23 Aug 2026. It was called arrival_xls_grid, the same name as the
    viewer above, so Python kept only this one and the Arrivals tab's View sheet
    button - which reads .grid - got a payload with no .grid in it and showed
    'Empty sheet' on files with 200 rows in them.

    file          - File docname or file_url of the attachment
    communication - Communication docname; the first spreadsheet attached to it is used
    arrival       - Port Arrival docname; the first spreadsheet attached to it is used
    """
    if not frappe.has_permission("Port Arrival", "read"):
        frappe.throw("You do not have permission to view arrival sheets.")

    max_rows = int(max_rows or 600)
    max_cols = int(max_cols or 30)

    fdoc = None
    if file:
        if frappe.db.exists("File", file):
            fdoc = frappe.get_doc("File", file)
        else:
            nm = frappe.db.get_value("File", {"file_url": file}, "name")
            if nm:
                fdoc = frappe.get_doc("File", nm)
    if fdoc is None and communication:
        rows = frappe.get_all(
            "File",
            filters={"attached_to_doctype": "Communication", "attached_to_name": communication},
            fields=["name", "file_name"],
            order_by="creation asc",
            limit_page_length=0,
        )
        for r in rows:
            if (r.get("file_name") or "").lower().endswith((".xls", ".xlsx")):
                fdoc = frappe.get_doc("File", r["name"])
                break
    if fdoc is None and arrival:
        rows = frappe.get_all(
            "File",
            filters={"attached_to_doctype": "Port Arrival", "attached_to_name": arrival},
            fields=["name", "file_name"],
            order_by="creation asc",
            limit_page_length=0,
        )
        for r in rows:
            if (r.get("file_name") or "").lower().endswith((".xls", ".xlsx")):
                fdoc = frappe.get_doc("File", r["name"])
                break
    if fdoc is None:
        return {"ok": 0, "error": "No spreadsheet attached to this record."}

    try:
        content = fdoc.get_content()
    except Exception as e:
        return {"ok": 0, "error": "Could not read the attachment: %s" % e}

    try:
        import xlrd
    except Exception:
        return {"ok": 0, "error": "The .xls reader (xlrd) is not installed on this bench."}

    try:
        wb = xlrd.open_workbook(file_contents=content)
    except Exception as e:
        return {"ok": 0, "error": "This file is not a readable .xls: %s" % e}

    single = wb.nsheets == 1
    out_sheets = []
    for sh in wb.sheets():
        used = False
        try:
            used = bool(_xls_is_dolphin_sheet(sh, single))
        except Exception:
            used = False
        nrows = min(sh.nrows, max_rows)
        ncols = min(sh.ncols, max_cols)
        grid = []
        for r in range(nrows):
            line = []
            for c in range(ncols):
                line.append(_xls_cell_out(sh, r, c, wb))
            if any(x != "" for x in line):
                grid.append({"n": r + 1, "c": line})
        out_sheets.append({
            "name": sh.name,
            "rows": grid,
            "nrows": sh.nrows,
            "ncols": sh.ncols,
            "shown_rows": nrows,
            "shown_cols": ncols,
            "truncated": bool(sh.nrows > nrows or sh.ncols > ncols),
            "is_arrival_sheet": used,
        })

    return {
        "ok": 1,
        "file_name": fdoc.file_name,
        "file_url": fdoc.file_url,
        "sheets": out_sheets,
    }


# ---------------------------------------------------------------------------
# EXPORT, IN ONE PIECE  (21 Aug 2026, evening)
#
# What was wrong, and it was proven by running it rather than by reading it:
# `sd_mark_exported` set the lot's status with frappe.db.set_value. A direct database
# write does not fire DocType events, so the After Save cascade that was supposed to move
# the blocks never ran. Tested live on XIAMENBLESS-BL-260727-02 and restored immediately:
#
#     lot     Ready -> Shipped -> Ready
#     blocks  56 Dispatched/Transported -> 56 Dispatched/Transported -> unchanged
#
# So the instant a shipment was marked exported the three records disagreed:
# document Exported, lot Shipped, blocks still Dispatched — and the board went on reading
# 0 Exported after a ship had sailed. This is the same class of fault as the one that lost
# 56 blocks last week, only with the polarity reversed.
#
# These two functions do the whole thing in one place, in one action, in both directions:
#   * the customs fields are compulsory, as they always were
#   * document, lot and every block move together, or nothing moves
#   * each block's previous status is recorded on it, so the way back restores exactly
#     what was there instead of guessing
#   * both directions demand a person's name and write what they did onto the document
# ---------------------------------------------------------------------------

EXPORT_BLOCK_STATUS = "Shipped"


def _lot_block_names(lot_name):
    """Every quarry block on a lot, in order, de-duplicated."""
    rows = frappe.get_all(
        "Shipment Lot Block",
        filters={"parent": lot_name, "parenttype": "Export Shipment Lot"},
        fields=["block"],
        order_by="idx asc",
    )
    out = []
    for r in rows:
        b = _s(r.get("block"))
        if b and b not in out and frappe.db.exists("Quarry Block", b):
            out.append(b)
    return out


def _sd_lot(sd_name):
    lot = _s(frappe.db.get_value("Shipping Document", sd_name, "source_lot"))
    if not lot:
        lot = _s(frappe.db.get_value("Export Shipment Lot", {"shipping_document": sd_name}, "name"))
    return lot


@frappe.whitelist()
def export_shipment(shipping_document=None, person=None, note=None, dry_run=0):
    """Mark a shipment exported: document, lot and every block, together.

    dry_run=1 reports exactly what would move and changes nothing."""
    sd = _s(shipping_document)
    person = _s(person)
    dry = _s(dry_run) in ("1", "true", "True")

    if not sd:
        frappe.throw("No shipping document given.")
    if not frappe.db.exists("Shipping Document", sd):
        frappe.throw("Shipping Document " + sd + " not found.")
    if not person and not dry:
        frappe.throw("Choose your name before marking this exported.")

    row = frappe.db.get_value(
        "Shipping Document", sd,
        ["export_status", "shipping_bill_no", "sb_date", "bl_no", "bl_date"],
        as_dict=True) or {}

    if _s(row.get("export_status")) == "Exported":
        frappe.throw(sd + " is already marked EXPORTED.")

    # the customs fields stay compulsory - this is the rule that has been doing its job
    missing = []
    for fname, label in (("shipping_bill_no", "Shipping Bill No"), ("sb_date", "SB Date")):
        if not _s(row.get(fname)):
            missing.append(label)
    if missing:
        frappe.throw(
            "Cannot mark " + sd + " as exported. These are compulsory on the invoice and "
            "are still empty: " + ", ".join(missing) + ".")

    lot = _sd_lot(sd)
    blocks = _lot_block_names(lot) if lot else []

    before = {}
    for b in blocks:
        before[b] = _s(frappe.db.get_value("Quarry Block", b, "status"))

    plan = {
        "shipping_document": sd,
        "lot": lot,
        "blocks": len(blocks),
        "block_statuses_now": before,
        "would_set_blocks_to": EXPORT_BLOCK_STATUS,
        "dry_run": True,
    }
    if dry:
        return plan

    stamp = frappe.utils.now()
    frappe.db.set_value("Shipping Document", sd, "export_status", "Exported")
    frappe.db.set_value("Shipping Document", sd, "exported_on", stamp)
    frappe.db.set_value("Shipping Document", sd, "exported_by_person", person)

    if lot:
        frappe.db.set_value("Export Shipment Lot", lot, "status", "Shipped")

    moved = 0
    for b in blocks:
        # remembered so the way back restores what was actually there
        frappe.db.set_value("Quarry Block", b, "status_before_export", before.get(b) or "")
        frappe.db.set_value("Quarry Block", b, "status", EXPORT_BLOCK_STATUS)
        moved += 1

    line = ("MARKED AS EXPORTED by " + person + " | login: " + frappe.session.user
            + " | lot: " + (lot or "-") + " | blocks moved to " + EXPORT_BLOCK_STATUS
            + ": " + str(moved))
    if _s(note):
        line += " | note: " + _s(note)
    frappe.get_doc({
        "doctype": "Comment", "comment_type": "Comment",
        "reference_doctype": "Shipping Document", "reference_name": sd,
        "content": line,
    }).insert(ignore_permissions=True)

    frappe.db.commit()
    return {"ok": 1, "status": "Exported", "shipping_document": sd,
            "lot": lot, "blocks_moved": moved}


@frappe.whitelist()
def unexport_shipment(shipping_document=None, person=None, reason=None, dry_run=0):
    """Undo an export: document back to Draft, lot back to Ready, and every block back to
    whatever it read before the export. A reason is compulsory, as it is on every other
    way back in this system."""
    sd = _s(shipping_document)
    person = _s(person)
    reason = _s(reason)
    dry = _s(dry_run) in ("1", "true", "True")

    if not sd:
        frappe.throw("No shipping document given.")
    if not frappe.db.exists("Shipping Document", sd):
        frappe.throw("Shipping Document " + sd + " not found.")
    if not dry:
        if not person:
            frappe.throw("Choose your name before undoing an export.")
        if not reason:
            frappe.throw("Say why this export is being undone. It is written onto the document.")

    cur = _s(frappe.db.get_value("Shipping Document", sd, "export_status"))
    if cur != "Exported":
        frappe.throw(sd + " is not marked exported, so there is nothing to undo.")

    lot = _sd_lot(sd)
    blocks = _lot_block_names(lot) if lot else []

    restore, stuck = {}, []
    for b in blocks:
        prev = _s(frappe.db.get_value("Quarry Block", b, "status_before_export"))
        if prev:
            restore[b] = prev
        else:
            stuck.append(b)

    if dry:
        return {"shipping_document": sd, "lot": lot, "blocks": len(blocks),
                "would_restore": restore, "no_previous_status_recorded": stuck,
                "dry_run": True}

    frappe.db.set_value("Shipping Document", sd, "export_status", "Draft")
    if lot and _s(frappe.db.get_value("Export Shipment Lot", lot, "status")) == "Shipped":
        frappe.db.set_value("Export Shipment Lot", lot, "status", "Ready")

    for b, prev in restore.items():
        frappe.db.set_value("Quarry Block", b, "status", prev)
        frappe.db.set_value("Quarry Block", b, "status_before_export", "")

    line = ("EXPORT UNDONE by " + person + " | login: " + frappe.session.user
            + " | reason: " + reason + " | lot: " + (lot or "-")
            + " | blocks put back: " + str(len(restore)))
    if stuck:
        line += (" | NO PREVIOUS STATUS RECORDED, left untouched: " + ", ".join(stuck[:60]))
    frappe.get_doc({
        "doctype": "Comment", "comment_type": "Comment",
        "reference_doctype": "Shipping Document", "reference_name": sd,
        "content": line,
    }).insert(ignore_permissions=True)

    frappe.db.commit()
    return {"ok": 1, "status": "Draft", "shipping_document": sd, "lot": lot,
            "blocks_put_back": len(restore), "left_untouched": stuck}


# ============================================================================
# AUTO-RECONCILE AND THE WAY BACK — 22 Aug 2026
#
# His words, in order, across one evening:
#   "all matched and verified should move to at port by default"
#   "if there is any issue it should have an option to send it back to
#    reconcilation till resolved good idea?"
#   "auto reconcilation and moving to port is the key it just occured to me"
#   "that is the solution our goal has to be that!"
#
# So: the app settles what it can settle and moves those blocks itself. A person
# is only asked about what the app genuinely cannot decide.
#
# The way back is built FIRST and on purpose. Automatic movement is only safe
# when one click puts a block back and records why — the same rule that makes
# Mark as Exported safe. `status_before_at_port` is written on the way in so the
# way back restores exactly what was there, never a guess.
# ============================================================================

AT_PORT_STATUS = "At Port"
AUTO_TOL_MT = 1.0          # his standing rule: one tonne, and inside it is matched

# A BLOCK CANNOT GROW ON THE WAY TO THE PORT. 23 Aug 2026.
#
# Measured across the 27 Jul sheet, our challan size against the port's, block by
# block: 9 the same, 19 where the port is 70-90% of ours, 28 below 70%, and NOT
# ONE where the port is bigger. A difference that only ever runs one way is not
# measurement error and it is not crossed numbers - both of those scatter in both
# directions. It is material coming off the block: it is dressed and squared at
# the port, and the port measures what is standing there afterwards.
#
# So "Out by 8.33 MT" was the wrong thing to say. Our figure is the rough quarry
# block; theirs is the dressed block. The honest reading:
#
#   port HEAVIER than we dispatched, past the tonne -> impossible, always flag
#   port lighter                                    -> dressing loss, not a fault
#   port very much lighter                          -> worth a look, not an alarm
#
# WHAT I DO NOT KNOW, and he does: how much loss is ordinary for his stone. Until
# he tells me, "very much lighter" is set at less than half, which nobody would
# call ordinary dressing. One number, one place, easy to change.
DRESSING_LOOK_AT_IT = 0.50

# 25 Aug 2026. How far the port's volume may sit from ours before the MATCH
# itself is doubted rather than the weight. His rule is that size neither varies
# nor compresses, so a block at half or double our volume is a different block.
# Deliberately wide, because this must only ever fire on matches that are wrong.
SIZE_MATCH_LOW = 0.55
SIZE_MATCH_HIGH = 1.45


def _prev_field_ready():
    """Make sure Quarry Block has somewhere to remember what it was before At Port.
    Created once, never fails the caller."""
    try:
        if frappe.get_meta("Quarry Block").has_field("status_before_at_port"):
            return True
        from frappe.custom.doctype.custom_field.custom_field import create_custom_field
        create_custom_field("Quarry Block", {
            "fieldname": "status_before_at_port",
            "label": "Status Before At Port",
            "fieldtype": "Data",
            "hidden": 1,
            "read_only": 1,
            "no_copy": 1,
        }, ignore_validate=True)
        frappe.clear_cache(doctype="Quarry Block")
        return frappe.get_meta("Quarry Block").has_field("status_before_at_port")
    except Exception:
        return False


def _size_conflict(r):
    """True only when BOTH sides gave a dimension and they disagree. A missing
    figure is a missing figure, never a conflict."""
    for ours, theirs in (("dc_l", "pt_l"), ("dc_w", "pt_w"), ("dc_h", "pt_h")):
        a, b = r.get(ours), r.get(theirs)
        if a in (None, "", 0) or b in (None, "", 0):
            continue
        try:
            if abs(float(a) - float(b)) > 0.5:
                return True
        except Exception:
            continue
    return False


def _challans_out_of_tolerance(rows):
    """Which challans are genuinely out by more than a tonne?

    22 Aug 2026, his question: "does the blocks weight ton difference with the port
    and ours goes auto to at port? it must go, only lesser or higher than 1 ton
    should be waiting for user decision is it correct?"

    Yes - with one correction, which is his own earlier rule. A PER-BLOCK weight
    from the agency is their arbitrary split of one truck weighing, so a per-block
    gap means nothing and must never hold a block up. The tonne is tested where the
    number is real: CHALLAN TOTAL against CHALLAN TOTAL, and only when they have
    sent every row of that challan. An incomplete challan gets no verdict at all.
    """
    tot = {}
    for r in rows or []:
        dc = _s(r.get("dc"))
        if not dc:
            continue
        t = tot.setdefault(dc, {"ours": 0.0, "theirs": 0.0, "rows": 0, "with_port": 0})
        t["rows"] += 1
        try:
            t["ours"] += float(r.get("ton") or 0)
        except Exception:
            pass
        nw = r.get("net_wt")
        if nw:
            t["with_port"] += 1
            try:
                t["theirs"] += float(nw)
            except Exception:
                pass
    bad = set()
    for dc, t in tot.items():
        if not t["with_port"] or t["with_port"] != t["rows"]:
            continue          # incomplete: a missing row is a missing row, not a gap
        if not t["ours"] or not t["theirs"]:
            continue
        if abs(t["theirs"] - t["ours"]) > AUTO_TOL_MT:
            bad.add(dc)
    return bad


def _classify_for_auto(rows):
    """Split the ledger into what the app may settle by itself and what it may not.

    verified   the agency sent a row for this block and its challan total is inside
               the one tonne tolerance. A size difference does NOT stop it - it is
               marked size_flag=1 and goes through
    noconflict the block is on a challan of ours, nothing contradicts it, but the
               agency has not sent a row - reconciliation, per his rule that blocks
               with no arrival details wait there
    conflict   the challan total is out by more than a tonne, or there is no challan
               at all - a person decides, always
    settled    already At Port, in a lot, or loaded

    23 Aug 2026: size is no longer a holding reason. Only three things hold a block
    now - a duplicate, no arrival row, and a challan more than a tonne out.
    """
    out = {"verified": [], "noconflict": [], "conflict": [], "settled": []}
    bad_dc = _challans_out_of_tolerance(rows)
    for r in rows or []:
        st = _s(r.get("state"))
        if st in ("port", "lot", "load"):
            out["settled"].append(r)
            continue
        # 23 Aug 2026. A size difference used to hold the block here. His rule:
        #   "now that measurement is not port concern if there is variation in the
        #    port measurement it is actually just a typo error since they dont
        #    measure at all you can highlight but is should go to at port"
        # So it is marked and carried, never a reason to stop the block. The mark
        # travels on the row so the screen can highlight it at the port.
        r["size_flag"] = 1 if _size_conflict(r) else 0
        if _s(r.get("dc")) in bad_dc:
            out["conflict"].append(r)
            continue
        got_port = any(r.get(k) for k in ("pt_l", "pt_w", "pt_h", "pt_cbm", "net_wt"))
        if got_port and r.get("arrival"):
            out["verified"].append(r)
        elif r.get("dc"):
            out["noconflict"].append(r)
        else:
            out["conflict"].append(r)
    return out


@frappe.whitelist()
def auto_settle_preview():
    """What would the app settle by itself right now? Changes nothing."""
    rows = ledger_view() or []
    if isinstance(rows, dict):
        rows = rows.get("rows") or []
    g = _classify_for_auto(rows)

    def names(k, cap=400):
        return [_s(x.get("export_block_no") or x.get("block_no")) for x in g[k][:cap]]

    return {
        "verified": len(g["verified"]),
        "noconflict": len(g["noconflict"]),
        "conflict": len(g["conflict"]),
        "settled": len(g["settled"]),
        "total": len(rows),
        "verified_blocks": names("verified"),
        "noconflict_blocks": names("noconflict"),
        "conflict_blocks": names("conflict"),
        "tolerance_mt": AUTO_TOL_MT,
    }


# ============================================================================
# THE AT PORT / RECONCILIATION SPLIT - 24 Aug 2026
#
# His rules, in his words, and the whole of the policy this endpoint carries:
#
#   "Port will not just go on making mistakes ... there will be some small
#    mistakes or typo mistakes and weight will vary since there method of
#    weighing is different"
#   "without Dc and arrivals dont jump to conclusions this is my strict
#    instruction"
#
# So the line is drawn once, here, and the screen only draws what it is told:
#
#   RECONCILIATION is where a PERSON decides. Exactly three things put a block
#   there - a duplicate, no arrival row, and a challan more than a tonne out
#   (which includes having no challan at all).
#
#   AT PORT holds only stone that has been decided. Nothing drifts into it.
#
# A SIZE DIFFERENCE IS NOT ONE OF THE THREE. It travels with the block as a
# note and the block passes. Weight variation inside the tonne is not a
# difference at all - two weighing methods, and theirs is the final one.
#
# One block appears in exactly ONE group, and the group IS the reason it is
# held. No block is in two places, and no block is held for a reason the screen
# does not name.
# ============================================================================


def _cbm_disagrees(l, w, h, cbm):
    """Does a stated size multiply out to the CBM printed beside it?

    His ask: "under reconcilation you have to highlight this major difference
    and ask user to choose which is correct". This answers only the arithmetic
    question. It never decides which side is right - a person does that.
    """
    try:
        l, w, h, cbm = float(l or 0), float(w or 0), float(h or 0), float(cbm or 0)
    except Exception:
        return None
    if not (l and w and h and cbm):
        return None
    calc = (l * w * h) / 1000000.0
    if calc <= 0:
        return None
    # 5% or 0.15 CBM, whichever is larger - below that it is rounding, not a
    # difference worth stopping a person for.
    if abs(calc - cbm) <= max(0.15, calc * 0.05):
        return None
    return {"stated": round(cbm, 3), "from_size": round(calc, 3),
            "gap": round(cbm - calc, 3)}


def _worklist_row(r):
    """The one shape the Reconciliation screen reads. Nothing is computed twice
    in Javascript that the classifier has already decided here."""
    ours = float(r.get("ton") or 0)
    theirs = float(r.get("net_wt") or 0)
    out = {
        "block_no": _s(r.get("export_block_no") or r.get("block_no")),
        "quarry_block_no": _s(r.get("quarry_block_no")),
        "dc": _s(r.get("dc")),
        "arrival": _s(r.get("arrival")),
        "arrival_confirmed": 1 if r.get("arrival_confirmed") else 0,
        "state": _s(r.get("state")),
        "lot": _s(r.get("lot")),
        "dc_l": r.get("dc_l"), "dc_w": r.get("dc_w"), "dc_h": r.get("dc_h"),
        "dc_cbm": r.get("dc_cbm"), "ours_mt": ours or None,
        "pt_l": r.get("pt_l"), "pt_w": r.get("pt_w"), "pt_h": r.get("pt_h"),
        "pt_cbm": r.get("pt_cbm"), "theirs_mt": theirs or None,
        "size_flag": 1 if r.get("size_flag") else 0,
    }
    # Which reading does not multiply out - ours, theirs, or neither. Both are
    # reported: the screen asks which is correct, it does not guess.
    ours_cbm = _cbm_disagrees(r.get("dc_l"), r.get("dc_w"), r.get("dc_h"), r.get("dc_cbm"))
    port_cbm = _cbm_disagrees(r.get("pt_l"), r.get("pt_w"), r.get("pt_h"), r.get("pt_cbm"))
    if ours_cbm:
        out["cbm_ours"] = ours_cbm
    if port_cbm:
        out["cbm_port"] = port_cbm
    out["cbm_flag"] = 1 if (ours_cbm or port_cbm) else 0
    return out


@frappe.whitelist()
def reconcile_worklist():
    """Reconciliation as a worklist: every held block under the reason it is held.

    Groups, and nothing else is a group:
      no_arrival   on a submitted challan of ours, nothing contradicts it, but
                   the agency has sent no arrival row. It waits.
      over_tonne   its challan total is out by more than one tonne. Grouped by
                   CHALLAN, because that is the level the tolerance is measured
                   at, with the direction shown - a port figure HEAVIER than we
                   dispatched cannot happen, a lighter one is dressing loss.
      no_challan   at the port on no submitted challan at all.
      duplicate    the same block number on more than one arrival row.

    `ready` is what the app would settle by itself. It is reported so the screen
    can say so, and it is NOT part of the worklist - nobody is asked about it.
    """
    rows = ledger_view() or []
    if isinstance(rows, dict):
        rows = rows.get("rows") or []
    g = _classify_for_auto(rows)
    bad_dc = _challans_out_of_tolerance(rows)

    no_arrival, over_tonne, no_challan = [], [], []
    for r in g["conflict"]:
        (over_tonne if _s(r.get("dc")) in bad_dc else no_challan).append(_worklist_row(r))
    for r in g["noconflict"]:
        no_arrival.append(_worklist_row(r))

    # The tonne is a CHALLAN measurement, so the screen shows the challan.
    by_dc = {}
    for r in over_tonne:
        by_dc.setdefault(r["dc"], {"dc": r["dc"], "blocks": [],
                                   "ours_mt": 0.0, "theirs_mt": 0.0})
        d = by_dc[r["dc"]]
        d["blocks"].append(r)
        d["ours_mt"] += float(r.get("ours_mt") or 0)
        d["theirs_mt"] += float(r.get("theirs_mt") or 0)
    challans = []
    for d in by_dc.values():
        gap = d["theirs_mt"] - d["ours_mt"]
        d["ours_mt"] = round(d["ours_mt"], 3)
        d["theirs_mt"] = round(d["theirs_mt"], 3)
        d["gap_mt"] = round(gap, 3)
        # Their figure heavier than what we dispatched cannot be dressing loss.
        # Their figure lighter is ordinary until it is very much lighter.
        if gap > 0:
            d["kind"] = "heavier"
        elif d["ours_mt"] and (d["theirs_mt"] / d["ours_mt"]) < DRESSING_LOOK_AT_IT:
            d["kind"] = "much_lighter"
        else:
            d["kind"] = "lighter"
        challans.append(d)
    challans.sort(key=lambda x: -abs(x["gap_mt"]))

    try:
        dups = duplicate_rows() or []
    except Exception:
        dups = []

    # The size note follows the BLOCK, wherever it has got to - a block that
    # reached At Port yesterday still carries it. The classifier only stamps
    # size_flag on rows it is deciding, so the note is read straight off the
    # ledger here instead.
    size_notes = [_s(r.get("export_block_no") or r.get("block_no"))
                  for r in rows if _size_conflict(r)]
    size_notes = [b for b in size_notes if b]

    return {
        "tolerance_mt": AUTO_TOL_MT,
        "ready": len(g["verified"]),
        # 24 Aug 2026. The numbers of the ready blocks, not just how many. A
        # ready block is nobody's question - the app will settle it - but it
        # still has a HOME, and anything checking that every block has one has
        # to be able to see it. Without this a ready block looks homeless.
        "ready_blocks": [_s(r.get("export_block_no") or r.get("block_no"))
                         for r in g["verified"]][:400],
        "settled": len(g["settled"]),
        "total": len(rows),
        "held": len(no_arrival) + len(over_tonne) + len(no_challan) + len(dups),
        "groups": {
            "no_arrival": {"count": len(no_arrival), "blocks": no_arrival[:400]},
            "over_tonne": {"count": len(over_tonne), "challans": challans},
            "no_challan": {"count": len(no_challan), "blocks": no_challan[:200]},
            "duplicate": {"count": len(dups), "blocks": dups[:200]},
        },
        # Carried, never held: the note that travels to At Port with the block.
        # The screen marks these rows quietly; it never stops one.
        "size_notes": len(size_notes),
        "size_note_blocks": size_notes[:1000],
        "cbm_questions": [r for r in (no_arrival + over_tonne + no_challan)
                          if r.get("cbm_flag")][:200],
    }


@frappe.whitelist()
def auto_settle_at_port(include_noconflict=0, person=None, dry_run=0, note=None):
    """Move everything the app can settle to At Port, by itself.

    include_noconflict=0  only blocks the agency has confirmed
    include_noconflict=1  also blocks with nothing contradicting them, where the
                          agency simply never sent a row
    Every block carries the reason, and every block remembers what it was, so
    send_back_to_reconcile can put it back exactly.
    """
    # ========================================================================
    # 22 Aug 2026. His words:
    #   "this is too dangerous when we automate to match.. even 1 block is error
    #    we had to pay huge penalty in Lakhs of rupees and dollars"
    #   "so rather it should be strong and foolproof with no room for errors"
    #
    # So the automatic path is now the NARROW one on purpose: only blocks the
    # agency's own sheet confirms. Everything else is a person's decision with a
    # reason - which is what he already asked for, and what his team is best
    # placed to make. include_noconflict is kept for that deliberate path but it
    # now demands a named person and a written reason; it can no longer be a
    # casual click.
    # ========================================================================
    dry = _s(dry_run) in ("1", "true", "True")
    inc = _s(include_noconflict) in ("1", "true", "True")
    person = _s(person) or frappe.session.user
    if inc and not dry:
        if not _s(person) or _s(person) == "Administrator":
            frappe.throw("Moving blocks the agency has not confirmed needs a named person.")
        if len(_s(note or "")) < 4:
            frappe.throw(
                "Moving blocks the agency has not confirmed needs a written reason. "
                "Only the agency-confirmed blocks move without one.")

    rows = ledger_view() or []
    if isinstance(rows, dict):
        rows = rows.get("rows") or []
    g = _classify_for_auto(rows)

    todo = list(g["verified"]) + (list(g["noconflict"]) if inc else [])
    plan = {
        "would_move": len(todo),
        "verified": len(g["verified"]),
        "noconflict": len(g["noconflict"]),
        "left_for_a_person": len(g["conflict"]),
        "include_noconflict": 1 if inc else 0,
    }
    if dry:
        plan["blocks"] = [_s(x.get("export_block_no") or x.get("block_no")) for x in todo[:400]]
        plan["dry_run"] = True
        return plan

    _prev_field_ready()
    from dolphin_theme.block_resolve import try_resolve, set_status

    moved, refused, already = [], [], []
    for r in todo:
        bn = _s(r.get("export_block_no") or r.get("block_no"))
        if not bn:
            continue
        # ==================================================================
        # 25 Aug 2026. THE NUMBER WAS NEVER THE IDENTITY. THE RECORD IS.
        #
        # His rule that the tolerance should settle a block by itself could
        # not work, and this is why. Measured live today, all 16 blocks the
        # app wanted to settle were refused as "ambiguous - No block answers
        # to 1150." They were right to be refused on the NUMBER: 1150 is
        # block 1001564's EXPORT number and, at the same time, block
        # 1001474's QUARRY number. Every one of the 16 collides that way.
        # 1364 is one block's export number and another's quarry number, and
        # THAT block's export number is 1160, which is a third block's
        # quarry number. A chain of them.
        #
        # But the ledger row already knows exactly which record it is - `qb`
        # comes off the challan's own link to the Quarry Block. Throwing that
        # away and re-resolving a bare number is how a screen that knows the
        # answer manages to ask an ambiguous question.
        #
        # So: use the record when we have it, and demand only that the record
        # ANSWERS TO the number we are about to report. Fall back to
        # resolving the number when there is no record, where the old guards
        # all still apply. The refusals below are unchanged in strength - a
        # wrong match here costs lakhs - they are simply no longer applied to
        # a question we did not need to ask.
        # ==================================================================
        hit, why = None, ""
        qbid = _s(r.get("qb"))
        if qbid:
            try:
                _rec = frappe.db.get_value(
                    "Quarry Block", qbid,
                    ["name", "block_number", "export_block_no"], as_dict=True)
            except Exception:
                _rec = None
            if _rec and _s(bn) in {_s(_rec.get("export_block_no")),
                                   _s(_rec.get("block_number"))}:
                hit = {"name": _s(_rec.get("name")),
                       "block_number": _rec.get("block_number"),
                       "export_block_no": _rec.get("export_block_no")}
            elif _rec:
                refused.append({"block": bn,
                                "why": "the challan points at a block that does not "
                                       "answer to this number",
                                "detail": "record " + qbid})
                continue

        if not hit:
            # FOOLPROOF IDENTITY, the number-only path. Four separate refusals,
            # any one of which stops the block:
            #   1. the number must resolve, and never through a record id
            #   2. the resolution must be unambiguous (try_resolve enforces that)
            #   3. the block it landed on must itself answer to the number sent
            #   4. the number must not be some OTHER block's record id - on this
            #      site 62 numbers are, and those are precisely the ones that
            #      produce a confident, silent, wrong match (identity_guard.py)
            try:
                from dolphin_theme.identity_guard import check_number
                _chk = check_number(bn)
                if not _chk.get("ok"):
                    refused.append({"block": bn, "why": _chk.get("reason") or "refused",
                                    "detail": _chk.get("message")})
                    continue
            except Exception:
                pass
            hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            refused.append({"block": bn, "why": why})
            continue
        name = hit["name"]
        if _s(name) == _s(bn):
            refused.append({"block": bn, "why": "that is a record id, not a block number"})
            continue
        _answers = {_s(hit.get("export_block_no")), _s(hit.get("block_number"))}
        if _s(bn) not in _answers:
            refused.append({"block": bn,
                            "why": "resolved to a block that does not answer to this number"})
            continue
        cur = _s(frappe.db.get_value("Quarry Block", name, "status"))
        if cur == AT_PORT_STATUS:
            already.append(bn)
            continue
        confirmed = r in g["verified"]
        reason = ("auto-reconciled: " +
                  ("the agency's sheet agrees with ours"
                   if confirmed else
                   "nothing contradicts this block; the agency never sent a row") +
                  " (tolerance " + str(AUTO_TOL_MT) + " MT) - settled without a person by " + person)
        try:
            frappe.db.set_value("Quarry Block", name, "status_before_at_port", cur or "")
        except Exception:
            pass
        res = set_status(name, AT_PORT_STATUS, reason, machine="server (auto-reconcile)",
                         actor=person)
        if res.get("ok"):
            moved.append(bn)
        else:
            refused.append({"block": bn, "why": res.get("error") or "refused"})

    frappe.db.commit()
    return {"ok": 1, "moved": len(moved), "already_at_port": len(already),
            "refused": refused, "left_for_a_person": len(g["conflict"]),
            "moved_blocks": moved[:400]}


@frappe.whitelist()
def send_back_to_reconcile(blocks=None, reason=None, person=None, dry_run=0):
    """The way back from At Port. His words: "if there is any issue it should have
    an option to send it back to reconcilation till resolved".

    Restores the exact status the block held before it went to At Port, refuses
    without a reason, and writes who and why onto the block's own history."""
    if isinstance(blocks, str):
        blocks = _json.loads(blocks)
    blocks = blocks or []
    dry = _s(dry_run) in ("1", "true", "True")
    reason = _s(reason)
    person = _s(person) or frappe.session.user

    if not blocks:
        frappe.throw("No blocks given to send back.")
    if not dry and not reason:
        frappe.throw("Say why this block is going back to Reconcile. "
                     "Every way back in this system asks, and this one is no different.")

    from dolphin_theme.block_resolve import try_resolve, set_status

    plan, done, refused = [], [], []
    for it in blocks:
        bn = _s(it.get("block_no") if isinstance(it, dict) else it)
        if not bn:
            continue
        hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            refused.append({"block": bn, "why": why})
            continue
        name = hit["name"]
        cur = _s(frappe.db.get_value("Quarry Block", name, "status"))
        if cur != AT_PORT_STATUS:
            refused.append({"block": bn, "why": "not at port (reads " + (cur or "blank") + ")"})
            continue
        prev = ""
        try:
            prev = _s(frappe.db.get_value("Quarry Block", name, "status_before_at_port"))
        except Exception:
            prev = ""
        if not prev:
            # never guess. Dispatched/Transported is where a block sits between a
            # submitted challan and the port, and it is the only safe default.
            prev = "Dispatched/Transported"
        plan.append({"block": bn, "from": cur, "back_to": prev})
        if dry:
            continue
        res = set_status(name, prev,
                         "sent back to Reconcile: " + reason + " - by " + person,
                         machine="server (send back to reconcile)",
                         allow_backwards=True, actor=person)
        if res.get("ok"):
            try:
                frappe.db.set_value("Quarry Block", name, "status_before_at_port", "")
            except Exception:
                pass
            done.append(bn)
        else:
            refused.append({"block": bn, "why": res.get("error") or "refused"})

    if dry:
        return {"dry_run": True, "plan": plan, "refused": refused}
    frappe.db.commit()
    return {"ok": 1, "sent_back": len(done), "blocks": done, "refused": refused}


# ============================================================================
# THE CHALLAN COMPARISON, DONE HONESTLY — 22 Aug 2026
#
# His words:
#   "shows ask the agency missing meaning it is not reconciled right? usually
#    that is not the case a Dc cannot be partially received please cross check if
#    wrong numbers I mean internal id's etc are checking vs the port export
#    numbers? because practically once a truck unloads all the blocks in the Dc
#    should be there it cannot miss so easily since each block weigh in tons"
#
# He was right. Measured before writing this: of the 61 challans in the ledger,
# **every block of every one of them is on an arrival sheet — 0 partially
# received**. Challan DC-DAEG-004 was reported as "2 rows missing"; its three
# blocks 1387 / 1353 / 1365 are all on ARR-27Jul2026-NA with weights 8.37, 7.28
# and 7.29. The agency had sent the whole truck.
#
# The old `dc_weight_check` matched raw numbers, so it found the rows keyed one
# way and missed the rows keyed the other — the agency's sheet carries the
# Quarry Block record id while the challan carries the export number. That is the
# same identity fault that made Move to At Port silently do nothing.
#
# This version resolves BOTH sides through the one resolver that refuses to
# guess, so a row is matched by the block it means, not by the digits it happens
# to carry. A truck is still weighed as a truck: total against total, 1 tonne.
# ============================================================================


def _arrival_rows_by_block():
    """Every arrival row, keyed by the Quarry Block it actually refers to.

    Keyed by resolved block name, so a sheet written in record ids and a challan
    written in export numbers land on the same key."""
    # 22 Aug 2026, evening. He asked "is this correct?" of a screen showing
    # "Out by 8.33 MT" on a challan whose every block read "not on their sheet".
    # It was not. Two faults, both here:
    #   * this pulled rows from DRAFT arrival sheets and treated them as a
    #     confirmed weighing. ledger_view has refused to do that since 17 Aug;
    #     the comparison did not, so the two screens contradicted each other.
    #   * a block can sit on several sheets (198 duplicates). Taking "the first
    #     row found" can assemble one challan's agency total out of rows from
    #     different trucks on different days. That produced -34.82 MT on a
    #     two-block truck.
    # Both are fixed: docstatus travels with every row, and the caller may only
    # give a verdict when every row came from ONE sheet.
    _ds = _arrival_docstatus()
    rows = frappe.get_all(
        "Port Arrival Block",
        fields=["name", "parent", "block_no", "length", "width", "height", "cbm", "net_wt"],
        limit_page_length=0,
    )
    for _r in rows:
        _r["submitted"] = 1 if _ds.get(_r.get("parent")) == 1 else 0
    from dolphin_theme.block_resolve import try_resolve

    seen, out, id_only = {}, {}, set()
    for r in rows:
        key = _s(r.get("block_no"))
        if not key:
            continue
        if key not in seen:
            # ================================================================
            # 22 Aug 2026 - ORDER MATTERS, AND GETTING IT WRONG PICKS THE WRONG
            # BLOCK. Proved on live data:
            #
            #   "1353" as a BLOCK NUMBER  -> Quarry Block 1865
            #   "1353" as a RECORD ID     -> a different block (export 803)
            #
            # Same digits, two different blocks. So a number is ALWAYS read as a
            # block number first. Only when it means nothing as a block number -
            # as "1387" does, which is a record id and nothing else - is it read
            # as a record id, and then it is marked so the caller knows.
            # ================================================================
            hit, why = try_resolve(key, allow_record_name=False)
            via_id = False
            if not hit:
                hit, why = try_resolve(key, allow_record_name=True)
                via_id = bool(hit)
            seen[key] = (hit or {}).get("name")
            if via_id and hit:
                id_only.add(_s(hit.get("name")))
        name = seen[key]
        if not name:
            continue
        out.setdefault(name, []).append(r)
    # id_only is not returned today, but it is the list worth watching: those are
    # the arrival rows that only made sense as a record id.
    return out


def _arrival_rows_by_number():
    """Every arrival row, keyed by THE NUMBER THE AGENCY ACTUALLY WROTE.

    23 Aug 2026. His words, and he was right: "usually that is not the case a Dc
    cannot be partially received ... practically once a truck unloads all the
    blocks in the Dc should be there it cannot miss so easily since each block
    weigh in tons".

    Measured on the live site, of the 110 blocks the screen was reporting as
    never sent by the agency:

        81  were on the agency's sheet, written EXACTLY as the challan writes them
         6  were on it under the block's other number
        23  were genuinely not there

    So 87 of 110 "ask the agency" messages were our matching, not their paperwork -
    exactly what he said. The cause: matching went challan number -> block RECORD
    -> arrival rows for that record. Two lookups, and either one failing loses a
    row whose digits sit in plain sight on both sheets.

    A person does not do that. A person reads the number. So does this now, and
    the record is used only when the plain number finds nothing.

    THE REASON THIS IS SAFE TODAY AND WOULD NOT HAVE BEEN YESTERDAY: until the
    seven-digit rename, 62 block numbers were also some other block's record id,
    so "the number as written" could mean two blocks. It cannot any more. And a
    number that is STILL claimed by more than one block master is left out of this
    index entirely rather than guessed - his rule: "even 1 block is error we had to
    pay huge penalty in Lakhs of rupees and dollars".
    """
    _ds = _arrival_docstatus()
    rows = frappe.get_all(
        "Port Arrival Block",
        fields=["name", "parent", "block_no", "length", "width", "height", "cbm", "net_wt"],
        limit_page_length=0,
    )
    # which numbers are unambiguous across every block master
    owners = {}
    for b in frappe.get_all("Quarry Block",
                            fields=["name", "block_number", "export_block_no"],
                            limit_page_length=0):
        for f in ("block_number", "export_block_no"):
            v = _s(b.get(f))
            if v:
                owners.setdefault(v, set()).add(_s(b.name))
    out = {}
    for r in rows:
        key = _s(r.get("block_no"))
        if not key:
            continue
        if len(owners.get(key, set())) > 1:
            continue                      # two blocks answer to it - never guess
        r["submitted"] = 1 if _ds.get(r.get("parent")) == 1 else 0
        out.setdefault(key, []).append(r)
    return out


def _sheet_number_system():
    """Which number each agency sheet is written in, decided by the sheet itself.

    23 Aug 2026. After the first fix, 69 blocks were still unmatched - not because
    the digits were missing but because 103 numbers on this site are BOTH one
    block's export number and a different block's quarry number. The matcher
    refused to guess between them, which was right, but it left a third of the
    board grey.

    Then the sheets answered the question themselves. Counted row by row:

        ARR-27Jul2026-NA     56 rows   56 export numbers   0 quarry numbers
        ARR-30Jul2026-NA    200 rows  200 export numbers   0 quarry numbers
        ARR-30Jul2026-NA-2  193 rows  193 export numbers   0 quarry numbers
        ARR-05Aug2026-NA    200 rows  200 export numbers   0 quarry numbers
        ARR-13Aug2026-NA    200 rows  200 export numbers   0 quarry numbers

    Not one row on any sheet is a quarry-only number. The agency writes export
    numbers, because export numbers are what we give them. So on a sheet that
    reads that way, a number IS an export number and the block it belongs to is
    not in doubt any more.

    This is measured every time rather than assumed. A sheet only gets read as
    export numbers when at least four in five of its rows are export numbers and
    none of them is a quarry-only number. Any sheet that does not pass is left
    exactly as strict as before - his rule stands: "even 1 block is error we had
    to pay huge penalty in Lakhs of rupees and dollars".
    """
    exp, qry = {}, {}
    for b in frappe.get_all("Quarry Block",
                            fields=["name", "block_number", "export_block_no"],
                            limit_page_length=0):
        e, q = _s(b.get("export_block_no")), _s(b.get("block_number"))
        if e:
            exp.setdefault(e, set()).add(_s(b.name))
        if q:
            qry.setdefault(q, set()).add(_s(b.name))

    tally = {}
    for r in frappe.get_all("Port Arrival Block",
                            fields=["parent", "block_no"], limit_page_length=0):
        k = _s(r.get("block_no"))
        if not k:
            continue
        t = tally.setdefault(_s(r.get("parent")), {"rows": 0, "exp": 0, "qry_only": 0})
        t["rows"] += 1
        if k in exp:
            t["exp"] += 1
        elif k in qry:
            t["qry_only"] += 1

    system = {}
    for sheet, t in tally.items():
        if t["rows"] and not t["qry_only"] and t["exp"] >= 0.8 * t["rows"]:
            system[sheet] = "export"
        else:
            system[sheet] = "unclear"
    return system, exp, qry


def _other_number_index():
    """quarry number -> export number and back, so a challan written one way finds
    a sheet written the other way. Only where the pairing is unambiguous."""
    q2e, e2q, dupq, dupe = {}, {}, set(), set()
    for b in frappe.get_all("Quarry Block",
                            fields=["block_number", "export_block_no"],
                            limit_page_length=0):
        q, e = _s(b.get("block_number")), _s(b.get("export_block_no"))
        if not q or not e:
            continue
        if q in q2e and q2e[q] != e:
            dupq.add(q)
        if e in e2q and e2q[e] != q:
            dupe.add(e)
        q2e[q], e2q[e] = e, q
    for k in dupq:
        q2e.pop(k, None)
    for k in dupe:
        e2q.pop(k, None)
    return q2e, e2q


@frappe.whitelist()
def dc_weight_check_v2():
    """Challan total against challan total, with identity resolved on both sides.

    A challan gets a verdict only when the agency has a row for EVERY block on
    it. Anything else says plainly that our matching found nothing for those
    blocks - it never claims the agency failed to send them.
    """
    from dolphin_theme.block_resolve import try_resolve

    by_block = _arrival_rows_by_block()
    by_number = _arrival_rows_by_number()
    q2e, e2q = _other_number_index()

    # THE AGENCY SENDS A RUNNING STOCK LIST, NOT ONE SHEET PER TRUCK.
    #
    # 23 Aug 2026, counted on the live site:
    #
    #     ARR-30Jul2026-NA    199 distinct numbers
    #     ARR-30Jul2026-NA-2  192   - shares 192 with the one above
    #     ARR-05Aug2026-NA    200   - shares 199
    #     ARR-13Aug2026-NA    200   - shares 199
    #
    # Four sheets, the same 200 blocks. And the 31 blocks that were on the 27 Jul
    # sheet but have since dropped off it are, every one of them, already in a
    # shipment lot. So each sheet is the agency telling us what is standing at the
    # port THAT DAY - a block joins the list when it arrives and leaves it when it
    # is loaded.
    #
    # That is also what the 198 "duplicates" were: not duplicates at all, the same
    # list arriving again the following week.
    #
    # So a block is read from the LATEST sheet that carries it. One weight per
    # block, chosen by date and not by whichever row was found first - which is
    # what produced -34.82 MT on a two-block truck back on 22 Aug.
    _system, _expmap, _qrymap = _sheet_number_system()
    _ds_all = _arrival_docstatus()
    _when = {}
    for a in frappe.get_all("Port Arrival",
                            fields=["name", "arrival_date", "creation"],
                            limit_page_length=0):
        _when[_s(a.name)] = _s(a.get("arrival_date") or a.get("creation"))

    by_export, _seen_at = {}, {}
    for r in frappe.get_all(
            "Port Arrival Block",
            fields=["name", "parent", "block_no", "length", "width", "height",
                    "cbm", "net_wt"],
            limit_page_length=0):
        k = _s(r.get("block_no"))
        sheet = _s(r.get("parent"))
        if not k or _system.get(sheet) != "export":
            continue
        if len(_expmap.get(k, set())) != 1:
            continue                      # two blocks export under it - never guess
        r["submitted"] = 1 if _ds_all.get(sheet) == 1 else 0
        when = _when.get(sheet, "")
        # keep the newest row for this block, and only that one
        if k in _seen_at and _seen_at[k] >= when:
            continue
        _seen_at[k] = when
        by_export[k] = [r]

    def _agency_rows(written, quarry_no, export_no):
        """The agency's rows for one block on a challan, found the way a person
        finds them. Returns (rows, how) - `how` says which reading found them, so
        nothing is matched silently.

        1. the block's EXPORT number, on a sheet that is written in export
           numbers - the reading the sheets themselves prove is the right one
        2. the number written on the challan, as written
        3. the block's other number - challan in quarry numbers, sheet in export
           numbers, or the other way round
        4. only then through the block record, which is where the old matcher
           started and where 87 of 110 blocks were being lost
        """
        for k in (_s(export_no), _s(written)):
            if k and k in by_export:
                return by_export[k], "export number"
        w0 = _s(written)
        alt_e = q2e.get(_s(quarry_no)) or q2e.get(w0)
        if alt_e and alt_e in by_export:
            return by_export[alt_e], "export number"
        for k in (_s(written), _s(export_no), _s(quarry_no)):
            if k and k in by_number:
                return by_number[k], "number"
        w = _s(written)
        for alt in (q2e.get(w), e2q.get(w)):
            if alt and alt in by_number:
                return by_number[alt], "other number"
        hit, _why = try_resolve(w, allow_record_name=True)
        nm = _s((hit or {}).get("name"))
        if nm and nm in by_block:
            return by_block[nm], "block record"
        return None, ""

    # 22 Aug 2026: the real field names on this site are dc_no / delivery_challan_no /
    # vehicle. Read the meta rather than assuming - the first version guessed and the
    # query died with "Unknown column 'challan_no'". Never guess a field name.
    # 22 Aug 2026, his instruction, verbatim: "What ever Dc not submitted should
    # not be listed or considered here it will be in Dc list, Dc consolidated etc.
    # whatever we are speaking at port, port and stock is simply after dC
    # submitted dispatched so make it strict that anywhere it is same in draft is
    # not submitted and consider as done only after draft."
    #
    # A draft challan is DMG permit paperwork, not a dispatch. It has no business
    # on any port screen. ledger_view has filtered on docstatus 1 since 17 Aug;
    # this did not, which is why 135 challans appeared here when only 62 have
    # actually left the quarry.
    meta = frappe.get_meta("Delivery Challan")
    num_field = next((f for f in ("dc_no", "delivery_challan_no", "challan_no")
                      if meta.has_field(f)), None)
    veh_field = next((f for f in ("vehicle", "vehicle_no", "driver")
                      if meta.has_field(f)), None)
    fields = ["name", "docstatus"]
    if num_field:
        fields.append(num_field)
    if veh_field:
        fields.append(veh_field)
    challans = frappe.get_all("Delivery Challan", filters={"docstatus": 1},
                              fields=fields, limit_page_length=0)
    detail = []
    agree = flagged = incomplete = never = 0

    for c in challans:
        rows = frappe.get_all(
            "DC Block Row",
            filters={"parent": c["name"]},
            # 25 Aug 2026: the dimensions come too, because the SIZE is now a
            # gate on the match itself, not just a figure on a screen.
            fields=["block", "block_no", "export_block_no", "gross_tonnage",
                    "length_gross", "width_gross", "height_gross"],
            limit_page_length=0,
        )
        if not rows:
            continue
        ours = 0.0
        matched, unmatched = 0, []
        theirs = 0.0
        sheets, per_block, draft_sheet = set(), [], False
        for r in rows:
            try:
                ours += float(r.get("gross_tonnage") or 0)
            except Exception:
                pass
            key = (_s(r.get("export_block_no")) or _s(r.get("block_no"))
                   or _s(r.get("block")))
            arr, how = _agency_rows(key, _s(r.get("block_no")),
                                    _s(r.get("export_block_no")))
            if arr:
                matched += 1
                for a in arr:
                    sheets.add(_s(a.get("parent")))
                    if not a.get("submitted"):
                        draft_sheet = True
                a0 = arr[0]
                per_block.append({"block": key,
                                  "sheet": _s(a0.get("parent")),
                                  "found_by": how,
                                  "submitted": 1 if a0.get("submitted") else 0,
                                  "port_l": a0.get("length"), "port_w": a0.get("width"),
                                  "port_h": a0.get("height"), "port_mt": a0.get("net_wt"),
                                  "on_more_than_one_sheet": len(arr) > 1})
                try:
                    theirs += float(a0.get("net_wt") or 0)
                except Exception:
                    pass
            else:
                unmatched.append(key)
                per_block.append({"block": key, "sheet": None, "submitted": 0,
                                  "found_by": "", "port_l": None, "port_w": None,
                                  "port_h": None, "port_mt": None,
                                  "on_more_than_one_sheet": False})

        # A verdict is only honest when ALL of these hold. Any one missing and the
        # figures are shown for information with no verdict at all - never a red
        # number that cannot be justified.
        n = len(rows)
        # 23 Aug 2026: "all rows from ONE sheet" was the right guard when a sheet
        # was thought to be one truck's weighing. It is the wrong guard for a
        # running stock list, where a block naturally appears on every sheet since
        # it arrived - it refused a verdict on almost every challan for a reason
        # that is not a fault. What actually has to hold is that no block is
        # counted twice, and that is now guaranteed upstream: exactly one row per
        # block, taken from the newest sheet carrying it.
        one_sheet = True
        # ------------------------------------------------------------------
        # THE TRUCK RULE, and why a loose match now refuses a verdict.
        # 25 Aug 2026, his words, and he asked me to hold on to them:
        #
        #   "99 of 100 times no truck will have less number of blocks from what
        #    it originally had when loaded. The size neither varies, nor
        #    compresses, weight neither increases nor decreases."
        #
        #   "why is there a variation in weight you might ask, the answer is due
        #    to the packing stones they give for support adds as additional
        #    weight since the whole truck is weighed at port then tare weight and
        #    they divide the total weight divided by the number of the blocks
        #    since end of it what matters for them, us and client is total
        #    weight."
        #
        # So the agency does not weigh BLOCKS. They weigh the TRUCK, take off
        # tare, and divide by the number of blocks. A per-block agency weight is
        # an average of a truckload, and our own figure is specific gravity off a
        # sample cube. Two approximations of the same total.
        #
        # Therefore a large gap is almost never stone going astray. It is my
        # matching being wrong - and he had to tell me so about block 802:
        #
        #   "802 block etc is not the exact comparision you are getting the
        #    blocks wrong, there cannot be such a difference in the weight so
        #    increase the parameters to check ... you are getting them wrong just
        #    based on assumption"
        #
        # "Increase the parameters" is exactly this. A verdict is now given only
        # when EVERY block on the challan was found by a number that identifies
        # it - its export number, or the number written on the challan. A block
        # reached through the OTHER number, or through the block record, is a
        # guess, and a guess must never carry a weight into a total that gets a
        # red verdict printed under it.
        # ------------------------------------------------------------------
        STRONG = ("export number", "number")
        loose = [p for p in per_block
                 if p.get("sheet") and _s(p.get("found_by")) not in STRONG]

        # ------------------------------------------------------------------
        # THE SIZE GATE.  25 Aug 2026, his words:
        #   "I doubt these blocks are exact match many a times it is happening
        #    wrong blocks are being tried to match by you"
        #
        # He was right, and the screen he was looking at proved it: block 823 on
        # challan 0011 read ours 238x212x140 (7.06 CBM) against port 220x105x85
        # (1.96 CBM). Under his own truck rule - "the size neither varies, nor
        # compresses" - a block does not lose two thirds of its volume on a
        # lorry. That is not dressing loss and not a mistyped figure; it is the
        # wrong agency row attached to the right-looking number, and the -13.77
        # MT printed under it was fiction.
        #
        # So a size nowhere near ours REFUSES the match outright. The challan
        # gets no verdict and the offending blocks are named, which is the
        # honest thing to put on a screen: not "the port lost stone" but "we
        # matched the wrong row".
        # ------------------------------------------------------------------
        def _vol(a, b, c):
            try:
                a, b, c = float(a or 0), float(b or 0), float(c or 0)
            except Exception:
                return 0.0
            return round(a * b * c / 1e6, 3) if (a and b and c) else 0.0

        ours_by_block = {}
        for r in rows:
            k = (_s(r.get("export_block_no")) or _s(r.get("block_no"))
                 or _s(r.get("block")))
            if k:
                ours_by_block[k] = r
        wrong_size = []
        for p in per_block:
            if not p.get("sheet"):
                continue
            src = ours_by_block.get(_s(p.get("block"))) or {}
            oc = _vol(src.get("length_gross"), src.get("width_gross"),
                      src.get("height_gross"))
            tc = _vol(p.get("port_l"), p.get("port_w"), p.get("port_h"))
            if not (oc and tc):
                continue
            ratio = round(tc / oc, 3)
            if ratio < SIZE_MATCH_LOW or ratio > SIZE_MATCH_HIGH:
                wrong_size.append({"block": _s(p.get("block")),
                                   "our_cbm": oc, "their_cbm": tc,
                                   "ratio": ratio})
        if matched == 0:
            verdict, state = "Not sent yet", "never"
            never += 1
            diff = None
        elif matched < n:
            verdict, state = "Incomplete", "incomplete"
            incomplete += 1
            diff = None
        elif draft_sheet:
            verdict, state = "Unconfirmed sheet", "incomplete"
            incomplete += 1
            diff = None
        elif wrong_size:
            verdict, state = "Wrong block matched - no verdict", "incomplete"
            incomplete += 1
            diff = None
        elif loose:
            verdict, state = "Matched loosely - no verdict", "incomplete"
            incomplete += 1
            diff = None
        elif not one_sheet:
            verdict, state = "Rows from more than one sheet", "incomplete"
            incomplete += 1
            diff = None
        else:
            diff = round(theirs - ours, 3)
            kept = (theirs / ours) if ours else 1.0
            if diff > AUTO_TOL_MT:
                # heavier at the port than what left the quarry - cannot happen
                verdict, state = "FLAG", "flagged"
                flagged += 1
            elif kept < DRESSING_LOOK_AT_IT:
                verdict, state = "FLAG", "flagged"
                flagged += 1
            elif diff < -AUTO_TOL_MT:
                verdict, state = "Dressed", "agree"
                agree += 1
            else:
                verdict, state = "Agrees", "agree"
                agree += 1

        detail.append({
            "dc": c["name"],
            "challan_no": (c.get(num_field) if num_field else None) or c["name"],
            "our_vehicle": (c.get(veh_field) if veh_field else "") or "",
            "blocks": n,
            "agency_rows": matched,
            "our_total": round(ours, 3),
            "their_total": round(theirs, 3) if matched else None,
            "difference": diff,
            "verdict": verdict,
            "state": "submitted" if c.get("docstatus") == 1 else "draft",
            "unmatched": unmatched[:20],
            "sheets": sorted([x for x in sheets if x])[:6],
            "any_draft_sheet": 1 if draft_sheet else 0,
            # Named, so the screen can say WHICH blocks were only guessed at
            # rather than leaving "no verdict" looking like an app fault.
            "loose_matches": [{"block": p.get("block"), "found_by": p.get("found_by")}
                              for p in loose][:20],
            # Named, so the screen says "we matched the wrong row" rather than
            # implying the port lost stone.
            "wrong_size_matches": wrong_size[:20],
            "per_block": per_block,
        })

    detail.sort(key=lambda x: (0 if x["verdict"] == "FLAG" else
                               1 if x["verdict"] == "Incomplete" else
                               2 if x["verdict"] == "Agrees" else 3,
                               x["challan_no"]))
    return {
        "agree": agree, "flagged": flagged, "incomplete": incomplete,
        "agency_never_sent": never, "tolerance_mt": AUTO_TOL_MT,
        "detail": detail,
    }


# ==========================================================================
# SIZE vs CBM - ASK, THEN TAKE THE ANSWER.  24 Aug 2026
#
# His words, 23 Aug 2026:
#   "under reconcilation you have to highlight this major difference and ask
#    user to choose which is correct"
#   "these mistakes should not happen even at the packing list and invoice
#    level run this verification and give second chance to correct"
#
# Reconciliation could already SHOW the question - 322 x 189 x 185 does not
# make the 12.40 printed beside it. What it could not do was take the answer.
# This does. Two answers, and only two:
#
#   choice="size"  the size is right, so the CBM beside it is wrong. Recompute
#                  L x W x H / 1e6 and write that everywhere the block appears.
#   choice="cbm"   the CBM is right, so the size is a typo. Nothing is computed
#                  and nothing is guessed - the CBM stands and the size is left
#                  for a person to retype from the inspection sheet.
#
# The measurement is Dolphin's and the CBM is Dolphin's, so this only ever
# rewrites our own figures. It never touches the agency's weight and it never
# touches a Quarry Block reading - that is the record of what came out of the
# pit and it is never rewritten.
#
# SUBMITTED DOCUMENTS ARE REFUSED, not silently skipped. A submitted challan or
# invoice has left the building; changing its figures behind someone's back is
# exactly the financial damage this whole exercise exists to prevent. The reply
# names the document so a person can decide to amend it deliberately.
#
# The reason is compulsory, as it is on a measurement change.
# ==========================================================================

# parent doctype -> (child doctype, table field, L, W, H, CBM)
CBM_DOCS = [
    ("Delivery Challan",    "DC Block Row",       "dc_block_rows",
     "length_gross", "width_gross", "height_gross", "gross_volume"),
    ("Local Tax Invoice",   "Tax Invoice Block",  "blocks",
     "length_gross", "width_gross", "height_gross", "gross_volume"),
    ("Export Shipment Lot", "Shipment Lot Block", "blocks",
     "length", "width", "height", "cbm"),
    ("Shipping Document",   "Shipping Block",     "blocks",
     "length", "width", "height", "net_volume"),
]

# A CBM is printed to three decimals, so anything under a litre is rounding,
# not a disagreement. Above that, somebody has to say which figure is right.
CBM_TOL = 0.010


def _cbm_of(l, w, h):
    try:
        v = (float(l or 0) * float(w or 0) * float(h or 0)) / 1000000.0
    except Exception:
        return None
    return round(v, 3) if v > 0 else None


def _cbm_rows():
    """Every document row whose size does not multiply out to the CBM beside it."""
    out = []
    for parent_dt, child_dt, _tf, fl, fw, fh, fc in CBM_DOCS:
        try:
            rows = frappe.get_all(
                child_dt, filters={"parenttype": parent_dt},
                fields=["name", "parent", "block", "block_no",
                        fl, fw, fh, fc], limit_page_length=0)
        except Exception:
            continue
        if not rows:
            continue
        states = {d.name: d.docstatus for d in frappe.get_all(
            parent_dt, fields=["name", "docstatus"], limit_page_length=0)}
        for r in rows:
            computed = _cbm_of(r.get(fl), r.get(fw), r.get(fh))
            printed = r.get(fc)
            if computed is None or printed in (None, ""):
                continue
            try:
                printed = float(printed)
            except Exception:
                continue
            if abs(computed - printed) <= CBM_TOL:
                continue
            out.append({
                "parent_doctype": parent_dt, "child_doctype": child_dt,
                "document": _s(r.get("parent")), "row": _s(r.get("name")),
                "docstatus": states.get(_s(r.get("parent")), 0),
                "block": _s(r.get("block")),
                "block_no": _s(r.get("block_no")) or _s(r.get("block")),
                "size": [r.get(fl), r.get(fw), r.get(fh)],
                "cbm_printed": round(printed, 3),
                "cbm_from_size": computed,
                "gap": round(computed - printed, 3),
                "cbm_field": fc,
            })
    return out


@frappe.whitelist()
def size_vs_cbm_questions():
    """The question Reconciliation asks, per block. Changes nothing.

    One entry per block, with every document the disagreement reaches, so a
    person answers ONCE for a block rather than once per document and cannot
    give two different answers to the same question.
    """
    by_block = {}
    for r in _cbm_rows():
        key = r["block_no"] or r["block"]
        if not key:
            continue
        b = by_block.setdefault(key, {
            "block_no": key, "block": r["block"], "documents": [],
            "sizes_seen": [], "biggest_gap": 0.0, "any_submitted": 0})
        b["documents"].append(r)
        if r["size"] not in b["sizes_seen"]:
            b["sizes_seen"].append(r["size"])
        if abs(r["gap"]) > abs(b["biggest_gap"]):
            b["biggest_gap"] = r["gap"]
        if r["docstatus"] == 1:
            b["any_submitted"] = 1
    out = sorted(by_block.values(), key=lambda x: -abs(x["biggest_gap"]))
    return {
        "count": len(out),
        "tolerance_cbm": CBM_TOL,
        "blocks": out,
        "note": ("The size is ours and the CBM is ours. One of the two is "
                 "wrong. A person says which - nothing is guessed."),
    }


@frappe.whitelist()
def resolve_size_vs_cbm(block=None, choice=None, person=None, reason=None,
                        dry_run=0):
    """Take the answer to the size-vs-CBM question and write it through.

    choice="size"  the size is right -> recompute the CBM everywhere
    choice="cbm"   the CBM is right  -> write nothing, record that the size
                   needs retyping from the inspection sheet
    """
    block = _s(block)
    choice = _s(choice).lower()
    person = _s(person)
    reason = _s(reason)
    dry = _s(dry_run) in ("1", "true", "yes")

    if not block:
        frappe.throw("Which block?")
    if choice not in ("size", "cbm"):
        frappe.throw('choice must be "size" (the size is right) or "cbm" '
                     '(the CBM is right).')
    # The person is who is logged in - authenticated beats typed, and it cannot
    # be left blank or filled in with someone else's name. The REASON is what a
    # person has to supply, and it is compulsory exactly as it is on a
    # measurement change. His rule, 23 Aug 2026: "typing the reason is better to
    # understand in the long run so yes compulsory."
    if not person:
        person = _s(frappe.session.user)
    if not reason:
        frappe.throw("A written reason is required - this changes a figure "
                     "that money is raised on.")

    mine = [r for r in _cbm_rows()
            if (r["block_no"] == block or r["block"] == block)]
    if not mine:
        return {"block": block, "changed": 0,
                "message": "Nothing to answer - every document already "
                           "multiplies out for this block."}

    submitted = [r for r in mine if r["docstatus"] == 1]
    if submitted:
        return {
            "block": block, "changed": 0, "refused": 1,
            "submitted_documents": sorted({r["document"] for r in submitted}),
            "message": ("Refused. These documents are submitted, and a "
                        "submitted document is not edited behind someone's "
                        "back. Amend the document deliberately, then answer "
                        "this again."),
        }

    stamp = "{0} - {1} ({2})".format(frappe.utils.now(), reason, person)
    changed, skipped = [], []
    for r in mine:
        if choice == "cbm":
            skipped.append({"document": r["document"], "row": r["row"],
                            "why": "the CBM stands; the size needs retyping "
                                   "from the inspection sheet"})
            continue
        if not dry:
            frappe.db.set_value(r["child_doctype"], r["row"],
                                r["cbm_field"], r["cbm_from_size"],
                                update_modified=False)
        changed.append({"document": r["document"], "row": r["row"],
                        "from": r["cbm_printed"], "to": r["cbm_from_size"]})

    if not dry:
        for doc in sorted({r["document"] for r in mine}):
            parent_dt = next(x["parent_doctype"] for x in mine
                             if x["document"] == doc)
            try:
                frappe.get_doc(parent_dt, doc).add_comment(
                    "Comment",
                    "Size vs CBM answered for block {0}: {1} is correct. {2}"
                    .format(block,
                            "the size" if choice == "size" else "the CBM",
                            stamp))
            except Exception:
                pass
        frappe.db.commit()

    return {"block": block, "choice": choice, "dry_run": 1 if dry else 0,
            "changed": len(changed), "detail": changed, "left_alone": skipped,
            "recorded": stamp}


# ==========================================================================
# BEFORE A LOT IS BUILT.  24 Aug 2026
#
# His rule: validate export numbers and measurements BEFORE a lot, not after.
# The packing list is where a missing export number or a missing measurement
# currently surfaces, and the packing list is exactly where he said a mistake
# must never surface.
#
# Two questions, and only two. Neither is a matter of judgement:
#   * has this block got an export number
#   * has this block got a measurement in use
# A size that disagrees with its CBM is reported alongside as a WARNING, not a
# blocker - it has an owner (the question above) and it does not stop loading.
# ==========================================================================
@frappe.whitelist()
def lot_readiness(blocks=None):
    """Is every one of these blocks fit to go on a lot? Changes nothing."""
    if isinstance(blocks, str):
        try:
            blocks = frappe.parse_json(blocks)
        except Exception:
            blocks = [b.strip() for b in blocks.split(",") if b.strip()]
    blocks = [_s(b) for b in (blocks or []) if _s(b)]
    if not blocks:
        return {"ok": 1, "checked": 0, "blockers": [], "warnings": [],
                "message": "No blocks given."}

    qbs = frappe.get_all(
        "Quarry Block",
        filters=[["name", "in", blocks]],
        fields=["name", "block_number", "export_block_no",
                "length_gross", "width_gross", "height_gross"],
        limit_page_length=0)
    if len(qbs) < len(blocks):
        qbs += frappe.get_all(
            "Quarry Block",
            filters=[["block_number", "in", blocks]],
            fields=["name", "block_number", "export_block_no",
                    "length_gross", "width_gross", "height_gross"],
            limit_page_length=0)
    seen, uniq = set(), []
    for q in qbs:
        if q.name not in seen:
            seen.add(q.name)
            uniq.append(q)

    # the measurement in use: the NEWEST buyer inspection, quarry as fallback
    bi = {}
    if uniq:
        for r in frappe.get_all(
                "Buyer Inspection Block",
                filters=[["block", "in", [q.name for q in uniq]]],
                fields=["block", "parent", "creation", "length_gross",
                        "width_gross", "height_gross"],
                order_by="creation asc", limit_page_length=0):
            bi[_s(r.block)] = r          # ascending, so the last write is newest

    found = {q.name for q in uniq} | {_s(q.block_number) for q in uniq}
    blockers, warnings = [], []
    for b in blocks:
        if b not in found:
            blockers.append({"block": b, "why": "no such block"})

    for q in uniq:
        label = _s(q.export_block_no) or _s(q.block_number) or q.name
        if not _s(q.export_block_no):
            blockers.append({
                "block": _s(q.block_number) or q.name,
                "why": "no export number",
                "fix": "The packing list and the invoice are printed on the "
                       "export number. Give it one before it goes on a lot."})
        src = bi.get(q.name)
        dims = ([src.length_gross, src.width_gross, src.height_gross] if src
                else [q.length_gross, q.width_gross, q.height_gross])
        if not all(dims):
            blockers.append({
                "block": label, "why": "no measurement in use",
                "fix": "Neither a buyer inspection nor the quarry reading "
                       "gives this block a full size."})
        elif not src:
            warnings.append({
                "block": label, "why": "priced on the quarry reading",
                "detail": "No buyer inspection exists for this block."})

    qs = size_vs_cbm_questions() or {}
    open_q = {_s(x.get("block_no")) for x in (qs.get("blocks") or [])}
    for q in uniq:
        label = _s(q.export_block_no) or _s(q.block_number) or q.name
        if label in open_q or _s(q.block_number) in open_q:
            warnings.append({
                "block": label, "why": "size vs CBM unanswered",
                "detail": "Its size does not multiply out to the CBM beside "
                          "it on at least one document."})

    return {
        "ok": 1 if not blockers else 0,
        "checked": len(blocks),
        "blockers": blockers,
        "warnings": warnings,
        "message": ("Every block has an export number and a measurement."
                    if not blockers else
                    "{0} block(s) are not fit to go on a lot yet."
                    .format(len(blockers))),
    }


# ==========================================================================
# EMPTY ARRIVAL SHEETS.  24 Aug 2026
#
# [stated] "remove empty sheets and provide a button to remove for empty
#           sheets alone?"
#
# ARR-19Aug2026-NA was created by the email sync from a mail that carried
# nothing parsable: no rows, no stored spreadsheet. It cannot confirm anything
# and cannot move anything, but it sits in the list looking like a sheet, and a
# confirmed empty sheet would be a record asserting "the agency reported
# nothing" - worse than no record at all.
#
# THE DISTINCTION THAT MATTERS, and the reason this is not a one-line delete:
#
#   no rows AND no file   -> genuinely EMPTY. The email had no spreadsheet.
#                            Nothing was ever lost, so removing it loses
#                            nothing. This is the only thing the button deletes.
#
#   no rows BUT a file    -> a PARSE FAILURE, not an empty sheet. The agency
#                            DID send a spreadsheet and we failed to read it.
#                            Deleting that destroys the evidence of a bug and
#                            silently loses an arrival. REFUSED, by name, every
#                            time - it needs looking at, not removing.
#
# Every guard is re-checked on the server at the moment of deletion. The client
# says which sheets it means; it is never believed about whether they are empty.
# ==========================================================================
@frappe.whitelist()
def empty_arrivals():
    """Arrival sheets holding no blocks. Changes nothing.

    Split into what can safely be removed and what must not be.
    """
    sheets = frappe.get_all("Port Arrival",
                            fields=["name", "arrival_date", "docstatus",
                                    "email_sender", "source_sheet", "creation"],
                            limit_page_length=0)
    if not sheets:
        return {"removable": [], "look_at_these": [], "count": 0}

    counts = {}
    for r in frappe.get_all("Port Arrival Block",
                            filters={"parenttype": "Port Arrival"},
                            fields=["parent"], limit_page_length=0):
        counts[_s(r.parent)] = counts.get(_s(r.parent), 0) + 1

    files = {}
    for f in frappe.get_all("File",
                            filters={"attached_to_doctype": "Port Arrival"},
                            fields=["attached_to_name", "file_name"],
                            limit_page_length=0):
        files.setdefault(_s(f.attached_to_name), []).append(_s(f.file_name))

    removable, look = [], []
    for s in sheets:
        n = _s(s.name)
        if counts.get(n):
            continue                      # has blocks - not empty, not our business
        item = {"sheet": n, "arrival_date": _s(s.arrival_date),
                "sender": _s(s.email_sender), "source_sheet": _s(s.source_sheet),
                "created": _s(s.creation)[:16], "files": files.get(n, [])}
        if s.docstatus != 0:
            item["why_not"] = ("confirmed - a confirmed sheet is a record and is "
                               "never deleted")
            look.append(item)
        elif files.get(n):
            item["why_not"] = ("a spreadsheet IS attached but no rows were read - "
                               "this is a parsing failure, not an empty sheet. "
                               "Deleting it would lose a real arrival and hide "
                               "the bug that lost it.")
            look.append(item)
        else:
            removable.append(item)

    return {"removable": removable, "look_at_these": look,
            "count": len(removable),
            "note": ("Only a sheet with no rows AND no attached spreadsheet is "
                     "removable. Everything else is listed for a person.")}


@frappe.whitelist()
def delete_empty_arrivals(sheets=None, person=None, dry_run=0):
    """Remove arrival sheets that hold nothing and never held anything.

    Every guard is re-checked here. The caller says WHICH sheets; it is never
    believed about WHETHER they are empty.
    """
    if isinstance(sheets, str):
        try:
            sheets = frappe.parse_json(sheets)
        except Exception:
            sheets = [x.strip() for x in sheets.split(",") if x.strip()]
    dry = _s(dry_run) in ("1", "true", "yes")
    person = _s(person) or _s(frappe.session.user)

    state = empty_arrivals() or {}
    allowed = {_s(x.get("sheet")) for x in (state.get("removable") or [])}
    blocked = {_s(x.get("sheet")): _s(x.get("why_not"))
               for x in (state.get("look_at_these") or [])}

    wanted = [_s(s) for s in (sheets or []) if _s(s)] or sorted(allowed)

    removed, refused = [], []
    for n in wanted:
        if n in blocked:
            refused.append({"sheet": n, "why": blocked[n]})
            continue
        if n not in allowed:
            refused.append({"sheet": n,
                            "why": "not empty, or no such sheet - refused"})
            continue
        if dry:
            removed.append({"sheet": n, "dry_run": 1})
            continue
        try:
            frappe.delete_doc("Port Arrival", n, ignore_permissions=False,
                              delete_permanently=False)
            removed.append({"sheet": n, "removed_by": person})
        except Exception as e:
            refused.append({"sheet": n, "why": "delete failed: " + str(e)})

    if removed and not dry:
        frappe.db.commit()

    return {"removed": len(removed), "detail": removed,
            "refused": refused, "dry_run": 1 if dry else 0}


# ============================================================================
# ALL TRANSPORTED - the step straight after a challan is submitted.  25 Aug 2026
#
# [stated] "Create a separate tab before arrivals for all transported i.e the
#  next step after submit Dc and dont mix up with reconcilation it should show
#  only Dc and options to choose full dc or individual blocks to send it to dc
#  draft if in case due to some issue there is short loading may be all blocks
#  didnt fit and so 2 blocks letf out etc and make note once a DC is returned
#  after submit it shouldnt appear here in all transported too. So this Tab will
#  just show all transported and blocks dc Eye icon like earlier."
#
# So this is a CHALLAN screen, not a block soup. One row per submitted challan,
# its blocks underneath, and one thing you can do from here: send a whole
# challan or the blocks that did not fit back to a draft challan.
#
# WHY IT IS ITS OWN TAB. Short loading is not a reconciliation question. Nobody
# is comparing anything - the truck simply did not hold everything, somebody at
# the quarry knows which two were left behind, and the paperwork has to say so
# before the agency's sheet ever arrives. Mixing it into Reconciliation was
# asking a loading question on a weighing screen.
# ============================================================================

RETURNED_TO_STATUS = "In Stock"


def _returned_field_ready():
    """Somewhere on Quarry Block to remember what a block was before it was
    returned from a challan. Created once; never fails the caller."""
    try:
        meta = frappe.get_meta("Quarry Block")
        if meta.has_field("status_before_dc_return"):
            return True
        from frappe.custom.doctype.custom_field.custom_field import create_custom_field
        create_custom_field("Quarry Block", {
            "fieldname": "status_before_dc_return",
            "label": "Status Before DC Return",
            "fieldtype": "Data",
            "hidden": 1,
            "read_only": 1,
            "no_copy": 1,
        }, ignore_validate=True)
        frappe.clear_cache(doctype="Quarry Block")
        return frappe.get_meta("Quarry Block").has_field("status_before_dc_return")
    except Exception:
        return False


def _returned_blocks_index():
    """Every block number that has been returned from a challan and not undone.

    Kept as a Frappe cache-backed read of the block's own field rather than a
    separate ledger: one place to be wrong is better than two.
    """
    out = {}
    try:
        if not frappe.get_meta("Quarry Block").has_field("status_before_dc_return"):
            return out
        rows = frappe.get_all(
            "Quarry Block",
            filters={"status_before_dc_return": ["!=", ""]},
            fields=["name", "block_number", "export_block_no",
                    "status", "status_before_dc_return"],
            limit_page_length=0) or []
        for r in rows:
            for k in (_s(r.get("export_block_no")), _s(r.get("block_number"))):
                if k:
                    out[k] = r
    except Exception:
        return {}
    return out


@frappe.whitelist()
def transported_view():
    """Submitted challans and the stone on them, challan by challan.

    A challan appears here from the moment it is submitted. It leaves when every
    block on it has been returned - his rule: "once a DC is returned after submit
    it shouldnt appear here in all transported too".

    Nothing here is a judgement. The state of each block is reported as it is:
    on the road, at the port, in a lot, loaded, or returned.
    """
    rows = ledger_view() or []
    if isinstance(rows, dict):
        rows = rows.get("rows") or []
    returned = _returned_blocks_index()

    STATE_WORDS = {
        "await": "On the road",
        "port": "At port",
        "lot": "In a lot",
        "load": "Loaded",
    }

    by_dc = {}
    for r in rows:
        dc = _s(r.get("dc"))
        if not dc or _s(r.get("source")) != "dc":
            continue
        num = _s(r.get("export_block_no")) or _s(r.get("block_no"))
        d = by_dc.setdefault(dc, {
            "dc": dc, "blocks": [], "returned": 0, "on_road": 0,
            "truck": "", "port": "", "consignee": "", "mark": "",
            "ours_mt": 0.0, "cbm": 0.0,
        })
        if not d["truck"]:
            d["truck"] = _s(r.get("truck"))
        if not d["port"]:
            d["port"] = _s(r.get("port_code")) or _s(r.get("port"))
        if not d["consignee"]:
            d["consignee"] = _s(r.get("consignee"))
        if not d["mark"]:
            d["mark"] = _s(r.get("mark"))
        st = _s(r.get("state"))
        was_returned = num in returned
        if was_returned:
            d["returned"] += 1
        elif st == "await":
            d["on_road"] += 1
        d["ours_mt"] += flt(r.get("ton") or 0)
        d["cbm"] += flt(r.get("dc_cbm") or 0)
        d["blocks"].append({
            "block_no": num,
            "quarry_block_no": _s(r.get("quarry_block_no")),
            "qb": r.get("qb"),
            "state": st,
            "where": ("Returned - not loaded" if was_returned
                      else STATE_WORDS.get(st, st or "")),
            "returned": 1 if was_returned else 0,
            "l": r.get("dc_l"), "w": r.get("dc_w"), "h": r.get("dc_h"),
            "cbm": r.get("dc_cbm"), "mt": r.get("ton"),
            "grade": _s(r.get("grade")),
            "arrival": _s(r.get("arrival")),
        })

    dcs = []
    for d in by_dc.values():
        # Every block returned means the challan carried nothing. It goes.
        if d["blocks"] and d["returned"] >= len(d["blocks"]):
            continue
        d["count"] = len(d["blocks"])
        d["ours_mt"] = round(d["ours_mt"], 3)
        d["cbm"] = round(d["cbm"], 3)
        dcs.append(d)

    meta = {}
    try:
        names = [d["dc"] for d in dcs]
        if names:
            for x in frappe.get_all(
                    "Delivery Challan",
                    filters={"name": ["in", names]},
                    fields=["name", "delivery_challan_no", "dc_date", "docstatus"],
                    limit_page_length=0) or []:
                meta[_s(x.get("name"))] = x
    except Exception:
        meta = {}
    for d in dcs:
        m = meta.get(d["dc"]) or {}
        # THE NUMBER-TYPE TRAP, named on the row itself. DC-DCDG-070 is challan
        # 0033. The record id suffix looks exactly like a challan number and is
        # not one, and every screen that has ever confused the two has cost a day.
        d["challan_no"] = _s(m.get("delivery_challan_no"))
        d["dc_date"] = _s(m.get("dc_date"))
    dcs.sort(key=lambda x: (x.get("dc_date") or "", x.get("dc")), reverse=True)

    return {
        "challans": dcs,
        "count": len(dcs),
        "blocks": sum(d["count"] for d in dcs),
        "on_road": sum(d["on_road"] for d in dcs),
        "returned": sum(d["returned"] for d in dcs),
    }


@frappe.whitelist()
def return_from_dc(dc=None, blocks=None, reason=None, person=None, dry_run=0):
    """Short loading: send a whole challan, or the blocks that did not fit, back.

    His case, in his words: "due to some issue there is short loading may be all
    blocks didnt fit and so 2 blocks letf out".

    What this does and does not do, deliberately:
      * it does NOT cancel or amend the submitted challan. The challan is the
        record of what was written when the truck left, and cancelling a
        submitted document to correct a loading fact is a bigger, louder change
        than the fact deserves.
      * it returns each named block to stock, records on the block's own history
        who returned it, from which challan and why, and remembers exactly what
        the block was so the return can be undone.
      * it optionally puts the returned blocks straight onto a NEW DRAFT challan
        copied from the original, which is the thing he actually asked for -
        "send it to dc draft" - so they can go on the next truck.

    A reason is compulsory. Reversible by undo_return_from_dc.
    """
    blocks = frappe.parse_json(blocks) if isinstance(blocks, str) else (blocks or [])
    dry = _s(dry_run) in ("1", "true", "True", "yes")
    reason = _s(reason)
    person = _s(person) or _s(frappe.session.user)
    dc = _s(dc)

    if not dc:
        frappe.throw("No challan given.")
    if not dry and len(reason) < 4:
        frappe.throw("Sending stone back off a submitted challan needs a written "
                     "reason. It is the only record of why the truck went short.")

    doc = frappe.get_doc("Delivery Challan", dc)
    if cint(doc.docstatus) != 1:
        frappe.throw(
            "{0} is not submitted. Only a submitted challan has stone on the road "
            "to send back.".format(dc))

    on_challan = {}
    for r in (doc.get("dc_block_rows") or []):
        for k in (_s(r.get("export_block_no")), _s(r.get("block_no"))):
            if k:
                on_challan[k] = r
    wanted = [_s(b) for b in blocks if _s(b)] or sorted(on_challan.keys())

    from dolphin_theme.block_resolve import try_resolve, set_status

    plan, refused = [], []
    for bn in wanted:
        if bn not in on_challan:
            refused.append({"block": bn, "why": "not on this challan"})
            continue
        # THE RECORD IS THE IDENTITY, NOT THE NUMBER. 25 Aug 2026.
        # The challan row carries a LINK to the Quarry Block, so the identity is
        # already settled and does not have to be guessed from digits that two
        # different blocks may both answer to - which on this site they very
        # often do. The number is still required to match the record, so a wrong
        # link cannot slip through either.
        hit, why = None, ""
        linked = _s((on_challan.get(bn) or {}).get("block"))
        if linked:
            try:
                rec = frappe.db.get_value(
                    "Quarry Block", linked,
                    ["name", "block_number", "export_block_no"], as_dict=True)
            except Exception:
                rec = None
            if rec and _s(bn) in {_s(rec.get("export_block_no")),
                                  _s(rec.get("block_number"))}:
                hit = {"name": _s(rec.get("name")),
                       "block_number": rec.get("block_number"),
                       "export_block_no": rec.get("export_block_no")}
            elif rec:
                refused.append({"block": bn,
                                "why": "the challan row points at a block that does "
                                       "not answer to this number",
                                "detail": "record " + linked})
                continue
        if not hit:
            try:
                from dolphin_theme.identity_guard import check_number
                chk = check_number(bn)
                if not chk.get("ok"):
                    refused.append({"block": bn, "why": chk.get("reason") or "refused",
                                    "detail": chk.get("message")})
                    continue
            except Exception:
                pass
            hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            refused.append({"block": bn, "why": why})
            continue
        answers = {_s(hit.get("export_block_no")), _s(hit.get("block_number"))}
        if _s(bn) not in answers:
            refused.append({"block": bn,
                            "why": "resolved to a block that does not answer to this number"})
            continue
        cur = _s(frappe.db.get_value("Quarry Block", hit["name"], "status"))
        if cur in ("At Port", "In Export Shipment Lot", "Shipped", "Sold"):
            refused.append({"block": bn,
                            "why": "this block is already " + cur +
                                   " - it did reach the port, so it was not left behind"})
            continue
        plan.append({"block": bn, "name": hit["name"], "was": cur})

    if dry:
        return {"dry_run": 1, "dc": dc, "would_return": len(plan),
                "blocks": [p["block"] for p in plan], "refused": refused}

    _returned_field_ready()
    returned = []
    for p in plan:
        try:
            frappe.db.set_value("Quarry Block", p["name"],
                                "status_before_dc_return", p["was"] or "")
        except Exception:
            pass
        res = set_status(
            p["name"], RETURNED_TO_STATUS,
            "returned from {0} - not loaded (short loading) - {1}".format(dc, reason),
            machine="server (all transported)", actor=person, allow_backwards=True)
        if res.get("ok"):
            returned.append(p["block"])
        else:
            refused.append({"block": p["block"], "why": res.get("error") or "refused"})

    draft = None
    if returned:
        draft = _draft_challan_for(doc, returned, person, reason)
    frappe.db.commit()

    return {"ok": 1, "dc": dc, "returned": len(returned),
            "returned_blocks": returned, "draft_challan": draft,
            "refused": refused,
            "note": ("The submitted challan is untouched - it still says what was "
                     "written when the truck left. These blocks are back in stock"
                     + (" and on draft challan " + draft if draft else "") + ".")}


def _draft_challan_for(source, block_numbers, person, reason):
    """A new DRAFT challan carrying the blocks that did not fit.

    His rule stands: a draft challan does not exist yet. Nothing here is
    dispatched, counted or at port until somebody submits it.
    """
    try:
        d = frappe.new_doc("Delivery Challan")
        carry = ("sale_type", "export_country", "country_of_origin", "buyer",
                 "consignee", "export_consignee", "shipping_agency",
                 "description_of_goods", "port_of_loading", "place_of_receipt",
                 "pre_carriage_by", "despatched_through", "terms_of_delivery",
                 "place_of_loading_quarry", "shipping_mark", "destination",
                 "ql_no", "mdp_number", "prepared_by")
        for f in carry:
            try:
                if source.get(f):
                    d.set(f, source.get(f))
            except Exception:
                pass
        d.dc_date = frappe.utils.nowdate()
        d.remarks = ("Short loading. These blocks were on {0} and did not go. "
                     "Returned by {1}: {2}".format(source.name, person, reason))
        want = {_s(x) for x in block_numbers}
        for r in (source.get("dc_block_rows") or []):
            keys = {_s(r.get("export_block_no")), _s(r.get("block_no"))}
            if not (keys & want):
                continue
            row = d.append("dc_block_rows", {})
            for f in ("block", "grade", "source_inspection", "granite_size_category",
                      "block_no", "export_block_no", "block_number_input",
                      "length_gross", "width_gross", "height_gross",
                      "gross_volume", "gross_tonnage", "tonnage_factor"):
                try:
                    row.set(f, r.get(f))
                except Exception:
                    pass
        if not d.get("dc_block_rows"):
            return None
        d.flags.ignore_permissions = True
        d.insert(ignore_permissions=True)
        return d.name
    except Exception:
        # A draft that could not be made must never swallow the return itself.
        # The blocks are back in stock either way; say so rather than pretend.
        return None


@frappe.whitelist()
def undo_return_from_dc(dc=None, blocks=None, reason=None, person=None,
                        remove_draft=None):
    """Put back what return_from_dc took off. His rule: everything reverses.

    Restores the exact status each block held before it was returned. The draft
    challan made at the time is removed only if it is still a draft and only if
    the caller names it - nobody's later work is thrown away silently.
    """
    blocks = frappe.parse_json(blocks) if isinstance(blocks, str) else (blocks or [])
    reason = _s(reason)
    person = _s(person) or _s(frappe.session.user)
    if not blocks:
        frappe.throw("No blocks given to put back.")
    if len(reason) < 4:
        frappe.throw("Undoing a return needs a written reason too.")

    from dolphin_theme.block_resolve import try_resolve, set_status
    # The blocks we are undoing are exactly the ones carrying a
    # status_before_dc_return, and that index is keyed by BOTH of a block's
    # numbers - so the record is found without resolving an ambiguous number.
    idx = _returned_blocks_index()
    restored, refused = [], []
    for bn in [_s(b) for b in blocks if _s(b)]:
        rec = idx.get(bn)
        if rec:
            hit = {"name": _s(rec.get("name"))}
        else:
            hit, why = try_resolve(bn, allow_record_name=False)
        if not hit:
            refused.append({"block": bn, "why": why})
            continue
        prev = _s(frappe.db.get_value("Quarry Block", hit["name"],
                                      "status_before_dc_return"))
        if not prev:
            refused.append({"block": bn, "why": "this block was not returned from a challan"})
            continue
        res = set_status(
            hit["name"], prev,
            "return undone{0} - put back to {1} by {2}: {3}".format(
                (" (" + _s(dc) + ")") if _s(dc) else "", prev, person, reason),
            machine="server (all transported)", actor=person, allow_backwards=True)
        if res.get("ok"):
            try:
                frappe.db.set_value("Quarry Block", hit["name"],
                                    "status_before_dc_return", "")
            except Exception:
                pass
            restored.append(bn)
        else:
            refused.append({"block": bn, "why": res.get("error") or "refused"})

    removed_draft = None
    rd = _s(remove_draft)
    if rd:
        try:
            if cint(frappe.db.get_value("Delivery Challan", rd, "docstatus")) == 0:
                frappe.delete_doc("Delivery Challan", rd, ignore_permissions=True,
                                  delete_permanently=False)
                removed_draft = rd
        except Exception:
            removed_draft = None

    frappe.db.commit()
    return {"ok": 1, "restored": len(restored), "restored_blocks": restored,
            "refused": refused, "removed_draft": removed_draft}
