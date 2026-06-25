# dolphin_theme/api.py
# Day 37 build — elevated, no-permission-block server methods for Dolphin International.
#
# Deploy path in repo:  dolphin_theme/api.py   (callable as dolphin_theme.api.<fn>)
#
# Provides:
#   1. bulk_import_quarry_blocks(rows, dry_run, mode)  -> elevated Quarry Block import
#                                                         (works for ANY logged-in user,
#                                                          no doctype permission needed)
#   2. log_edit(ref_doctype, ref_name, action, device_id, device_name)
#                                                       -> append-only device/IP audit
#   3. get_machines()        -> dropdown list for the "name this machine" UI
#   4. get_known_locations() -> ip->location map (for reference / admin)
#
# Design notes:
#   * Client scripts do NOT run on bulk import, so gross_volume / gross_tonnage are
#     recomputed here, server-side, from the raw L/W/H + the DMG factor.
#   * tonnage_factor is resolved from dmg_tonnage_factor -> "DMG Tonnage Factor Master".factor_value
#   * All writes use ignore_permissions=True on purpose: the whole point is that
#     operators on shared branch logins can import / be audited WITHOUT being granted
#     write permission on the underlying doctypes. The method itself is the gate.

import base64
import io
import json

import frappe
from frappe.utils import flt, cint, getdate, now_datetime


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _r2(x):
	return round(flt(x) + 0.0, 2)


def _factor_for(dmg_name):
	"""Return the numeric specific-gravity factor for a DMG Tonnage Factor Master row."""
	if not dmg_name:
		return 0.0
	val = frappe.db.get_value("DMG Tonnage Factor Master", dmg_name, "factor_value")
	if val:
		return flt(val)
	# the master name is often itself the number (e.g. "2.70")
	try:
		return flt(dmg_name)
	except Exception:
		return 0.0


def _link_ok(doctype, value):
	if value in (None, ""):
		return True
	return bool(frappe.db.exists(doctype, value))


# field map for the Quarry Block template (header label -> Quarry Block fieldname).
# The UI maps headers to these fieldnames before sending, so here we just accept fieldnames.
_QB_FIELDS = [
	"block_number", "export_block_no", "block_suffix", "date_produced",
	"pit", "gangman", "length_gross", "width_gross", "height_gross",
	"status", "granite_quality_grade", "granite_size_category",
	"dmg_tonnage_factor",
]

_QB_LINKS = {
	"pit": "Pit",
	"gangman": "Gangman",
	"granite_quality_grade": "Granite Grade",
	"granite_size_category": "Granite Size Category",
	"dmg_tonnage_factor": "DMG Tonnage Factor Master",
}


# header label (lower-cased) -> Quarry Block fieldname, for parsing the xlsx template
_QB_HEADER_MAP = {
	"block number": "block_number",
	"quarry block number": "block_number",
	"export block no": "export_block_no",
	"export no": "export_block_no",
	"block suffix": "block_suffix",
	"date produced": "date_produced",
	"pit": "pit",
	"gangman": "gangman",
	"length gross": "length_gross",
	"width gross": "width_gross",
	"height gross": "height_gross",
	"status": "status",
	"granite_quality_grade": "granite_quality_grade",
	"grade": "granite_quality_grade",
	"granite size category": "granite_size_category",
	"size": "granite_size_category",
	"dmg tonnage factor": "dmg_tonnage_factor",
	"specific gravity": "dmg_tonnage_factor",
}


def _map_header(label):
	if label is None:
		return None
	t = str(label).strip().lower()
	if t in _QB_HEADER_MAP:
		return _QB_HEADER_MAP[t]
	for key, fn in _QB_HEADER_MAP.items():
		if t.startswith(key):
			return fn
	return None


def _parse_quarry_xlsx(file_content_b64):
	"""Decode a base64 xlsx and return a list of row dicts keyed by Quarry Block fieldname."""
	from openpyxl import load_workbook

	data = base64.b64decode(file_content_b64.split(",")[-1])
	wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
	ws = wb["Quarry Block"] if "Quarry Block" in wb.sheetnames else wb[wb.sheetnames[0]]

	rows, headers = [], None
	for r in ws.iter_rows(values_only=True):
		if headers is None:
			headers = [_map_header(c) for c in r]
			continue
		row = {}
		for fn, val in zip(headers, r):
			if fn and val not in (None, ""):
				row[fn] = val
		if row:
			rows.append(row)
	return rows


# --------------------------------------------------------------------------- #
# 1. Quarry Block import (elevated, idempotent, dry-run capable)
# --------------------------------------------------------------------------- #
@frappe.whitelist()
def bulk_import_quarry_blocks(rows=None, file_content=None, dry_run=1, mode="upsert"):
	"""Import Quarry Blocks from an uploaded xlsx (preferred) or a list of row dicts.

	file_content : base64 of the filled Quarry Block template (xlsx). Parsed server-side.
	rows         : alternatively, JSON string/list of {fieldname: value, ...}.
	dry_run      : 1 = validate + compute only, write nothing. 0 = actually write.
	mode         : "upsert" (update existing block_number, else create) or "create".
	Returns      : summary {dry_run, mode, total, created, updated, skipped, errors[]}.
	"""
	dry_run = cint(dry_run)
	if file_content:
		rows = _parse_quarry_xlsx(file_content)
	elif isinstance(rows, str):
		rows = json.loads(rows)
	rows = rows or []

	out = {"dry_run": dry_run, "mode": mode, "total": 0,
	       "created": 0, "updated": 0, "skipped": 0, "errors": []}

	for idx, raw in enumerate(rows, start=2):  # row 2 = first data row in the sheet
		row = {k: (v.strip() if isinstance(v, str) else v) for k, v in (raw or {}).items()}
		bno = row.get("block_number")
		if not bno:
			# completely blank line -> skip silently; partially filled -> error
			if any(v not in (None, "") for v in row.values()):
				out["errors"].append({"row": idx, "block_number": "", "message": "missing Block Number"})
				out["skipped"] += 1
			continue

		out["total"] += 1

		# validate link fields
		bad = []
		for fn, dt in _QB_LINKS.items():
			if not _link_ok(dt, row.get(fn)):
				bad.append("{0} '{1}' not found".format(fn, row.get(fn)))
		if bad:
			out["errors"].append({"row": idx, "block_number": bno, "message": "; ".join(bad)})
			out["skipped"] += 1
			continue

		# compute volume / tonnage
		L, W, H = cint(row.get("length_gross")), cint(row.get("width_gross")), cint(row.get("height_gross"))
		factor = _factor_for(row.get("dmg_tonnage_factor"))
		gross_volume = _r2(L * W * H / 1000000.0) if (L and W and H) else 0.0
		gross_tonnage = _r2(gross_volume * factor) if (gross_volume and factor) else 0.0

		values = {
			"block_number": bno,
			"export_block_no": row.get("export_block_no"),
			"block_suffix": row.get("block_suffix"),
			"pit": row.get("pit"),
			"gangman": row.get("gangman"),
			"length_gross": L, "width_gross": W, "height_gross": H,
			"status": row.get("status") or "In Stock",
			"granite_quality_grade": row.get("granite_quality_grade"),
			"granite_size_category": row.get("granite_size_category"),
			"dmg_tonnage_factor": row.get("dmg_tonnage_factor"),
			"tonnage_factor": factor,
			"gross_volume": gross_volume,
			"gross_tonnage": gross_tonnage,
		}
		if row.get("date_produced"):
			try:
				values["date_produced"] = getdate(row.get("date_produced"))
			except Exception:
				out["errors"].append({"row": idx, "block_number": bno, "message": "bad Date Produced"})

		# decide create vs update
		existing = frappe.get_all("Quarry Block", filters={"block_number": bno}, pluck="name")
		try:
			if mode == "upsert" and existing:
				if len(existing) > 1:
					out["errors"].append({"row": idx, "block_number": bno,
					                       "message": "duplicate block_number ({0} records) — skipped".format(len(existing))})
					out["skipped"] += 1
					continue
				if not dry_run:
					doc = frappe.get_doc("Quarry Block", existing[0])
					doc.update({k: v for k, v in values.items() if v not in (None, "")})
					doc.save(ignore_permissions=True)
				out["updated"] += 1
			else:
				if not dry_run:
					doc = frappe.get_doc(dict(doctype="Quarry Block", **{k: v for k, v in values.items() if v not in (None, "")}))
					doc.insert(ignore_permissions=True)
				out["created"] += 1
		except Exception as e:
			out["errors"].append({"row": idx, "block_number": bno, "message": str(e)[:300]})
			out["skipped"] += 1

	if not dry_run:
		frappe.db.commit()
	return out


# --------------------------------------------------------------------------- #
# 2. Device / IP audit  (append-only)
# --------------------------------------------------------------------------- #
@frappe.whitelist()
def log_edit(ref_doctype, ref_name, action, device_id=None, device_name=None):
	"""Append one audit entry. Never raises — auditing must not break a save."""
	try:
		ip = getattr(frappe.local, "request_ip", None) or ""
		log = frappe.get_doc({
			"doctype": "Dolphin Edit Log",
			"ref_doctype": ref_doctype,
			"ref_name": ref_name,
			"action": action,
			"edit_user": frappe.session.user,
			"device_id": device_id or "",
			"device_name": device_name or "(un-named device)",
			"ip_address": ip,
			"location": _location_for(ip),
			"event_time": now_datetime(),
		})
		log.flags.ignore_permissions = True
		log.insert(ignore_permissions=True)
		frappe.db.commit()
		return {"ok": 1}
	except Exception:
		frappe.log_error(title="Dolphin log_edit failed")
		return {"ok": 0}


def _location_for(ip):
	if not ip:
		return "unknown"
	loc = frappe.db.get_value("Dolphin Known Location", {"ip_address": ip}, "location_name")
	return loc or "unknown (flag)"


# --------------------------------------------------------------------------- #
# 3 + 4. small read helpers for the UI
# --------------------------------------------------------------------------- #
@frappe.whitelist()
def get_machines():
	return frappe.get_all("Dolphin Machine", fields=["name", "label", "branch"], order_by="branch, label")


@frappe.whitelist()
def get_known_locations():
	return frappe.get_all("Dolphin Known Location", fields=["ip_address", "location_name"], order_by="location_name")
