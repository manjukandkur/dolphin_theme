"""
Dolphin International -- live import-template generator.

Builds the Quarry Block / Quarry Inspection / Buyer Inspection import
templates ON DEMAND from the current masters, so the dropdowns
(pits, grades, sizes, gangmen, DMG factors, inspectors, in-stock blocks)
and the PitMap / BlockMap auto-fill are always up to date at the moment
of download.

Three whitelisted endpoints (call them from the list "Template" button):
    /api/method/dolphin_theme.template_generator.quarry_block_template
    /api/method/dolphin_theme.template_generator.quarry_inspection_template
    /api/method/dolphin_theme.template_generator.buyer_inspection_template

Purely additive: one file, no hooks change, nothing existing is touched.

NOTE (Day 41 recovery): the in-stock block-number dropdown on the Buyer
Inspection template was added in commit bbee130 and then lost on 2026-06-19
by an "Add files via upload" folder re-upload from a stale copy. This file
restores it AND adds BlockMap auto-fill (L/W/H + Source QI). Do NOT re-upload
the whole folder over this -- edit single files only, or these features get
clobbered again.
"""

import io

import frappe
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Light red used to flag a missing mandatory cell right in the sheet.
REQ_FILL = PatternFill("solid", fgColor="FFFFC7CE")

# How many rows to pre-arm with formulas / dropdowns.
FORMULA_ROWS = 300
DV_ROWS = 500

NAVY = "FF0F2540"
GREY = "FFF1F1EE"
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
REF_FONT = Font(bold=True, color="FF555555", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
REF_FILL = PatternFill("solid", fgColor=GREY)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# --------------------------------------------------------------------------
# master readers (all defensive -- a missing master never crashes the build)
# --------------------------------------------------------------------------
def _names(doctype, order_by="name asc"):
	try:
		return [d.name for d in frappe.get_all(doctype, fields=["name"], order_by=order_by)]
	except Exception:
		return []


def _select_options(doctype, fieldname, fallback=None):
	try:
		opts = (frappe.get_meta(doctype).get_field(fieldname).options or "").split("\n")
		opts = [o.strip() for o in opts if o.strip()]
		return opts or (fallback or [])
	except Exception:
		return fallback or []


def _pits():
	"""Returns ordered list of dicts: {pit, gangman, factor}."""
	try:
		rows = frappe.get_all(
			"Pit",
			fields=["name", "default_gangman", "specific_gravity"],
			limit_page_length=0,
		)
	except Exception:
		rows = []

	def sort_key(r):
		try:
			return (0, int(str(r.name)))
		except Exception:
			return (1, str(r.name))

	rows.sort(key=sort_key)
	out = []
	for r in rows:
		out.append(
			{
				"pit": str(r.name),
				"gangman": r.get("default_gangman") or "",
				"factor": r.get("specific_gravity") or 2.6,
			}
		)
	return out


def _instock_blocks():
	"""Current in-stock Quarry Blocks with dims + source QI, sorted by block number.

	Returns list of dicts: {block_number, name, l, w, h, qi}.
	"""
	try:
		rows = frappe.get_all(
			"Quarry Block",
			filters={"status": "In Stock"},
			fields=[
				"name",
				"block_number",
				"length_gross",
				"width_gross",
				"height_gross",
				"source_quarry_inspection",
				"pit",
			],
			limit_page_length=0,
		)
	except Exception:
		rows = []

	# pit -> specific gravity, so each block's tonnage uses its own pit's SG
	pit_sg = {}
	try:
		for p in frappe.get_all("Pit", fields=["name", "specific_gravity"], limit_page_length=0):
			pit_sg[str(p.name)] = p.get("specific_gravity") or 2.6
	except Exception:
		pit_sg = {}

	def sort_key(r):
		try:
			return (0, int(str(r.block_number)))
		except Exception:
			return (1, str(r.block_number))

	rows = [r for r in rows if r.get("block_number")]
	rows.sort(key=sort_key)
	out = []
	for r in rows:
		out.append(
			{
				"block_number": str(r.block_number),
				"name": str(r.name),
				"l": r.get("length_gross") or "",
				"w": r.get("width_gross") or "",
				"h": r.get("height_gross") or "",
				"qi": r.get("source_quarry_inspection") or "",
				"sg": pit_sg.get(str(r.get("pit")), 2.6),
			}
		)
	return out


# --------------------------------------------------------------------------
# sheet helpers
# --------------------------------------------------------------------------
def _write_headers(ws, headers, ref_cols=()):
	for c, label in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=c, value=label)
		if c in ref_cols:
			cell.fill, cell.font = REF_FILL, REF_FONT
		else:
			cell.fill, cell.font = HEADER_FILL, HEADER_FONT
		cell.alignment = CENTER
		ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(label)) + 2)
	ws.row_dimensions[1].height = 30
	ws.freeze_panes = "A2"


def _list_dv(ws_lists, col_idx, header, values):
	"""Drop the values onto a hidden Lists sheet, return a range DV string.

	Using a sheet range (not an inline list) avoids the 255-char limit and any
	value that happens to contain a comma.
	"""
	letter = get_column_letter(col_idx)
	ws_lists.cell(row=1, column=col_idx, value=header)
	for i, v in enumerate(values, start=2):
		ws_lists.cell(row=i, column=col_idx, value=v)
	last = max(2, len(values) + 1)
	ref = "Lists!${0}$2:${0}${1}".format(letter, last)
	return DataValidation(type="list", formula1=ref, allow_blank=True, showErrorMessage=True)


def _whole_dv():
	return DataValidation(
		type="whole", operator="between", formula1="1", formula2="999", allow_blank=True
	)


def _date_col(ws, letter):
	for r in range(2, DV_ROWS + 1):
		ws[letter + str(r)].number_format = "DD-MM-YYYY"


def _apply(ws, dv, col_letter):
	ws.add_data_validation(dv)
	dv.add("{0}2:{0}{1}".format(col_letter, DV_ROWS))


def _pitmap(wb, pits):
	ws = wb.create_sheet("PitMap")
	ws.append(["Pit", "Gangman", "Factor"])
	for p in pits:
		ws.append([p["pit"], p["gangman"], p["factor"]])
	ws.sheet_state = "hidden"
	return ws


def _blockmap(wb, blocks):
	"""Hidden sheet: block_number -> name, L, W, H, SG(from pit) for VLOOKUP autofill."""
	ws = wb.create_sheet("BlockMap")
	ws.append(["Block Number", "Block Name", "L", "W", "H", "SG"])
	for b in blocks:
		ws.append([b["block_number"], b["name"], b["l"], b["w"], b["h"], b["sg"]])
	ws.sheet_state = "hidden"
	return ws


def _readme(wb, title, lines):
	ws = wb.create_sheet("READ ME")
	ws["A1"] = title
	ws["A1"].font = Font(bold=True, size=14)
	for i, ln in enumerate(lines, start=3):
		ws.cell(row=i, column=1, value=ln)
	ws.column_dimensions["A"].width = 110


def _send(wb, filename):
	buf = io.BytesIO()
	wb.save(buf)
	frappe.response["filename"] = filename
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"
	frappe.response[
		"content_type"
	] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --------------------------------------------------------------------------
# 1. QUARRY BLOCK
# --------------------------------------------------------------------------
@frappe.whitelist()
def quarry_block_template():
	pits = _pits()
	statuses = _select_options(
		"Quarry Block",
		"status",
		fallback=[
			"In Stock",
			"Buyer Marked",
			"In Delivery Challan",
			"Dispatched/Transported",
			"At Port",
			"At Bannikoppa Station yard",
			"Shipped",
		],
	)
	grades = _names("Granite Grade")
	sizes = _names("Granite Size Category")

	wb = Workbook()
	ws = wb.active
	ws.title = "Quarry Block"
	lists = wb.create_sheet("Lists")
	lists.sheet_state = "hidden"

	headers = [
		"Quarry Block Number",
		"Date Produced (DD-MM-YYYY)",
		"Pit",
		"Gangman",
		"Length Gross",
		"Width Gross",
		"Height Gross",
		"Status",
		"granite_quality_grade",
		"Granite Size Category",
		"dmg_tonnage_factor",
		"Gross Volume",
		"Gross Tonnage",
	]
	_write_headers(ws, headers, ref_cols={4, 11, 12, 13})
	_date_col(ws, "B")

	_apply(ws, _list_dv(lists, 1, "Pit", [p["pit"] for p in pits]), "C")
	_apply(ws, _list_dv(lists, 2, "Status", statuses), "H")
	_apply(ws, _list_dv(lists, 3, "Grade", grades), "I")
	_apply(ws, _list_dv(lists, 4, "Size", sizes), "J")
	for col in ("E", "F", "G"):
		_apply(ws, _whole_dv(), col)

	for r in range(2, FORMULA_ROWS + 2):
		ws["D{0}".format(r)] = '=IFERROR(VLOOKUP($C{0}&"",PitMap!$A:$C,2,FALSE),"")'.format(r)
		ws["K{0}".format(r)] = '=IFERROR(TEXT(VLOOKUP($C{0}&"",PitMap!$A:$C,3,FALSE),"0.00"),"")'.format(r)
		ws["L{0}".format(r)] = '=IF(AND(E{0}>0,F{0}>0,G{0}>0),ROUND(E{0}*F{0}*G{0}/1000000,4),"")'.format(r)
		ws["M{0}".format(r)] = '=IF(AND(ISNUMBER(L{0}),C{0}<>""),ROUND(L{0}*IFERROR(VLOOKUP($C{0}&"",PitMap!$A:$C,3,FALSE),0),2),"")'.format(r)

	_pitmap(wb, pits)
	_send(wb, "Dolphin_QuarryBlock_Template.xlsx")


# --------------------------------------------------------------------------
# 2. QUARRY INSPECTION  (parent + Block Rows child)
# --------------------------------------------------------------------------
@frappe.whitelist()
def quarry_inspection_template():
	pits = _pits()
	grades = _names("Granite Grade")
	inspectors = _names("Quarry Inspector")
	supervisors = _names("Supervisor")

	wb = Workbook()
	ws = wb.active
	ws.title = "Quarry Inspection"
	lists = wb.create_sheet("Lists")
	lists.sheet_state = "hidden"

	headers = [
		"Report No",
		"Report Date",
		"Inspector (Quarry Inspector)",
		"Supervisor (Supervisor)",
		"Quarry Block No (Block Rows)",
		"Pit (Block Rows)",
		"Gangman (Block Rows)",
		"DMG Tonnage Factor (Block Rows)",
		"L (Block Rows)",
		"W (Block Rows)",
		"H (Block Rows)",
		"Grade (Block Rows)",
		"Remarks",
		"Gross Vol (Block Rows)",
		"Gross Tonnage (Block Rows)",
	]
	_write_headers(ws, headers, ref_cols={7, 8, 14, 15})
	_date_col(ws, "B")
	# Pre-fill the first row's Report Date so the importer always has a parent anchor.
	# A blank Report Date on the first row makes the Inspector/Supervisor multiselect
	# import crash ('NoneType' has no attribute 'get'); this prevents that.
	try:
		ws["B2"] = frappe.utils.getdate(frappe.utils.nowdate())
	except Exception:
		pass

	# --- Mandatory-field guidance shown right inside the sheet ---------------
	# Header notes (hover) marking which fields are required and which are one-time.
	ws["B1"].comment = Comment("REQUIRED. Report Date - fill ONCE on the first row only.", "Dolphin")
	ws["C1"].comment = Comment("REQUIRED. Inspector - fill ONCE on the first row only.", "Dolphin")
	ws["D1"].comment = Comment("Optional. Supervisor - fill once on the first row only.", "Dolphin")
	ws["E1"].comment = Comment("REQUIRED for every block row.", "Dolphin")
	for col in ("I", "J", "K"):
		ws[col + "1"].comment = Comment("REQUIRED for every block row (L / W / H).", "Dolphin")
	# Red highlight when a required cell is missing.
	last = FORMULA_ROWS + 1
	ws.conditional_formatting.add("C2", FormulaRule(formula=['$C$2=""'], fill=REQ_FILL))
	ws.conditional_formatting.add(
		"E2:E{0}".format(last),
		FormulaRule(formula=['AND(OR(I2<>"",J2<>"",K2<>""),E2="")'], fill=REQ_FILL),
	)
	ws.conditional_formatting.add(
		"I2:K{0}".format(last),
		FormulaRule(formula=['AND($E2<>"",I2="")'], fill=REQ_FILL),
	)

	_apply(ws, _list_dv(lists, 1, "Inspector", inspectors), "C")
	_apply(ws, _list_dv(lists, 2, "Supervisor", supervisors), "D")
	_apply(ws, _list_dv(lists, 3, "Pit", [p["pit"] for p in pits]), "F")
	_apply(ws, _list_dv(lists, 4, "Grade", grades), "L")
	for col in ("I", "J", "K"):
		_apply(ws, _whole_dv(), col)

	for r in range(2, FORMULA_ROWS + 2):
		ws["G{0}".format(r)] = '=IFERROR(VLOOKUP($F{0}&"",PitMap!$A:$C,2,FALSE),"")'.format(r)
		ws["H{0}".format(r)] = '=IFERROR(VLOOKUP($F{0}&"",PitMap!$A:$C,3,FALSE),"")'.format(r)
		ws["N{0}".format(r)] = '=IF(AND(I{0}>0,J{0}>0,K{0}>0),ROUND(I{0}*J{0}*K{0}/1000000,4),"")'.format(r)
		ws["O{0}".format(r)] = '=IF(AND(N{0}<>"",H{0}<>""),ROUND(N{0}*H{0},2),"")'.format(r)

	_pitmap(wb, pits)
	_readme(
		wb,
		"Quarry Inspection - Import Template",
		[
			"MANDATORY - fill ONCE on the FIRST data row only: Report Date, Inspector.",
			"MANDATORY for EVERY block row: Quarry Block No, L, W, H.",
			"Missing required cells turn RED automatically; headers have a 'Required' note (hover over them).",
			"Report No is OPTIONAL - leave blank and the server auto-numbers the report.",
			"Report Date is pre-filled with today - change it if needed (first row only).",
			"Supervisor is optional; fill once on the first row.",
			"Every column ending in '(Block Rows)' repeats per block; leave the parent columns blank on rows 2..n to group blocks into ONE report.",
			"Gangman, DMG factor, Gross Vol and Gross Tonnage fill in automatically.",
		],
	)
	_send(wb, "Dolphin_QuarryInspection_Template.xlsx")


# --------------------------------------------------------------------------
# 3. BUYER INSPECTION  (parent + Block Rows child)
# --------------------------------------------------------------------------
@frappe.whitelist()
def buyer_inspection_template():
	sale_types = _select_options("Buyer Inspection", "sale_type", fallback=["Export", "Local"])
	dmg = _names("DMG Tonnage Factor Master")
	buyer_inspectors = _names("Buyer Inspector")
	sizes = _names("Granite Size Category")
	# Consignee: show COMPANY NAMES; a hidden column resolves to the code via ConsigneeMap.
	try:
		_crows = frappe.get_all("Export Consignee", fields=["name", "company_name"], limit_page_length=0)
	except Exception:
		_crows = []
	consignee_pairs = [((c.get("company_name") or c.get("name")), c.get("name")) for c in _crows]
	consignee_pairs = [p for p in consignee_pairs if p[0]]
	consignee_names = [p[0] for p in consignee_pairs]
	blocks = _instock_blocks()

	wb = Workbook()
	ws = wb.active
	ws.title = "Buyer Inspection"
	lists = wb.create_sheet("Lists")
	lists.sheet_state = "hidden"

	headers = [
		"Report No",
		"Report Date",
		"Sale Type",
		"Consignee",
		"Export Consignee (Buyer)",
		"dmg_tonnage_factor",
		"Buyer Inspector (Buyer Inspector)",
		"Block Number (Block Rows)",
		"Export No (Block Rows)",
		"L (Block Rows)",
		"W (Block Rows)",
		"H (Block Rows)",
		"Specific Gravity (Block Rows)",
		"Size (Block Rows)",
	]
	_write_headers(ws, headers, ref_cols={5, 10, 11, 12, 13})
	_date_col(ws, "B")
	ws.column_dimensions["E"].hidden = True

	_apply(ws, _list_dv(lists, 1, "SaleType", sale_types), "C")
	_apply(ws, _list_dv(lists, 2, "Consignee", consignee_names), "D")
	_apply(ws, _list_dv(lists, 3, "DMG", dmg), "F")
	_apply(ws, _list_dv(lists, 4, "BuyerInspector", buyer_inspectors), "G")
	_apply(ws, _list_dv(lists, 5, "InStockBlocks", [b["block_number"] for b in blocks]), "H")
	_apply(ws, _list_dv(lists, 6, "Size", sizes), "N")
	for col in ("J", "K", "L"):
		_apply(ws, _whole_dv(), col)

	for r in range(2, FORMULA_ROWS + 2):
		ws["E{0}".format(r)] = '=IFERROR(VLOOKUP($D{0}&"",ConsigneeMap!$A:$B,2,FALSE),"")'.format(r)
		ws["J{0}".format(r)] = '=IFERROR(VLOOKUP($H{0}&"",BlockMap!$A:$F,3,FALSE),"")'.format(r)
		ws["K{0}".format(r)] = '=IFERROR(VLOOKUP($H{0}&"",BlockMap!$A:$F,4,FALSE),"")'.format(r)
		ws["L{0}".format(r)] = '=IFERROR(VLOOKUP($H{0}&"",BlockMap!$A:$F,5,FALSE),"")'.format(r)
		ws["M{0}".format(r)] = '=IFERROR(VLOOKUP($H{0}&"",BlockMap!$A:$F,6,FALSE),"")'.format(r)

	cmap = wb.create_sheet("ConsigneeMap")
	cmap.append(["Company", "Code"])
	for company, code in consignee_pairs:
		cmap.append([company, code])
	cmap.sheet_state = "hidden"

	_blockmap(wb, blocks)
	_readme(
		wb,
		"Buyer Inspection - Import Template",
		[
			"Report No is OPTIONAL - leave blank and the server auto-numbers the report.",
			"Fill Report Date, Sale Type, Consignee, Specific Gravity and Buyer Inspector ONLY on the FIRST row.",
			"Consignee is a dropdown of consignee COMPANY NAMES - pick one (a hidden column fills the code).",
			"Block Number is a dropdown of CURRENT in-stock block numbers - pick or type to search.",
			"L/W/H and Specific Gravity pre-fill from stock; L/W/H are editable to override.",
			"Size: leave blank to use the block stock size, or pick A/B/C to override.",
			"Leave parent columns blank on rows 2..n to group all blocks into ONE report.",
		],
	)
	_send(wb, "Dolphin_BuyerInspection_Template.xlsx")


@frappe.whitelist()
def export_shipment_lot_template():
	"""Live 'Import XLS' template for Export Shipment Lot.

	A single 'Export Block No' column with a dropdown of the EXPORT numbers
	that are currently At Port and not yet in any lot (refreshed from masters
	on every download). Import matches EXPORT numbers ONLY -- quarry block
	numbers are not accepted."""
	try:
		from dolphin_theme.api_arrivals import at_port_available
		avail = at_port_available(None)
	except Exception:
		avail = []
	rows = []
	for b in avail:
		ex = str(b.get("export_block_no") or "").strip()
		if ex:
			dims = "x".join([str(b.get(k) or "") for k in ("length", "width", "height")])
			rows.append((ex, dims, b.get("cbm")))
	rows.sort(key=lambda r: r[0])
	export_nos = [r[0] for r in rows]

	wb = Workbook()
	ws = wb.active
	ws.title = "Import Blocks"
	_write_headers(ws, ["Export Block No", "Notes (optional - ignored)"])
	ws.column_dimensions["A"].width = 22
	ws.column_dimensions["B"].width = 46

	ws_lists = wb.create_sheet("Lists")
	ws_lists.sheet_state = "hidden"
	if export_nos:
		dv = _list_dv(ws_lists, 1, "Export Block No", export_nos)
		_apply(ws, dv, "A")

	ref = wb.create_sheet("AtPort (ref)")
	ref.append(["Export Block No", "Size (LxWxH)", "CBM"])
	for r in rows:
		ref.append([r[0], r[1], r[2]])
	for c in range(1, 4):
		ref.cell(row=1, column=c).fill = HEADER_FILL
		ref.cell(row=1, column=c).font = HEADER_FONT
	ref.column_dimensions["A"].width = 20
	ref.column_dimensions["B"].width = 20
	ref.column_dimensions["C"].width = 10

	_readme(wb, "Export Shipment Lot - Import XLS template", [
		"1. Put EXPORT block numbers in the 'Export Block No' column (one per row).",
		"   Only EXPORT numbers are matched - quarry block numbers are NOT accepted.",
		"   Pick from the dropdown (current At-Port blocks) or type the number.",
		"2. The 'Notes' column is only for your reference and is ignored on import.",
		"3. On the Export Shipment Lot form click 'Import XLS' and upload this file.",
		"",
		"The system matches each export number to a block that is At Port and not",
		"already in a lot, then fills dimensions / CBM / net tonnage / source DC",
		"automatically. Numbers that are not At Port (or unknown) are listed back.",
		"",
		"Generated LIVE - the dropdown reflects the " + str(len(export_nos)) +
		" export block(s) currently available At Port.",
	])
	_send(wb, "Export_Shipment_Lot_Import_Template.xlsx")
